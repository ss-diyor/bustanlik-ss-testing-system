import asyncio
import os
from aiogram import Router, F, Bot
from aiogram.types import (
    Message,
    CallbackQuery,
    FSInputFile,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    WebAppInfo,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.exceptions import TelegramBadRequest

from config import ADMIN_PASSWORD, MAX_SAVOL, ADMIN_IDS


async def safe_edit(callback: CallbackQuery, text: str, **kwargs):
    """edit_text ni xavfsiz chaqiradi.

    Telegram 'message is not modified' xatosini jim o'tkazib yuboradi —
    foydalanuvchi bir tugmani ikki marta bosganida sodir bo'ladi.
    """
    try:
        await callback.message.edit_text(text, **kwargs)
    except TelegramBadRequest as e:
        if "message is not modified" in str(e):
            pass
        else:
            raise
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use("Agg")  # GUI yo'q muhitda ishlashi uchun
from certificate import CertificateGenerator


def create_class_chart(stats_data):
    """Sinflar bo'yicha o'rtacha ballar grafigini yasash"""
    if not stats_data:
        return None

    # Ma'lumotlarni tayyorlash
    classes = [item["sinf"] for item in stats_data]
    avg_scores = [float(item["avg_score"]) for item in stats_data]

    # Grafik yaratish
    plt.figure(figsize=(10, 6))
    bars = plt.bar(classes, avg_scores, color="#90EE90", alpha=0.8, width=0.6)

    # Grafikni sozlash
    plt.title(
        "Sinflar bo'yicha o'rtacha ballar",
        fontsize=16,
        fontweight="bold",
        pad=20,
    )
    plt.xlabel("Sinflar", fontsize=12)
    plt.ylabel("O'rtacha ball", fontsize=12)
    plt.grid(axis="y", alpha=0.3)

    # Y o'qini sozlash
    max_score = max(avg_scores) if avg_scores else 100
    plt.ylim(0, max_score * 1.1)

    # Har bir bar ustida qiymatni ko'rsatish
    for bar, score in zip(bars, avg_scores):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 1,
            f"{score:.1f}",
            ha="center",
            va="bottom",
            fontweight="bold",
        )

    # Grafikni saqlash
    plt.tight_layout()
    filename = "class_stats.png"
    plt.savefig(filename, dpi=100, bbox_inches="tight")
    plt.close()

    return filename


def create_direction_chart(stats_data):
    """Yo'nalishlar bo'yicha o'rtacha ballar grafigini yasash"""
    if not stats_data:
        return None

    # Ma'lumotlarni tayyorlash
    directions = [item["yonalish"] or "Umumiy" for item in stats_data]
    avg_scores = [float(item["avg_score"]) for item in stats_data]

    # Grafik yaratish
    plt.figure(figsize=(12, 6))
    bars = plt.bar(
        directions, avg_scores, color="#87CEEB", alpha=0.8, width=0.6
    )

    # Grafikni sozlash
    plt.title(
        "Yo'nalishlar bo'yicha o'rtacha ballar",
        fontsize=16,
        fontweight="bold",
        pad=20,
    )
    plt.xlabel("Yo'nalishlar", fontsize=12)
    plt.ylabel("O'rtacha ball", fontsize=12)
    plt.grid(axis="y", alpha=0.3)

    # Y o'qini sozlash
    max_score = max(avg_scores) if avg_scores else 100
    plt.ylim(0, max_score * 1.1)

    # Har bir bar ustida qiymatni ko'rsatish
    for bar, score in zip(bars, avg_scores):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 1,
            f"{score:.1f}",
            ha="center",
            va="bottom",
            fontweight="bold",
        )

    # X o'qidagi yozuvlarni burish
    plt.xticks(rotation=45, ha="right")

    # Grafikni saqlash
    plt.tight_layout()
    filename = "direction_stats.png"
    plt.savefig(filename, dpi=100, bbox_inches="tight")
    plt.close()

    return filename


from database import (
    init_db,
    talaba_topish,
    talaba_qosh,
    talaba_ochir,
    talaba_user_id_ol,
    talaba_user_id_yangila,
    talaba_hammasi,
    talaba_hammasi_by_maktab,
    statistika,
    statistika_by_maktab,
    yonalish_ol,
    yonalish_qosh,
    yonalish_ochir,
    sinf_ol,
    sinf_ol_batafsil,
    sinf_qosh,
    sinf_ochir,
    sinf_ochir_id,
    sinf_talabalar_soni,
    sinf_id_dan_full_nomi,
    kalit_ol,
    kalit_qosh,
    kalit_ochir,
    kalit_holat_ozgartir,
    kalit_tahrirla,
    ball_hisobla,
    natija_qosh,
    talaba_filtrlangan,
    talaba_filtrlangan_by_maktab,
    get_all_students_for_excel,
    get_all_user_ids,
    get_user_ids_by_maktab,
    get_pending_requests,
    update_request_status,
    get_all_in_class,
    get_overall_ranking,
    appeal_ol_id,
    delete_all_data,
    sinf_maktabga_bogla,
    maktab_sinflari_ol,
    get_student_rank,
    get_score_difference,
    get_class_comparison,
    get_class_comparison_by_maktab,
    get_avg_score_by_direction,
    get_avg_score_by_direction_by_maktab,
    get_most_improved_students,
    get_most_improved_students_by_maktab,
    get_most_declined_students,
    get_most_declined_students_by_maktab,
    get_sorted_students_for_excel,
    get_sorted_students_for_excel_by_maktab,
    get_setting,
    set_setting,
    guruhlar_ol,
    guruh_qosh,
    appeal_qosh,
    appeal_javob_ber,
    appeals_ol,
    oqituvchi_ol,
    oqituvchi_qosh,
    oqituvchi_ochir,
    oqituvchilar_hammasi,
    talaba_songi_natija,
    maktab_qosh,
    maktablar_ol,
    maktab_ochir,
    reminder_qosh,
    kutilayotgan_reminders_ol,
    reminder_ochir,
    admin_session_start,
    admin_session_end,
    get_active_admin_sessions,
    is_admin_already_active,
    create_admins_table,
    add_admin,
    remove_admin,
    get_all_admins,
    is_admin_in_db,
    get_all_registered_users,
    # Yangi funktsiyalar
    talaba_tahrirlash,
    maktab_statistikasi,
    bitta_natija_ochir,
    sinf_transferi,
    sinf_transferi_by_id,
    sinf_nomi_ol_by_id,
    bitiruvchilarni_arxivlash,
    bitiruvchilarni_arxivdan_chiqarish,
    dublikatlarni_topish,
    dublikatlarni_birlashtir,
    talaba_natijalari,
    maktablar_ol,
    get_all_students,
    chatbot_oxirgi_foydalanuvchilar,
    chatbot_bugungi_son,
    is_notification_enabled,
    get_sinf_detail_stats,
    get_two_sinf_comparison,
)
from id_generator import (
    keyod_yarat, keyod_settings_ol, keyod_settings_saqla,
    keyod_preview_list, keyod_ommaviy_yarat, kodlar_txt_fayl,
)
from keyboards import (
    admin_menu_keyboard,
    admin_inline_menu,
    yonalish_keyboard,
    tasdiqlash_keyboard,
    baza_tozalash_keyboard,
    yonalish_boshqarish_keyboard,
    yonalish_ochirish_keyboard,
    sinf_keyboard,
    sinf_boshqarish_keyboard,
    sinf_ochirish_keyboard,
    kalit_boshqarish_keyboard,
    kalit_actions_keyboard,
    kalit_yonalish_tanlash_keyboard,
    oquvchilar_filtrlash_keyboard,
    sinf_tanlash_keyboard,
    yonalish_tanlash_keyboard,
    filter_actions_keyboard,
    settings_keyboard,
    request_actions_keyboard,
    ranking_keyboard,
    sinf_tanlash_ranking_keyboard,
    sinf_prefix_tanlash_keyboard,
    oqituvchi_boshqarish_keyboard,
    oqituvchi_menu_keyboard,
    oqituvchi_ochirish_keyboard,
    reminder_boshqarish_keyboard,
    maktab_boshqarish_keyboard,
    guruh_boshqarish_keyboard,
    reminder_list_keyboard,
    maktab_list_keyboard,
    maktab_detail_keyboard,
    appeals_keyboard,
    appeal_action_keyboard,
    stats_keyboard,
    admin_management_keyboard,
    # Yangi klaviaturalar
    talaba_tahrirlash_keyboard,
    talaba_edit_options_keyboard,
    maktab_statistikasi_keyboard,
    maktab_solishtirish_keyboard,
    maktab_comp2_keyboard,
    natijalar_ochirish_keyboard,
    sinf_transferi_keyboard,
    sinf_tanlash_transfer_keyboard,
    yangi_sinf_tanlash_keyboard,
    bitiruvchilar_arxivlash_keyboard,
    sinf_arxivlash_keyboard,
    sinf_arxivdan_chiqarish_keyboard,
    dublikatlar_keyboard,
    dublikat_birlashtirish_keyboard,
    # PDF export klaviaturalar
    pdf_export_keyboard,
    sinf_tanlash_pdf_keyboard,
    sinf_tanlash_pdf_maktab_keyboard,
    # Excel export klaviaturalar
    excel_export_keyboard,
    sinf_tanlash_excel_keyboard,
    sinf_tanlash_excel_maktab_keyboard,
    # Google Sheets export klaviaturalar
    sheets_export_keyboard,
    maktab_tanlash_keyboard,
    broadcast_cancel_keyboard,
    broadcast_confirm_keyboard,
    broadcast_target_keyboard,
    broadcast_maktab_tanlash_keyboard,
    shaxsiy_xabar_confirm_keyboard,
    sinf_taqqoslash_birinchi_keyboard,
    sinf_taqqoslash_ikkinchi_keyboard,
    sinf_taqqoslash_natija_keyboard,
)

router = Router()

# ─────────────────────────────────────────
# FSM holatlari
# ─────────────────────────────────────────


class AdminLogin(StatesGroup):
    parol_kutish = State()


class AdminAdd(StatesGroup):
    user_id_kutish = State()
    tasdiq_kutish = State()


class TalabaQosh(StatesGroup):
    kod_kutish = State()
    yonalish_kutish = State()
    maktab_kutish = State()
    sinf_kutish = State()
    ismlar_kutish = State()
    majburiy_kutish = State()
    asosiy1_kutish = State()
    asosiy2_kutish = State()
    tasdiq_kutish = State()


class IdGenSozlama(StatesGroup):
    prefix_kutish   = State()
    digits_kutish   = State()
    separator_kutish = State()


class ExcelImport(StatesGroup):
    fayl_kutish = State()


class KalitBoshqar(StatesGroup):
    nomi_kutish = State()
    kalit_kutish = State()
    edit_kalit_kutish = State()
    yonalish_nomi_kutish = State()  # Yo'nalishga kalit: test nomi
    yonalish_kalit_kutish = State()  # Yo'nalishga kalit: kalitlar matni


class BittaNatijaOchirish(StatesGroup):
    kod_kutish = State()


class TalabaTahrirlash(StatesGroup):
    tanlash = State()
    ism_kutish = State()
    sinf_kutish = State()
    yonalish_kutish = State()


class SinfTransferi(StatesGroup):
    eski_sinf = State()
    yangi_sinf = State()


class DublikatBirlashtirish(StatesGroup):
    tanlash = State()


class KodQidirish(StatesGroup):
    kod_kutish = State()


class YonalishBoshqar(StatesGroup):
    nomi_kutish = State()


class SinfBoshqar(StatesGroup):
    nomi_kutish = State()
    maktab_kutish = State()


class Broadcast(StatesGroup):
    maktab_tanlash = State()
    xabar_kutish = State()
    tasdiq_kutish = State()


class ShaxsiyXabar(StatesGroup):
    kod_kutish = State()
    xabar_kutish = State()
    tasdiq_kutish = State()


class MurojaatJavob(StatesGroup):
    javob_kutish = State()


class OqituvchiQosh(StatesGroup):
    user_id_kutish = State()
    ismlar_kutish = State()
    sinf_kutish = State()


class NatijaTahrirlash(StatesGroup):
    kod_kutish = State()
    majburiy_kutish = State()
    asosiy1_kutish = State()
    asosiy2_kutish = State()


class ReminderAdd(StatesGroup):
    xabar_kutish = State()
    vaqt_kutish = State()


class MaktabAdd(StatesGroup):
    nomi_kutish = State()


class MaktabSinfAdd(StatesGroup):
    maktab_id = State()
    sinf_nomi = State()


class TalabaOchir(StatesGroup):
    kod_kutish = State()
    tasdiq_kutish = State()


class AppealReply(StatesGroup):
    appeal_id = State()
    javob_kutish = State()

class MiniTestState(StatesGroup):
    pdf_kutish = State()
    keys_kutish = State()
    nomi_kutish = State()
    fan_kutish = State()
    tasdiq_kutish = State()


# ─────────────────────────────────────────
# Yordamchi funksiyalar
# ─────────────────────────────────────────


async def is_admin(message: Message, state: FSMContext):
    data = await state.get_data()
    return data.get("admin") is True


async def admin_tekshir(state: FSMContext, user_id: int = None) -> bool:
    data = await state.get_data()
    if data.get("admin") is True:
        return True
    if user_id:
        try:
            from config import ADMIN_IDS

            if int(user_id) in [int(i) for i in ADMIN_IDS]:
                await state.update_data(admin=True)
                return True
        except Exception:
            pass
    return False


def is_admin_id(user_id: int) -> bool:
    from config import ADMIN_IDS

    return user_id in ADMIN_IDS


def _son_tekshir(text: str):
    try:
        son = int(text)
        if 0 <= son <= MAX_SAVOL:
            return son
    except ValueError:
        pass
    return None


def _normalize_name(value: str) -> str:
    return " ".join(str(value).strip().lower().split())


def _resolve_maktab(value, maktablar: list[dict]):
    """Maktab qiymatini id yoki nom bo'yicha aniqlaydi."""
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw.lower() == "nan":
        return None

    by_name = {_normalize_name(m["nomi"]): m for m in maktablar}

    # Raqam bo'yicha topish
    try:
        maktab_id = int(float(raw))
        for m in maktablar:
            if int(m["id"]) == maktab_id:
                return m
    except ValueError:
        pass

    # Nom bo'yicha topish
    return by_name.get(_normalize_name(raw))


def _generate_certificate_file(cert_gen, ismlar, ball, sana, kod, sinf="", maktab=""):
    """Sertifikatni alohida executor'da generatsiya qilish uchun.
    Returns: (pdf_path, cert_hash)
    """
    return cert_gen.generate(ismlar, ball, sana, kod, sinf=sinf, maktab=maktab)


def _parse_sinf_maktab(sinf_field: str) -> tuple[str, str]:
    """
    talabalar.sinf maydonidan sinf va maktab nomini ajratib oladi.
    Format: "11-A - Bo'stonliq ITMA"  →  ("11-A", "Bo'stonliq ITMA")
    Agar " - " bo'lmasa faqat sinf qaytariladi.
    """
    if " - " in sinf_field:
        parts = sinf_field.split(" - ", 1)
        return parts[0].strip(), parts[1].strip()
    return sinf_field.strip(), ""


def _split_long_text(text: str, max_len: int = 3500):
    """Uzun matnni Telegram limitidan oshmasdan bo'lib beradi."""
    if len(text) <= max_len:
        return [text]

    chunks = []
    current = ""
    for line in text.splitlines(keepends=True):
        if len(current) + len(line) > max_len:
            if current:
                chunks.append(current)
                current = ""
            # Juda uzun bitta qator bo'lsa ham bo'lib yuboramiz
            while len(line) > max_len:
                chunks.append(line[:max_len])
                line = line[max_len:]
        current += line
    if current:
        chunks.append(current)
    return chunks


def _build_students_page(
    talabalar: list, title: str, page: int = 1, page_size: int = 20
):
    total = len(talabalar)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size
    visible = talabalar[start:end]

    text = f"{title} (jami: {total} ta)\n"
    text += f"📄 Sahifa: {page}/{total_pages}\n\n"
    for i, t in enumerate(visible, start=start + 1):
        text += f"{i}. {t['kod']} | {t['ismlar']} | {t['sinf']}\n"
    return text, page, total_pages


async def _ota_onalarga_bildirish(bot: Bot, talaba_kod: str, ball: float):
    """
    Yangi natija kiritilganda ulangan ota-onalarga avtomatik xabar yuboradi.
    admin.py ning ichidan chaqiriladi — parent.py moduliga delegatsiya qiladi.
    """
    try:
        from parent import ota_onalarga_xabar_yuborish
        await ota_onalarga_xabar_yuborish(bot, talaba_kod, ball)
    except Exception as e:
        logging.warning(f"Ota-onalarga bildirish xatosi ({talaba_kod}): {e}")


async def guruhlarga_yangi_oquvchi_yuborish(
    bot: Bot, talaba: dict, natija: dict = None
):
    """Guruhlarga yangi qo'shilgan o'quvchi ma'lumotlarini yuboradi."""
    from database import guruhlar_ol

    guruhlar = guruhlar_ol()
    if not guruhlar:
        return

    text = (
        f"🆕 <b>Yangi o'quvchi qo'shildi!</b>\n\n"
        f"👤 Ism: <b>{talaba['ismlar']}</b>\n"
        f"🆔 Kod: <b>{talaba['kod']}</b>\n"
        f"🏫 Sinf: <b>{talaba['sinf']}</b>\n"
        f"🎯 Yo'nalish: <b>{talaba['yonalish']}</b>"
    )
    if natija:
        text += f"\n📊 Birinchi natija: <b>{natija['umumiy_ball']} ball</b>"

    for g in guruhlar:
        try:
            await bot.send_message(g["chat_id"], text, parse_mode="HTML")
        except Exception:
            pass


async def bildirishnoma_yuborish(
    bot: Bot, talaba_kod: str, natija: dict, talaba: dict
):
    """
    Talabaning Telegram user_id si bo'lsa, yangi natija haqida xabar yuboradi.
    Sertifikat ham generatsiya qilinadi.
    """
    user_id = talaba_user_id_ol(talaba_kod)
    if not user_id:
        return  # Talaba profilini ulamagan — bildirishnoma yo'q

    from database import get_student_rank, get_score_difference

    try:
        ranks = get_student_rank(talaba_kod)
        diff = get_score_difference(talaba_kod)
    except Exception as e:
        print(f"Rank/Diff error: {e}")
        ranks = {"class": None, "overall": None}
        diff = None

    foiz = round((natija["umumiy_ball"] / 189) * 100, 1)

    status_text = ""
    if ranks["class"]:
        status_text = f"📍 Siz sinfda <b>{ranks['class']}-o'rindasiz</b>."

    diff_text = ""
    if diff is not None:
        if diff > 0:
            diff_text = f"\n📈 Natijangiz <b>+{diff} ballga</b> yaxshilandi! Tabriklaymiz! 🎉"
        elif diff < 0:
            diff_text = f"\n📉 Natijangiz <b>{diff} ballga</b> o'zgardi. Keyingi safar yanada yaxshoreq bo'ladi! 💪"
        else:
            diff_text = "\n⏺ Natijangiz o'zgarmadi."

    matn = (
        f"✅ <b>Natija chiqdi!</b>\n\n"
        f"👤 Ism: <b>{talaba['ismlar']}</b>\n"
        f"🆔 Kod: <b>{talaba['kod']}</b>\n\n"
        f"📊 <b>Natijalar:</b>\n"
        f"  📘 Majburiy: {natija['majburiy']}/30 → <b>{natija['majburiy'] * 1.1:.1f}</b> ball\n"
        f"  📗 1-asosiy: {natija['asosiy_1']}/30 → <b>{natija['asosiy_1'] * 3.1:.1f}</b> ball\n"
        f"  📙 2-asosiy: {natija['asosiy_2']}/30 → <b>{natija['asosiy_2'] * 2.1:.1f}</b> ball\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"🏆 <b>Umumiy ball: {natija['umumiy_ball']} / 189</b>\n"
        f"📈 Foiz: <b>{foiz}%</b>\n\n"
        f"{status_text}{diff_text}\n\n"
        f"📜 Quyida sizning natijangiz aks etgan sertifikat biriktirilgan."
    )

    if is_notification_enabled(user_id, "notify_results"):
        # Sertifikat yaratish (Blocking bo'lgani uchun executor ishlatamiz)
        try:
            cert_gen = CertificateGenerator.from_db()
            sana = str(natija.get("test_sanasi", "Noaniq"))[:10]
            _sinf, _maktab = _parse_sinf_maktab(talaba.get("sinf", ""))

            loop = asyncio.get_event_loop()
            cert_path, cert_hash = await loop.run_in_executor(
                None,
                _generate_certificate_file,
                cert_gen,
                talaba["ismlar"],
                natija["umumiy_ball"],
                sana,
                talaba["kod"],
                _sinf,
                _maktab,
            )

            # Hash'ni database'ga saqlaymiz (blockchain tekshiruvi uchun)
            try:
                from database import get_connection, release_connection
                _conn = get_connection()
                _cur = _conn.cursor()
                _cur.execute(
                    "UPDATE talabalar SET cert_hash = %s WHERE kod = %s",
                    (cert_hash, talaba["kod"]),
                )
                _conn.commit()
                _cur.close()
                release_connection(_conn)
            except Exception as _he:
                print(f"cert_hash saqlashda xato: {_he}")

            await bot.send_document(
                user_id, FSInputFile(cert_path), caption=matn, parse_mode="HTML"
            )
            # Sertifikatni yuborgandan keyin o'chirib tashlaymiz
            if os.path.exists(cert_path):
                os.remove(cert_path)
        except Exception as e:
            # Agar sertifikatda xato bo'lsa, faqat matnni yuboramiz
            try:
                await bot.send_message(user_id, matn, parse_mode="HTML")
            except Exception:
                pass
    
    # Ota-onalarga ham xabar yuboramiz
    try:
        from parent import ota_onalarga_xabar_yuborish
        await ota_onalarga_xabar_yuborish(bot, talaba_kod, natija["umumiy_ball"])
    except Exception as e:
        print(f"Ota-onalarga xabar yuborishda xato: {e}")


# ─────────────────────────────────────────
# /admin va Kirish
# ─────────────────────────────────────────


@router.message(F.text == "/panel")
async def admin_panel_cmd(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "📊 Admin Web Panel:",
        reply_markup=admin_inline_menu()
    )


@router.message(F.text == "/admin")
async def admin_start(message: Message, state: FSMContext):
    if await admin_tekshir(state, message.from_user.id):
        await message.answer(
            "✅ Siz allaqachon admin rejimidasiz.",
            reply_markup=admin_menu_keyboard(),
        )
        await message.answer(
            "📊 Admin Web Panel:",
            reply_markup=admin_inline_menu()
        )
        return
    await state.set_state(AdminLogin.parol_kutish)
    await message.answer("🔐 Admin paroli kiriting:")


@router.message(AdminLogin.parol_kutish)
async def parol_tekshir(message: Message, state: FSMContext):
    if message.text.strip() == ADMIN_PASSWORD:
        # Admin sessiyasini boshlash
        from database import admin_session_start, get_active_admin_sessions

        admin_session_start(message.from_user.id, message.from_user.full_name)

        # Faol adminlarni tekshirish
        active_sessions = get_active_admin_sessions()
        if len(active_sessions) > 1:
            # Boshqa adminlar ham borligi haqida xabar
            other_admins = [
                s
                for s in active_sessions
                if s["admin_id"] != message.from_user.id
            ]
            warning_text = f"⚠️ <b>Diqqat!</b>\n\n"
            warning_text += (
                f"Sizdan tashqari {len(other_admins)} ta admin faol:\n"
            )
            for admin in other_admins:
                warning_text += f"• {admin['admin_name']}\n"
            warning_text += "\nIltimos, o'z ishlaringiz moslashtiring."
            await message.answer(warning_text, parse_mode="HTML")

        await state.update_data(admin=True)
        await state.set_state(None)
        await message.answer(
            "✅ Xush kelibsiz, admin!", 
            reply_markup=admin_menu_keyboard()
        )
        await message.answer(
            "📊 Web panelni ochish uchun quyidagi tugmani bosing:",
            reply_markup=admin_inline_menu()
        )
    else:
        await message.answer(
            "❌ Noto'g'ri parol. Qayta urinib ko'ring yoki /start bosing."
        )
        await state.clear()


@router.message(F.text == "🚪 Chiqish")
async def chiqish(message: Message, state: FSMContext):
    # Admin sessiyasini tugatish
    from database import admin_session_end

    admin_session_end(message.from_user.id)

    await state.clear()
    await message.answer("Admin rejimidan chiqdingiz.", reply_markup=None)

@router.message(F.text == "🌐 Web Admin Panel")
async def web_admin_panel_msg(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id): return
    await message.answer("📊 Admin Web Panel:", reply_markup=admin_inline_menu())

@router.message(F.text == "📝 Mashq Quiz (Web)")
async def web_quiz_panel_msg(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id): return
    from config import WEBAPP_URL
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 Quizlarni boshqarish", web_app=WebAppInfo(url=f"{WEBAPP_URL}/admin#quiz"))]
    ])
    await message.answer("📝 Mashq quizlarini boshqarish (Web):", reply_markup=kb)

# ── Mini-test yuborish ──────────────────────────────────────

@router.message(F.text == "📦 Mini-test yuborish")
async def mini_test_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id): return
    await state.set_state(MiniTestState.pdf_kutish)
    await message.answer(
        "📦 <b>Yangi mini-test yaratish</b>\n\n"
        "1️⃣ Test savollari yozilgan <b>PDF faylni</b> yuboring.\n"
        "(Agar PDF kerak bo'lmasa <code>-</code> deb yozing)",
        parse_mode="HTML"
    )

@router.message(MiniTestState.pdf_kutish)
async def mini_test_pdf(message: Message, state: FSMContext):
    if message.document and message.document.mime_type == "application/pdf":
        await state.update_data(pdf_file_id=message.document.file_id)
    elif message.text == "-":
        await state.update_data(pdf_file_id=None)
    else:
        await message.answer("❌ Iltimos, PDF fayl yuboring yoki <code>-</code> deb yozing.")
        return
        
    await state.set_state(MiniTestState.keys_kutish)
    await message.answer(
        "2️⃣ <b>Javoblar kalitini yuboring.</b>\n\n"
        "Format: <code>1a,2b,3c...</code> yoki Excel fayl yuboring."
    )

@router.message(MiniTestState.keys_kutish)
async def mini_test_keys(message: Message, state: FSMContext):
    keys = {}
    if message.text:
        # 1a,2b,3c formatini parse qilish
        try:
            parts = message.text.replace(" ", "").split(",")
            for p in parts:
                num = "".join(filter(str.isdigit, p))
                ans = "".join(filter(str.isalpha, p)).lower()
                if num and ans:
                    keys[int(num)] = ans
        except Exception:
            await message.answer("❌ Kalitlar formati noto'g'ri. Misol: 1a,2b,3c")
            return
    elif message.document:
        # Excel'dan o'qish (ixtiyoriy, hozircha matnli format ustuvor)
        await message.answer("⚠️ Excel formati hozircha qo'llab-quvvatlanmaydi. Iltimos matn ko'rinishida yuboring.")
        return
    
    if not keys:
        await message.answer("❌ Kalitlar topilmadi.")
        return
        
    await state.update_data(keys=keys)
    await state.set_state(MiniTestState.nomi_kutish)
    await message.answer("3️⃣ <b>Test nomini kiriting:</b>\nMasalan: <i>Matematika 1-blok</i>")

@router.message(MiniTestState.nomi_kutish)
async def mini_test_nomi(message: Message, state: FSMContext):
    await state.update_data(nomi=message.text.strip())
    await state.set_state(MiniTestState.fan_kutish)
    await message.answer("4️⃣ <b>Fan nomini kiriting:</b>")

@router.message(MiniTestState.fan_kutish)
async def mini_test_fan(message: Message, state: FSMContext):
    await state.update_data(fan=message.text.strip())
    data = await state.get_data()
    
    text = (
        f"✅ <b>Mini-test tayyor!</b>\n\n"
        f"📝 Nomi: <b>{data['nomi']}</b>\n"
        f"📚 Fan: <b>{data['fan']}</b>\n"
        f"📄 PDF: {'Biriktirildi' if data['pdf_file_id'] else 'Yo\'q'}\n"
        f"🔑 Kalitlar soni: <b>{len(data['keys'])}</b>\n\n"
        f"Tasdiqlaysizmi?"
    )
    await state.set_state(MiniTestState.tasdiq_kutish)
    await message.answer(text, parse_mode="HTML", reply_markup=tasdiqlash_keyboard())

@router.callback_query(F.data.startswith("tasdiq:"), MiniTestState.tasdiq_kutish)
async def mini_test_save(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if callback.data == "tasdiq:ha":
        data = await state.get_data()
        try:
            from database import mini_test_qosh, get_all_user_ids

            test_id = mini_test_qosh(
                data['nomi'], data['fan'], data['pdf_file_id'], data['keys'], callback.from_user.id
            )

            await callback.message.edit_text(f"✅ Mini-test saqlandi! (ID: {test_id})\nBroadcast boshlanmoqda...")

            # Broadcast logic
            user_ids = get_all_user_ids()
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🚀 Testni boshlash", callback_data=f"minitest_start:{test_id}")]
            ])

            text = (
                f"🔔 <b>YANGI MINI-TEST!</b>\n\n"
                f"📝 Nomi: <b>{data['nomi']}</b>\n"
                f"📚 Fan: <b>{data['fan']}</b>\n\n"
                f"Testni boshlash uchun quyidagi tugmani bosing."
            )

            asyncio.create_task(run_broadcast_simple(callback.bot, user_ids, text, kb))
        except Exception as e:
            logging.error(f"Mini-test saqlashda xato: {e}")
            await callback.message.edit_text(f"❌ Xatolik yuz berdi: {e}\n\nIltimos qayta urinib ko'ring.")
    else:
        await callback.message.edit_text("❌ Bekor qilindi.")

    await state.clear()

async def run_broadcast_simple(bot, user_ids, text, kb=None):
    for uid in user_ids:
        try:
            await bot.send_message(uid, text, parse_mode="HTML", reply_markup=kb)
            await asyncio.sleep(0.05)
        except Exception:
            continue


# ─────────────────────────────────────────
# Adminlarni boshqarish
# ─────────────────────────────────────────


@router.message(F.text == "👥 Adminlarni boshqarish")
async def admin_management_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "Adminlarni boshqarish menyusi:",
        reply_markup=admin_management_keyboard(),
    )


@router.callback_query(F.data.startswith("admin_manage:"))
async def admin_management_handler(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "list":
        admins = get_all_admins()
        if not admins:
            text = "⚠️ Hozircha adminlar yo'q."
        else:
            text = "👥 <b>Adminlar ro'yxati:</b>\n\n"
            for i, admin in enumerate(admins, 1):
                status = "✅ Faol" if admin["is_active"] else "❌ Faol emas"
                username = (
                    f"@{admin['username']}"
                    if admin["username"]
                    else "Username yo'q"
                )
                text += f"{i}. <b>{admin['full_name']}</b> ({admin['user_id']}) - {status}\n"
                text += f"   👤 {username}\n"
        await callback.message.edit_text(
            text, parse_mode="HTML", reply_markup=admin_management_keyboard()
        )

    elif action == "add":
        await state.set_state(AdminAdd.user_id_kutish)
        await callback.message.edit_text(
            "🆕 Yangi admin qo'shish uchun foydalanuvchi Telegram ID sini kiriting:\n\n"
            "📝 Masalan: 123456789\n\n"
            "❌ Bekor qilish uchun /cancel ni bosing.",
            reply_markup=None,
        )

    elif action == "remove":
        admins = get_all_admins()
        if not admins:
            await callback.answer("❌ Adminlar yo'q", show_alert=True)
            return

        buttons = []
        for admin in admins:
            if admin["is_active"]:
                buttons.append(
                    [
                        InlineKeyboardButton(
                            text=f"❌ {admin['full_name']} ({admin['user_id']})",
                            callback_data=f"admin_remove:{admin['user_id']}",
                        )
                    ]
                )
        buttons.append(
            [
                InlineKeyboardButton(
                    text="🔙 Orqaga", callback_data="admin_manage:back"
                )
            ]
        )

        await callback.message.edit_text(
            "❌ O'chiriladigan adminni tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
        )

    elif action == "back":
        await callback.message.edit_text("✅ Admin boshqaruvi yopildi.")
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )

    await callback.answer()


@router.message(AdminAdd.user_id_kutish)
async def admin_add_user_id(message: Message, state: FSMContext):
    if message.text == "/cancel":
        await state.clear()
        await message.answer(
            "❌ Admin qo'shish bekor qilindi.",
            reply_markup=admin_menu_keyboard(),
        )
        return

    try:
        user_id = int(message.text.strip())
        await state.update_data(user_id=user_id)
        await state.set_state(AdminAdd.tasdiq_kutish)

        await message.answer(
            f"🆕 <b>Yangi admin:</b> {user_id}\n\n"
            "✅ Tasdiqlaysizmi?\n\n"
            "✅ Ha - Admin qilish\n"
            "❌ Yo'q - Bekor qilish",
            reply_markup=tasdiqlash_keyboard(),
        )
    except ValueError:
        await message.answer("❌ Noto'g'ri ID format. Faqat raqam kiriting.")


@router.callback_query(F.data.startswith("admin_remove:"))
async def admin_remove_handler(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    user_id = int(callback.data.split(":")[1])

    # O'zini o'chirib bo'lmaydi
    if user_id == callback.from_user.id:
        await callback.answer(
            "❌ O'zingizni o'chira olmaysiz!", show_alert=True
        )
        return

    if remove_admin(user_id):
        await callback.answer(f"✅ Admin o'chirildi!", show_alert=True)
        # Admin ro'yxatini yangilab ko'rsatish
        admins = get_all_admins()
        if admins:
            text = "👥 <b>Adminlar ro'yxati:</b>\n\n"
            for i, admin in enumerate(admins, 1):
                status = "✅ Faol" if admin["is_active"] else "❌ Faol emas"
                username = (
                    f"@{admin['username']}"
                    if admin["username"]
                    else "Username yo'q"
                )
                text += f"{i}. <b>{admin['full_name']}</b> ({admin['user_id']}) - {status}\n"
                text += f"   👤 {username}\n"
        else:
            text = "⚠️ Hozircha adminlar yo'q."

        await callback.message.edit_text(
            text, parse_mode="HTML", reply_markup=admin_management_keyboard()
        )
    else:
        await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)


@router.callback_query(AdminAdd.tasdiq_kutish, F.data.startswith("tasdiq:"))
async def admin_add_tasdiq(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]
    data = await state.get_data()

    if action == "ha":
        user_id = data.get("user_id")
        if user_id:
            # Bu user_id ga tegishli user ma'lumotlarini olish (masalan, bot orqali)
            # Bu yerda oddiy qilib user_id va ismni qo'shamiz
            if add_admin(user_id, username=None, full_name=f"Admin {user_id}"):
                await callback.message.edit_text(
                    f"✅ <b>Yangi admin qo'shildi!</b>\n\n"
                    f"👤 User ID: <code>{user_id}</code>\n"
                    f"📝 To'liq ma'lumotlar uchun admin avval botdan foydalanishi kerak.",
                    parse_mode="HTML",
                )
                await callback.message.answer(
                    "Asosiy menyu:", reply_markup=admin_menu_keyboard()
                )
            else:
                await callback.message.edit_text(
                    "❌ Admin qo'shishda xatolik yuz berdi.",
                )
                await callback.message.answer(
                    "Asosiy menyu:", reply_markup=admin_menu_keyboard()
                )
        else:
            await callback.message.edit_text(
                "❌ Ma'lumotlar topilmadi.",
            )
            await callback.message.answer(
                "Asosiy menyu:", reply_markup=admin_menu_keyboard()
            )
    else:
        await callback.message.edit_text(
            "❌ Admin qo'shish bekor qilindi.",
        )
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )

    await state.clear()
    await callback.answer()


@router.message(F.text == "🔑 Parol")
async def admin_login(message: Message, state: FSMContext):
    await state.set_state(AdminLogin.parol_kutish)
    await message.answer("🔑 Admin parolini kiriting:")


# ─────────────────────────────────────────
# Yo'nalishlarni boshqarish
# ─────────────────────────────────────────


@router.message(F.text == "⚙️ Yo'nalishlarni boshqarish")
async def yonalish_boshqar_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "Yo'nalishlarni boshqarish menyusi:",
        reply_markup=yonalish_boshqarish_keyboard(),
    )


@router.callback_query(F.data.startswith("yonalish_boshqar:"))
async def yonalish_boshqar_actions(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "ro'yxat":
        yonalishlar = yonalish_ol()
        if not yonalishlar:
            text = "⚠️ Hozircha yo'nalishlar yo'q."
        else:
            text = "📋 <b>Mavjud yo'nalishlar ro'yxati:</b>\n\n"
            for i, y in enumerate(yonalishlar, 1):
                text += f"{i}. {y}\n"
        await callback.message.edit_text(
            text,
            parse_mode="HTML",
            reply_markup=yonalish_boshqarish_keyboard(),
        )

    elif action == "qosh":
        await state.set_state(YonalishBoshqar.nomi_kutish)
        await callback.message.edit_text(
            "📝 Yangi yo'nalish nomini kiriting:\n(Masalan: Matematika + Fizika)"
        )

    elif action == "ochir":
        await callback.message.edit_text(
            "❌ O'chirmoqchi bo'lgan yo'nalishni tanlang:",
            reply_markup=yonalish_ochirish_keyboard(),
        )

    elif action == "orqaga":
        await state.set_state(None)
        await callback.message.edit_text(
            "Yo'nalishlarni boshqarish menyusi:",
            reply_markup=yonalish_boshqarish_keyboard(),
        )
    await callback.answer()


@router.message(YonalishBoshqar.nomi_kutish)
async def yonalish_nomi_saqla(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    nomi = message.text.strip()
    if yonalish_qosh(nomi):
        await message.answer(
            f"✅ Yangi yo'nalish qo'shildi: <b>{nomi}</b>",
            parse_mode="HTML",
            reply_markup=admin_menu_keyboard(),
        )
    else:
        await message.answer("❌ Bu yo'nalish allaqachon mavjud.")
    await state.set_state(None)


@router.callback_query(F.data.startswith("yonalish_ochir:"))
async def yonalish_ochir_tasdiq(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    nomi = callback.data.split(":")[1]
    yonalish_ochir(nomi)
    await callback.answer(f"✅ {nomi} o'chirildi")
    await callback.message.edit_text(
        "❌ O'chirmoqchi bo'lgan yo'nalishni tanlang:",
        reply_markup=yonalish_ochirish_keyboard(),
    )


# ─────────────────────────────────────────
# Sinflarni boshqarish
# ─────────────────────────────────────────


@router.callback_query(F.data == "no_action")
async def no_action_handler(callback: CallbackQuery):
    """Sarlavha tugmalar uchun — hech narsa qilmaydi."""
    await callback.answer()


@router.callback_query(F.data == "cancel:admin_menu")
async def cancel_to_admin_menu(callback: CallbackQuery, state: FSMContext):
    """Orqaga admin menyusiga qaytish."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    await state.clear()
    await callback.message.edit_text(
        "✅ Admin bosh menyusiga qaytdingiz.",
        reply_markup=admin_menu_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("sinf_page:"))
async def sinf_page_navigate(callback: CallbackQuery):
    """Sinf tanlash klaviaturasida sahifa almashtirish."""
    page = int(callback.data.split(":")[1])
    await callback.message.edit_reply_markup(
        reply_markup=sinf_keyboard(page=page)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("sinf_ochir_page:"))
async def sinf_ochir_page_navigate(callback: CallbackQuery, state: FSMContext):
    """Sinf o'chirish klaviaturasida sahifa almashtirish."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    page = int(callback.data.split(":")[1])
    await callback.message.edit_reply_markup(
        reply_markup=sinf_ochirish_keyboard(page=page)
    )
    await callback.answer()


@router.message(F.text == "🏫 Sinflarni boshqarish")
async def sinf_boshqar_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "Sinflarni boshqarish menyusi:",
        reply_markup=sinf_boshqarish_keyboard(),
    )


@router.callback_query(F.data.startswith("sinf_boshqar:"))
async def sinf_boshqar_actions(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "ro'yxat":
        sinflar = sinf_ol_batafsil()
        if not sinflar:
            text = "⚠️ Hozircha sinflar yo'q."
        else:
            text = "📋 <b>Mavjud sinflar ro'yxati:</b>\n\n"
            joriy_maktab = None
            num = 0
            for s in sinflar:
                if s["maktab_nomi"] != joriy_maktab:
                    joriy_maktab = s["maktab_nomi"]
                    text += f"\n🏫 <b>{joriy_maktab}</b>\n"
                num += 1
                text += f"  {num}. {s['nomi']}\n"
        await callback.message.edit_text(
            text, parse_mode="HTML", reply_markup=sinf_boshqarish_keyboard()
        )

    elif action == "qosh":
        await state.set_state(SinfBoshqar.nomi_kutish)
        await callback.message.edit_text(
            "📝 Yangi sinf nomini kiriting:\n(Masalan: 11-A, 10-B)"
        )

    elif action == "ochir":
        await callback.message.edit_text(
            "❌ O'chirmoqchi bo'lgan sinfni tanlang:\n\n"
            "⚠️ <i>Agar sinfda o'quvchilar bo'lsa, o'chirish tasdiqlanadi.</i>",
            parse_mode="HTML",
            reply_markup=sinf_ochirish_keyboard(),
        )

    elif action == "orqaga":
        await state.set_state(None)
        await callback.message.edit_text(
            "Sinflarni boshqarish menyusi:",
            reply_markup=sinf_boshqarish_keyboard(),
        )
    await callback.answer()


@router.message(SinfBoshqar.nomi_kutish)
async def sinf_nomi_kutish(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    nomi = " ".join(message.text.strip().split())
    if " - " in nomi:
        nomi = nomi.split(" - ", 1)[0].strip()
    await state.update_data(sinf_nomi=nomi)

    maktablar = maktablar_ol()
    await state.set_state(SinfBoshqar.maktab_kutish)
    from keyboards import maktab_tanlash_keyboard

    await message.answer(
        f"🏫 <b>{nomi}</b> sinfi qaysi maktabga tegishli?",
        parse_mode="HTML",
        reply_markup=maktab_tanlash_keyboard(maktablar, "sinf_maktab"),
    )


@router.callback_query(
    SinfBoshqar.maktab_kutish, F.data.startswith("sinf_maktab:")
)
async def sinf_maktab_saqla(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    maktab_id = int(callback.data.split(":")[1])
    data = await state.get_data()
    nomi = data.get("sinf_nomi")

    if sinf_qosh(nomi, maktab_id):
        maktablar = maktablar_ol()
        maktab_nomi = next(
            (m["nomi"] for m in maktablar if m["id"] == maktab_id), "Noma'lum"
        )
        await callback.message.edit_text(
            f"✅ Yangi sinf qo'shildi: <b>{nomi} - {maktab_nomi}</b>",
            parse_mode="HTML",
        )
        await callback.message.answer(
            "Admin menyusi:", reply_markup=admin_menu_keyboard()
        )
    else:
        await callback.message.edit_text(
            "❌ Bu sinf ushbu maktabda allaqachon mavjud."
        )

    await state.clear()
    await callback.answer()


@router.callback_query(F.data.startswith("sinf_ochir_id:"))
async def sinf_ochir_tasdiq(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    sinf_id = int(callback.data.split(":")[1])

    # Sinf to'liq nomini olamiz
    full_nomi = sinf_id_dan_full_nomi(sinf_id)
    if not full_nomi:
        await callback.answer("❌ Sinf topilmadi.", show_alert=True)
        return

    # O'quvchilar sonini tekshiramiz
    talabalar_soni = sinf_talabalar_soni(sinf_id)

    if talabalar_soni > 0:
        # Tasdiqlash so'raymiz
        await state.update_data(
            ochiriladigan_sinf_id=sinf_id, ochiriladigan_sinf_nomi=full_nomi
        )
        await callback.message.edit_text(
            f"⚠️ <b>{full_nomi}</b> sinfida <b>{talabalar_soni} ta o'quvchi</b> mavjud.\n\n"
            f"O'quvchilar ma'lumotlari saqlanadi, faqat sinf o'chiriladi.\n\n"
            f"Davom etasizmi?",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Ha, o'chirish",
                            callback_data=f"sinf_ochir_tasdiqlandi:{sinf_id}",
                        ),
                        InlineKeyboardButton(
                            text="❌ Bekor qilish",
                            callback_data="sinf_boshqar:ochir",
                        ),
                    ]
                ]
            ),
        )
    else:
        # To'g'ridan-to'g'ri o'chiramiz
        sinf_ochir_id(sinf_id)
        await callback.answer(f"✅ {full_nomi} o'chirildi")
        await callback.message.edit_text(
            "❌ O'chirmoqchi bo'lgan sinfni tanlang:",
            reply_markup=sinf_ochirish_keyboard(),
        )


@router.callback_query(F.data.startswith("sinf_ochir_tasdiqlandi:"))
async def sinf_ochir_tasdiqlandi(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    sinf_id = int(callback.data.split(":")[1])
    full_nomi = sinf_id_dan_full_nomi(sinf_id) or "Sinf"
    sinf_ochir_id(sinf_id)
    await callback.answer(f"✅ {full_nomi} o'chirildi", show_alert=True)
    await callback.message.edit_text(
        "❌ O'chirmoqchi bo'lgan sinfni tanlang:",
        reply_markup=sinf_ochirish_keyboard(),
    )


# ─────────────────────────────────────────
# Test kalitlarini boshqarish
# ─────────────────────────────────────────


@router.message(F.text == "🔑 Test kalitlarini boshqarish")
async def kalit_boshqar_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "Test kalitlarini boshqarish menyusi:",
        reply_markup=kalit_boshqarish_keyboard(),
    )


@router.callback_query(F.data.startswith("kalit_boshqar:"))
async def kalit_boshqar_actions(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "ro'yxat":
        kalitlar = kalit_ol()
        if not kalitlar:
            await callback.message.edit_text(
                "⚠️ Hozircha testlar yo'q.",
                reply_markup=kalit_boshqarish_keyboard(),
            )
        else:
            await callback.message.edit_text(
                "📋 Testni tanlang:",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text=f"{'🔓' if k['holat']=='ochiq' else '🔒'} {k['test_nomi']}",
                                callback_data=f"kalit_detail:{k['test_nomi']}",
                            )
                        ]
                        for k in kalitlar
                    ]
                    + [
                        [
                            InlineKeyboardButton(
                                text="🔙 Orqaga",
                                callback_data="kalit_boshqar:orqaga",
                            )
                        ]
                    ]
                ),
            )

    elif action == "qosh":
        await state.set_state(KalitBoshqar.nomi_kutish)
        await callback.message.edit_text("📝 Yangi test nomini kiriting:")

    elif action == "yonalish_qosh":
        await callback.message.edit_text(
            "🎯 Qaysi yo'nalish uchun kalit qo'shmoqchisiz?",
            reply_markup=kalit_yonalish_tanlash_keyboard(),
        )

    elif action == "ochir":
        kalitlar = kalit_ol()
        await callback.message.edit_text(
            "❌ O'chirmoqchi bo'lgan testni tanlang:",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text=f"❌ {k['test_nomi']}",
                            callback_data=f"kalit_del:{k['test_nomi']}",
                        )
                    ]
                    for k in kalitlar
                ]
                + [
                    [
                        InlineKeyboardButton(
                            text="🔙 Orqaga",
                            callback_data="kalit_boshqar:orqaga",
                        )
                    ]
                ]
            ),
        )

    elif action == "orqaga":
        await state.set_state(None)
        await callback.message.edit_text(
            "Test kalitlarini boshqarish menyusi:",
            reply_markup=kalit_boshqarish_keyboard(),
        )
    await callback.answer()


@router.callback_query(F.data.startswith("kalit_detail:"))
async def kalit_detail(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    test_nomi = callback.data.split(":")[1]
    kalitlar = kalit_ol()
    k = next((x for x in kalitlar if x["test_nomi"] == test_nomi), None)
    if not k:
        return

    text = (
        f"📋 <b>Test:</b> {k['test_nomi']}\n"
        f"🎯 <b>Yo'nalish:</b> {k.get('yonalish') or 'Umumiy'}\n"
        f"🔑 <b>Kalitlar:</b> <code>{k['kalitlar']}</code>\n"
        f"📊 <b>Holat:</b> {'Ochiq' if k['holat']=='ochiq' else 'Yopiq'}"
    )
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=kalit_actions_keyboard(k["test_nomi"], k["holat"]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("kalit_status:"))
async def kalit_status_change(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    test_nomi = callback.data.split(":")[1]
    test_info = kalit_ol(test_nomi)
    if not test_info:
        await callback.answer("❌ Test topilmadi", show_alert=True)
        return
    yangi_holat = "yopiq" if test_info.get("holat") == "ochiq" else "ochiq"
    kalit_holat_ozgartir(test_nomi, yangi_holat)
    await kalit_detail(callback, state)


@router.callback_query(F.data.startswith("kalit_edit:"))
async def kalit_edit_start(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    test_nomi = callback.data.split(":")[1]
    await state.update_data(edit_test_nomi=test_nomi)
    await state.set_state(KalitBoshqar.edit_kalit_kutish)
    await callback.message.edit_text(
        f"✏️ <b>{test_nomi}</b> uchun yangi kalitlarni kiriting:",
        parse_mode="HTML",
    )
    await callback.answer()


@router.message(KalitBoshqar.edit_kalit_kutish)
async def kalit_edit_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = await state.get_data()
    test_nomi = data.get("edit_test_nomi")
    if not test_nomi:
        await message.answer("❌ Tahrirlash ma'lumoti topilmadi.")
        await state.set_state(None)
        return
    kalit_tahrirla(test_nomi, message.text.strip().lower())
    await message.answer(
        f"✅ <b>{test_nomi}</b> kalitlari yangilandi.",
        parse_mode="HTML",
        reply_markup=admin_menu_keyboard(),
    )
    await state.set_state(None)


@router.callback_query(F.data.startswith("kalit_del:"))
async def kalit_delete(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    test_nomi = callback.data.split(":")[1]
    kalit_ochir(test_nomi)
    await callback.answer(f"✅ {test_nomi} o'chirildi")
    await kalit_boshqar_actions(callback, state)


@router.message(KalitBoshqar.nomi_kutish)
async def kalit_nomi_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.update_data(test_nomi=message.text.strip())
    await state.set_state(KalitBoshqar.kalit_kutish)
    await message.answer("🔑 Kalitlarni kiriting (masalan: abcd...):")


@router.message(KalitBoshqar.kalit_kutish)
async def kalit_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = await state.get_data()
    kalitlar = message.text.strip().lower()
    kalit_qosh(data["test_nomi"], kalitlar)
    await message.answer(
        f"✅ Test qo'shildi: <b>{data['test_nomi']}</b>",
        parse_mode="HTML",
        reply_markup=admin_menu_keyboard(),
    )
    await state.set_state(None)


@router.callback_query(F.data.startswith("kalit_yonalish:"))
async def kalit_yonalish_selected(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    yonalish = callback.data.split(":")[1]
    await state.update_data(yonalish_nomi=yonalish)
    await state.set_state(KalitBoshqar.yonalish_nomi_kutish)
    await callback.message.edit_text(
        f"🎯 Yo'nalish: {yonalish}\n📝 Test nomini kiriting:"
    )
    await callback.answer()


@router.message(KalitBoshqar.yonalish_nomi_kutish)
async def kalit_yonalish_nomi_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.update_data(test_nomi=message.text.strip())
    await state.set_state(KalitBoshqar.yonalish_kalit_kutish)
    await message.answer("🔑 Kalitlarni kiriting:")


@router.message(KalitBoshqar.yonalish_kalit_kutish)
async def kalit_yonalish_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = await state.get_data()
    kalit_qosh(
        data["test_nomi"], message.text.strip().lower(), data["yonalish_nomi"]
    )
    await message.answer(
        f"✅ Yo'nalishli test qo'shildi: <b>{data['test_nomi']}</b>",
        parse_mode="HTML",
        reply_markup=admin_menu_keyboard(),
    )
    await state.set_state(None)


# ─────────────────────────────────────────
# O'quvchi QO'SHISH (Bildirishnoma bilan)
# ─────────────────────────────────────────


@router.message(F.text == "➕ O'quvchi qo'shish")
async def talaba_qosh_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    cfg = keyod_settings_ol()
    namuna = keyod_preview_list(cfg["prefix"], cfg["separator"], cfg["digits"], count=1)
    await message.answer(
        "📝 <b>O'quvchi qo'shish</b>\n\n"
        "Shaxsiy kod qanday kiritilsin?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(
                text=f"🔢 Avtogenerasiya  ({namuna[0]}...)",
                callback_data="idgen_avto"
            )],
            [InlineKeyboardButton(
                text="✍️ Qo'lda kiritish",
                callback_data="idgen_qolda"
            )],
            [InlineKeyboardButton(
                text="⚙️ Generator sozlamalari",
                callback_data="idgen_sozlama"
            )],
        ]),
    )


@router.callback_query(F.data == "idgen_avto")
async def talaba_avto_kod(cb: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, cb.from_user.id):
        return
    kod = keyod_yarat()
    await state.update_data(kod=kod)
    await state.set_state(TalabaQosh.yonalish_kutish)
    await cb.message.edit_text(
        f"✅ Avtomatik kod: <b>{kod}</b>\n\n"
        "Yo'nalishni tanlang:",
        parse_mode="HTML",
        reply_markup=yonalish_keyboard(),
    )


@router.callback_query(F.data == "idgen_qolda")
async def talaba_qolda_kod_start(cb: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, cb.from_user.id):
        return
    await state.set_state(TalabaQosh.kod_kutish)
    await cb.message.edit_text(
        "✍️ O'quvchining shaxsiy <b>kodini</b> kiriting:", parse_mode="HTML"
    )


@router.message(TalabaQosh.kod_kutish)
async def talaba_kod(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    kod = message.text.strip().upper()
    await state.update_data(kod=kod)
    await state.set_state(TalabaQosh.yonalish_kutish)
    await message.answer(
        f"✅ Kod: <b>{kod}</b>\n\nYo'nalishni tanlang:",
        reply_markup=yonalish_keyboard(),
        parse_mode="HTML",
    )


# ── Generator sozlamalari ─────────────────────────────────────────────────────

@router.callback_query(F.data == "idgen_sozlama")
async def idgen_sozlama_menu(cb: CallbackQuery, state: FSMContext):
    cfg = keyod_settings_ol()
    namunalar = keyod_preview_list(cfg["prefix"], cfg["separator"], cfg["digits"])
    await cb.message.edit_text(
        "⚙️ <b>Kod generatori sozlamalari</b>\n\n"
        f"🔤 Prefiks:    <code>{cfg['prefix']}</code>\n"
        f"🔢 Raqam uzunligi: <code>{cfg['digits']}</code>\n"
        f"➖ Ajratuvchi: <code>{'(yoq)' if not cfg['separator'] else cfg['separator']}</code>\n\n"
        f"📋 Namuna: <b>{' | '.join(namunalar)}</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔤 Prefiksni o'zgartirish", callback_data="idgen_set_prefix")],
            [InlineKeyboardButton(text="🔢 Raqam uzunligini o'zgartirish", callback_data="idgen_set_digits")],
            [InlineKeyboardButton(text="➖ Ajratuvchini o'zgartirish", callback_data="idgen_set_separator")],
            [InlineKeyboardButton(text="❌ Yopish", callback_data="idgen_close")],
        ]),
    )


@router.callback_query(F.data == "idgen_set_prefix")
async def idgen_set_prefix(cb: CallbackQuery, state: FSMContext):
    await state.set_state(IdGenSozlama.prefix_kutish)
    cfg = keyod_settings_ol()
    await cb.message.edit_text(
        f"🔤 Yangi <b>prefiks</b> kiriting.\n"
        f"Hozirgi: <code>{cfg['prefix']}</code>\n\n"
        f"Masalan: <code>SS</code>, <code>BS</code>, <code>A</code>",
        parse_mode="HTML",
    )


@router.message(IdGenSozlama.prefix_kutish)
async def idgen_prefix_saqlash(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    prefix = message.text.strip().upper()
    if not prefix.isalpha() or len(prefix) > 5:
        await message.answer("❌ Prefiks faqat harflardan iborat bo'lsin (1–5 ta). Qayta kiriting:")
        return
    keyod_settings_saqla(prefix=prefix)
    await state.clear()
    cfg = keyod_settings_ol()
    namunalar = keyod_preview_list(cfg["prefix"], cfg["separator"], cfg["digits"])
    await message.answer(
        f"✅ Prefiks <b>{prefix}</b> ga o'zgartirildi.\n"
        f"📋 Namuna: <b>{' | '.join(namunalar)}</b>",
        parse_mode="HTML",
    )


@router.callback_query(F.data == "idgen_set_digits")
async def idgen_set_digits(cb: CallbackQuery, state: FSMContext):
    await state.set_state(IdGenSozlama.digits_kutish)
    cfg = keyod_settings_ol()
    await cb.message.edit_text(
        f"🔢 Raqam <b>uzunligini</b> kiriting (1–6).\n"
        f"Hozirgi: <code>{cfg['digits']}</code>\n\n"
        f"3 → <code>001</code>   4 → <code>0001</code>",
        parse_mode="HTML",
    )


@router.message(IdGenSozlama.digits_kutish)
async def idgen_digits_saqlash(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    try:
        digits = int(message.text.strip())
        if not (1 <= digits <= 6):
            raise ValueError
    except ValueError:
        await message.answer("❌ 1 dan 6 gacha raqam kiriting:")
        return
    keyod_settings_saqla(digits=digits)
    await state.clear()
    cfg = keyod_settings_ol()
    namunalar = keyod_preview_list(cfg["prefix"], cfg["separator"], cfg["digits"])
    await message.answer(
        f"✅ Raqam uzunligi <b>{digits}</b> ga o'zgartirildi.\n"
        f"📋 Namuna: <b>{' | '.join(namunalar)}</b>",
        parse_mode="HTML",
    )


@router.callback_query(F.data == "idgen_set_separator")
async def idgen_set_separator(cb: CallbackQuery, state: FSMContext):
    await state.set_state(IdGenSozlama.separator_kutish)
    cfg = keyod_settings_ol()
    cur_sep = cfg["separator"] or "(yo'q)"
    await cb.message.edit_text(
        f"➖ <b>Ajratuvchi belgi</b> kiriting.\n"
        f"Hozirgi: <code>{cur_sep}</code>\n\n"
        f"<code>-</code> → SS-001\n"
        f"<code>_</code> → SS_001\n"
        f"Yo'q bo'lsin desangiz <code>.</code> yozing → SS001",
        parse_mode="HTML",
    )


@router.message(IdGenSozlama.separator_kutish)
async def idgen_separator_saqlash(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    text = message.text.strip()
    separator = "" if text == "." else text[:2]
    keyod_settings_saqla(separator=separator)
    await state.clear()
    cfg = keyod_settings_ol()
    namunalar = keyod_preview_list(cfg["prefix"], cfg["separator"], cfg["digits"])
    sep_display = separator or "(yo'q)"
    await message.answer(
        f"✅ Ajratuvchi <b>{sep_display}</b> ga o'zgartirildi.\n"
        f"📋 Namuna: <b>{' | '.join(namunalar)}</b>",
        parse_mode="HTML",
    )


@router.callback_query(F.data == "idgen_close")
async def idgen_close(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.edit_text("✅ Sozlamalar yopildi.")


# ════════════════════════════════════════════════════════════
# 🔢 KOD GENERATORI — OMMAVIY (alohida ishchi vosita)
# Maqsad: yangi o'quvchilar uchun bazaga kod yozish va tarqatish
# Oqim: nechta? → yo'nalish → bazaga saqlash → fayl yuborish
# ════════════════════════════════════════════════════════════

class KodGeneratori(StatesGroup):
    son_kutish      = State()
    yonalish_kutish = State()


@router.message(F.text == "🔢 Kod generatori")
async def kod_gen_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    cfg = keyod_settings_ol()
    namunalar = keyod_preview_list(cfg["prefix"], cfg["separator"], cfg["digits"], count=3)
    await state.set_state(KodGeneratori.son_kutish)
    await message.answer(
        "🔢 <b>Kod generatori</b>\n\n"
        f"⚙️ Format: <code>{cfg['prefix']}{cfg['separator']}{'0' * cfg['digits']}</code>\n"
        f"📋 Keyingi bo'sh kodlar: <b>{' | '.join(namunalar)}</b>\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📌 <b>Qanday ishlaydi:</b>\n"
        "1. Siz sonni tanlaysiz\n"
        "2. Yo'nalish tanlanadi\n"
        "3. Kodlar <b>bazaga saqlanadi</b>\n"
        "4. O'quvchiga kod beriladi\n"
        "5. O'quvchi <code>ULASH_KOD</code> yozadi → profil ulandi ✅\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Nechta kod kerak?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="10", callback_data="kodgen_son:10"),
                InlineKeyboardButton(text="20", callback_data="kodgen_son:20"),
                InlineKeyboardButton(text="30", callback_data="kodgen_son:30"),
            ],
            [
                InlineKeyboardButton(text="50", callback_data="kodgen_son:50"),
                InlineKeyboardButton(text="100", callback_data="kodgen_son:100"),
                InlineKeyboardButton(text="200", callback_data="kodgen_son:200"),
            ],
            [InlineKeyboardButton(text="⚙️ Format sozlamalari", callback_data="idgen_sozlama")],
            [InlineKeyboardButton(text="❌ Bekor", callback_data="kodgen_bekor")],
        ]),
    )


@router.callback_query(F.data.startswith("kodgen_son:"))
async def kod_gen_son_cb(cb: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, cb.from_user.id):
        return
    son = int(cb.data.split(":")[1])
    await state.update_data(son=son)
    await _kodgen_yonalish_so_ra(cb.message, state, edit=True)


@router.message(KodGeneratori.son_kutish)
async def kod_gen_son_msg(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    try:
        son = int(message.text.strip())
        if not (1 <= son <= 500):
            raise ValueError
    except ValueError:
        await message.answer("❌ 1 dan 500 gacha butun son kiriting:")
        return
    await state.update_data(son=son)
    await _kodgen_yonalish_so_ra(message, state, edit=False)


async def _kodgen_yonalish_so_ra(msg, state: FSMContext, edit: bool):
    """Yo'nalish tanlash klaviaturasini ko'rsatadi."""
    data = await state.get_data()
    son  = data["son"]

    yonalishlar = yonalish_ol()

    buttons = []
    if yonalishlar:
        buttons = [
            [InlineKeyboardButton(text=y, callback_data=f"kodgen_yonalish:{y}")]
            for y in yonalishlar
        ]
    buttons.append([InlineKeyboardButton(
        text="⏭️ O'tkazib yuborish (Aniqlanmagan)",
        callback_data="kodgen_yonalish:Aniqlanmagan"
    )])
    buttons.append([InlineKeyboardButton(text="❌ Bekor", callback_data="kodgen_bekor")])

    text = (
        f"✅ Son: <b>{son} ta</b>\n\n"
        "🎯 Yo'nalishni tanlang <i>(ixtiyoriy)</i>:"
    )
    if edit:
        await msg.edit_text(text, parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    else:
        await state.set_state(KodGeneratori.yonalish_kutish)
        await msg.answer(text, parse_mode="HTML",
                         reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))


@router.callback_query(F.data.startswith("kodgen_yonalish:"))
async def kod_gen_yonalish(cb: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, cb.from_user.id):
        return
    yonalish = cb.data.split(":", 1)[1]
    data     = await state.get_data()
    son      = data.get("son", 0)
    await state.clear()

    wait = await cb.message.edit_text(
        f"⏳ <b>{son} ta</b> kod yaratilmoqda va bazaga saqlanmoqda...",
        parse_mode="HTML",
    )
    await _kod_gen_saqlash_va_yuborish(wait, son, yonalish)


async def _kod_gen_saqlash_va_yuborish(msg, son: int, yonalish: str):
    """Kodlarni bazaga saqlaydi va fayl yuboradi."""
    import io
    from aiogram.types import BufferedInputFile

    natija = keyod_ommaviy_yarat(son, yonalish)
    kodlar  = natija["saqlangan"]
    xatolar = natija["xato"]
    cfg     = keyod_settings_ol()

    if not kodlar:
        await msg.edit_text("❌ Hech qanday kod saqlanmadi. DB xatosini tekshiring.")
        return

    # ── Xabar matni ──
    preview_son = min(len(kodlar), 25)
    preview     = kodlar_txt_fayl(kodlar[:preview_son], ustun=5)
    qoldi_str   = f"\n<i>...va yana {len(kodlar) - preview_son} ta</i>" if len(kodlar) > preview_son else ""
    xato_str    = f"\n⚠️ {len(xatolar)} ta kod saqlanmadi (DB xato)" if xatolar else ""

    matn = (
        f"✅ <b>{len(kodlar)} ta kod bazaga saqlandi!</b>\n"
        f"🎯 Yo'nalish: <b>{yonalish}</b>\n"
        f"📋 <b>{kodlar[0]}</b> → <b>{kodlar[-1]}</b>\n"
        f"{xato_str}\n\n"
        f"<code>{preview}</code>{qoldi_str}\n\n"
        "📄 To'liq ro'yxat quyida 👇"
    )

    # ── .txt fayl ──
    txt_tarkib = (
        f"Shaxsiy kodlar ro'yxati\n"
        f"Yo'nalish : {yonalish}\n"
        f"Format    : {cfg['prefix']}{cfg['separator']}{'0' * cfg['digits']}\n"
        f"Jami      : {len(kodlar)} ta\n"
        f"{'─' * 38}\n\n"
        + kodlar_txt_fayl(kodlar, ustun=5)
        + "\n\n"
        "─────────────────────────────────────────\n"
        "Har bir o'quvchiga BITTA kod berilsin.\n"
        "Botga ulash: Telegram botga  ULASH_[KOD]  deb yuboring.\n"
        "Misol: ULASH_SS001\n"
    )

    # ── .xlsx fayl ──
    xlsx_fayl = None
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kodlar"

        ws.merge_cells("A1:D1")
        ws["A1"] = f"Shaxsiy kodlar — {yonalish} — {len(kodlar)} ta"
        ws["A1"].font      = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        sarlavhalar = ["№", "Shaxsiy kod", "O'quvchi ismi", "Imzo"]
        for col, s in enumerate(sarlavhalar, start=1):
            c = ws.cell(row=2, column=col, value=s)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="2E75B6")
            c.alignment = Alignment(horizontal="center")

        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for i, kod in enumerate(kodlar, start=1):
            row = i + 2
            for col, val in enumerate([i, kod, "", ""], start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = Alignment(horizontal="center")
                c.border    = border

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_fayl = BufferedInputFile(
            buf.getvalue(),
            filename=f"kodlar_{yonalish.replace(' ', '_')}.xlsx",
        )
    except Exception:
        pass

    # ── Yuborish ──
    try:
        await msg.edit_text(matn, parse_mode="HTML")
    except Exception:
        await msg.answer(matn, parse_mode="HTML")

    await msg.answer_document(
        BufferedInputFile(txt_tarkib.encode("utf-8"),
                          filename=f"kodlar_{yonalish.replace(' ', '_')}.txt"),
        caption="📄 <b>Barcha kodlar — matn (.txt)</b>",
        parse_mode="HTML",
    )
    if xlsx_fayl:
        await msg.answer_document(
            xlsx_fayl,
            caption=(
                "📊 <b>Barcha kodlar — Excel (.xlsx)</b>\n"
                "<i>Chop etish uchun: № | Kod | Ism | Imzo ustunlari</i>"
            ),
            parse_mode="HTML",
        )


@router.callback_query(F.data == "kodgen_bekor")
async def kodgen_bekor(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.edit_text("❌ Bekor qilindi.")


@router.callback_query(
    TalabaQosh.yonalish_kutish, F.data.startswith("yonalish_idx:")
)
async def talaba_yonalish(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    try:
        idx = int(callback.data.split(":", 1)[1])
        yonalishlar = yonalish_ol()
        yonalish = yonalishlar[idx]
    except (ValueError, IndexError):
        await callback.answer("❌ Yo'nalish topilmadi.", show_alert=True)
        return
    await state.update_data(yonalish=yonalish)
    await state.set_state(TalabaQosh.maktab_kutish)
    maktablar = maktablar_ol()
    if not maktablar:
        await callback.message.edit_text(
            "❌ Maktablar topilmadi. Avval '🏫 Maktablarni boshqarish' orqali maktab qo'shing."
        )
        await state.set_state(None)
        await callback.answer()
        return
    await callback.message.edit_text(
        f"🎯 Yo'nalish: <b>{yonalish}</b>\n\n🏫 Maktabni tanlang:",
        reply_markup=maktab_tanlash_keyboard(maktablar, "talaba_maktab"),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(
    TalabaQosh.maktab_kutish, F.data.startswith("talaba_maktab:")
)
async def talaba_maktab(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    try:
        maktab_id = int(callback.data.split(":")[1])
    except ValueError:
        await callback.answer("❌ Maktab topilmadi.", show_alert=True)
        return

    maktab = next((m for m in maktablar_ol() if int(m["id"]) == maktab_id), None)
    if not maktab:
        await callback.answer("❌ Maktab topilmadi.", show_alert=True)
        return

    await state.update_data(maktab_id=maktab_id, maktab_nomi=maktab["nomi"])
    await state.set_state(TalabaQosh.sinf_kutish)
    await callback.message.edit_text(
        f"🏫 Maktab: <b>{maktab['nomi']}</b>\n\nSinfni tanlang:",
        reply_markup=sinf_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(TalabaQosh.sinf_kutish, F.data.startswith("sinf_id:"))
async def talaba_sinf(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    sinf_id = int(callback.data.split(":")[1])
    sinf = sinf_id_dan_full_nomi(sinf_id)
    if not sinf:
        await callback.answer("❌ Sinf topilmadi.", show_alert=True)
        return
    data = await state.get_data()
    maktab_nomi = data.get("maktab_nomi")
    if maktab_nomi and not sinf.endswith(f" - {maktab_nomi}"):
        await callback.answer(
            "❌ Tanlangan sinf boshqa maktabga tegishli. Iltimos, to'g'ri sinfni tanlang.",
            show_alert=True,
        )
        return
    await state.update_data(sinf=sinf)
    await state.set_state(TalabaQosh.ismlar_kutish)
    await callback.message.edit_text(
        f"🏫 Sinf: <b>{sinf}</b>\n\n👤 O'quvchining ism-familiyasini kiriting:",
        parse_mode="HTML",
    )
    await callback.answer()


@router.message(TalabaQosh.ismlar_kutish)
async def talaba_ismlar(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    ismlar = message.text.strip()
    await state.update_data(ismlar=ismlar)
    await state.set_state(TalabaQosh.majburiy_kutish)
    await message.answer(
        f"👤 Ism: <b>{ismlar}</b>\n\n📘 <b>Majburiy fanlar</b> bo'yicha to'g'ri javoblar sonini kiriting (0-30):",
        parse_mode="HTML",
    )


@router.message(TalabaQosh.majburiy_kutish)
async def talaba_majburiy(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    son = _son_tekshir(message.text)
    if son is None:
        await message.answer(
            "❌ Noto'g'ri son. 0 dan 30 gacha raqam kiriting:"
        )
        return
    await state.update_data(majburiy=son)
    await state.set_state(TalabaQosh.asosiy1_kutish)
    await message.answer(
        "📗 <b>1-asosiy fan</b> bo'yicha to'g'ri javoblar sonini kiriting (0-30):",
        parse_mode="HTML",
    )


@router.message(TalabaQosh.asosiy1_kutish)
async def talaba_asosiy1(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    son = _son_tekshir(message.text)
    if son is None:
        await message.answer(
            "❌ Noto'g'ri son. 0 dan 30 gacha raqam kiriting:"
        )
        return
    await state.update_data(asosiy1=son)
    await state.set_state(TalabaQosh.asosiy2_kutish)
    await message.answer(
        "📙 <b>2-asosiy fan</b> bo'yicha to'g'ri javoblar sonini kiriting (0-30):",
        parse_mode="HTML",
    )


@router.message(TalabaQosh.asosiy2_kutish)
async def talaba_asosiy2(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    son = _son_tekshir(message.text)
    if son is None:
        await message.answer(
            "❌ Noto'g'ri son. 0 dan 30 gacha raqam kiriting:"
        )
        return
    await state.update_data(asosiy2=son)
    data = await state.get_data()

    ball = ball_hisobla(data["majburiy"], data["asosiy1"], data["asosiy2"])
    await state.update_data(umumiy_ball=ball)

    text = (
        f"🧐 <b>Ma'lumotlarni tekshiring:</b>\n\n"
        f"🆔 Kod: <b>{data['kod']}</b>\n"
        f"👤 Ism: <b>{data['ismlar']}</b>\n"
        f"🏫 Maktab: <b>{data.get('maktab_nomi', 'Nomalum')}</b>\n"
        f"🏫 Sinf: <b>{data['sinf']}</b>\n"
        f"🎯 Yo'nalish: <b>{data['yonalish']}</b>\n\n"
        f"📊 <b>Natijalar:</b>\n"
        f"  📘 Majburiy: {data['majburiy']}/30\n"
        f"  📗 1-asosiy: {data['asosiy1']}/30\n"
        f"  📙 2-asosiy: {data['asosiy2']}/30\n"
        f"━━━━━━━━━━━━━━\n"
        f"🏆 <b>Umumiy ball: {ball}</b>"
    )
    await state.set_state(TalabaQosh.tasdiq_kutish)
    await message.answer(
        text, parse_mode="HTML", reply_markup=tasdiqlash_keyboard()
    )


@router.callback_query(TalabaQosh.tasdiq_kutish, F.data.startswith("tasdiq:"))
async def talaba_tasdiq(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "ha":
        data = await state.get_data()
        talaba = {
            "kod": data["kod"],
            "ismlar": data["ismlar"],
            "yonalish": data["yonalish"],
            "sinf": data["sinf"],
        }
        natija = {
            "majburiy": data["majburiy"],
            "asosiy_1": data["asosiy1"],
            "asosiy_2": data["asosiy2"],
            "umumiy_ball": data["umumiy_ball"],
        }

        talaba_qosh(
            talaba["kod"],
            talaba["yonalish"],
            talaba["sinf"],
            talaba["ismlar"],
            data.get("maktab_id"),
        )
        natija_qosh(
            talaba["kod"],
            natija["majburiy"],
            natija["asosiy_1"],
            natija["asosiy_2"],
        )

        await callback.message.edit_text(
            "✅ Ma'lumotlar muvaffaqiyatli saqlandi!", reply_markup=None
        )
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )

        # Bildirishnomalar
        asyncio.create_task(
            bildirishnoma_yuborish(callback.bot, talaba["kod"], natija, talaba)
        )
        asyncio.create_task(
            guruhlarga_yangi_oquvchi_yuborish(callback.bot, talaba, natija)
        )
        # Ota-onalarga avtomatik xabarnoma
        asyncio.create_task(
            _ota_onalarga_bildirish(callback.bot, talaba["kod"], natija["umumiy_ball"])
        )

    else:
        await callback.message.edit_text(
            "❌ Bekor qilindi.", reply_markup=None
        )
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )

    await state.set_state(None)
    await callback.answer()


# ─────────────────────────────────────────
# Natijani Tahrirlash
# ─────────────────────────────────────────


@router.message(F.text == "✏️ Natijani tahrirlash")
async def natija_tahrirlash_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.set_state(NatijaTahrirlash.kod_kutish)
    await message.answer(
        "🔍 Tahrirlash uchun o'quvchi <b>kodini</b> kiriting:",
        parse_mode="HTML",
    )


@router.message(NatijaTahrirlash.kod_kutish)
async def natija_tahrir_kod(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)
    if not talaba:
        await message.answer("❌ Bunday kodli o'quvchi topilmadi.")
        await state.set_state(None)
        return

    await state.update_data(kod=kod)
    await state.set_state(NatijaTahrirlash.majburiy_kutish)
    await message.answer(
        f"👤 O'quvchi: <b>{talaba['ismlar']}</b>\n\n📘 <b>Majburiy fanlar</b> bo'yicha yangi <b>to'g'ri javoblar sonini</b> kiriting (0-30):",
        parse_mode="HTML",
    )


@router.message(NatijaTahrirlash.majburiy_kutish)
async def natija_tahrir_majburiy(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    son = _son_tekshir(message.text)
    if son is None:
        await message.answer("❌ Noto'g'ri son.")
        return
    await state.update_data(majburiy=son)
    await state.set_state(NatijaTahrirlash.asosiy1_kutish)
    await message.answer(
        "📗 <b>1-asosiy fan</b> bo'yicha yangi <b>to'g'ri javoblar sonini</b> kiriting (0-30):", parse_mode="HTML"
    )


@router.message(NatijaTahrirlash.asosiy1_kutish)
async def natija_tahrir_asosiy1(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    son = _son_tekshir(message.text)
    if son is None:
        await message.answer("❌ Noto'g'ri son.")
        return
    await state.update_data(asosiy1=son)
    await state.set_state(NatijaTahrirlash.asosiy2_kutish)
    await message.answer(
        "📙 <b>2-asosiy fan</b> bo'yicha yangi <b>to'g'ri javoblar sonini</b> kiriting (0-30):", parse_mode="HTML"
    )


@router.message(NatijaTahrirlash.asosiy2_kutish)
async def natija_tahrir_asosiy2(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    son = _son_tekshir(message.text)
    if son is None:
        await message.answer("❌ Noto'g'ri son.")
        return
    data = await state.get_data()
    ball = ball_hisobla(data["majburiy"], data["asosiy1"], son)

    # Ma'lumotlarni yangilash
    natija_qosh(data["kod"], data["majburiy"], data["asosiy1"], son)

    # O'quvchiga avtomatik xabarnoma (profil ulangan bo'lsa)
    try:
        talaba = talaba_topish(data["kod"])
        if talaba:
            natija = {
                "majburiy": data["majburiy"],
                "asosiy_1": data["asosiy1"],
                "asosiy_2": son,
                "umumiy_ball": ball,
            }
            asyncio.create_task(
                bildirishnoma_yuborish(message.bot, data["kod"], natija, talaba)
            )
    except Exception:
        pass

    # Ota-onalarga avtomatik xabarnoma
    asyncio.create_task(
        _ota_onalarga_bildirish(message.bot, data["kod"], ball)
    )

    await message.answer(
        f"✅ Natija yangilandi! Yangi ball: <b>{ball}</b>",
        parse_mode="HTML",
        reply_markup=admin_menu_keyboard(),
    )
    await state.set_state(None)


# ─────────────────────────────────────────
# Exceldan IMPORT (Bildirishnoma bilan)
# ─────────────────────────────────────────


@router.message(F.text == "📥 Exceldan import")
async def excel_import_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.set_state(ExcelImport.fayl_kutish)
    await message.answer(
        "📥 <b>Excel faylini yuboring.</b>\n\n"
        "Fayl ustunlari tartibi:\n"
        "Kod | Ism | Sinf | Yo'nalish | Maktab(raqam yoki nom) | Majburiy | Asosiy1 | Asosiy2\n\n"
        "⚠️ Maktab ustuni majburiy.",
        parse_mode="HTML",
    )


@router.message(ExcelImport.fayl_kutish, F.document)
async def excel_import_process(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    file_id = message.document.file_id
    file = await message.bot.get_file(file_id)
    file_path = file.file_path

    download_path = f"import_{message.from_user.id}.xlsx"
    await message.bot.download_file(file_path, download_path)

    try:
        df = pd.read_excel(download_path)
        normalized_columns = {
            str(col)
            .strip()
            .lower()
            .replace("'", "")
            .replace("`", "")
            .replace("’", "")
            .replace("ʻ", "")
            .replace("-", "")
            .replace("_", "")
            .replace(" ", ""): col
            for col in df.columns
        }

        alias_map = {
            "kod": ["kod", "idkod", "oquvchikodi"],
            "ismlar": ["ism", "ismlar", "ismfamiliya", "fio", "fullname"],
            "sinf": ["sinf", "class", "klass"],
            "yonalish": ["yonalish", "yunalish", "direction"],
            "maktab": [
                "maktab",
                "maktabid",
                "maktabraqami",
                "maktabnomi",
                "school",
                "schoolid",
                "schoolname",
            ],
            "majburiy": ["majburiy", "majburiyfan", "majburiyfanlar"],
            "asosiy1": ["asosiy1", "1asosiy", "asosiyfan1", "asosiybirinchi"],
            "asosiy2": ["asosiy2", "2asosiy", "asosiyfan2", "asosiyikkinchi"],
        }

        header_mode = True
        resolved_cols = {}
        for key, aliases in alias_map.items():
            found = None
            for alias in aliases:
                if alias in normalized_columns:
                    found = normalized_columns[alias]
                    break
            if found is None:
                header_mode = False
                break
            resolved_cols[key] = found

        maktablar = maktablar_ol()
        if not maktablar:
            await message.answer(
                "❌ Maktablar topilmadi. Avval maktablar ro'yxatini to'ldiring.",
                reply_markup=admin_menu_keyboard(),
            )
            await state.set_state(None)
            return

        if header_mode and "maktab" not in resolved_cols:
            await message.answer(
                "❌ Excelda <b>Maktab</b> ustuni topilmadi. "
                "Maktab raqami yoki nomi ko'rsatilishi shart.",
                parse_mode="HTML",
                reply_markup=admin_menu_keyboard(),
            )
            await state.set_state(None)
            return
        if not header_mode and len(df.columns) < 8:
            await message.answer(
                "❌ Excel format noto'g'ri. Header bo'lmasa kamida 8 ta ustun bo'lishi kerak:\n"
                "Kod | Ism | Sinf | Yo'nalish | Maktab | Majburiy | Asosiy1 | Asosiy2",
                reply_markup=admin_menu_keyboard(),
            )
            await state.set_state(None)
            return

        await message.answer(
            "⏳ <b>Import boshlandi...</b>\n\nKatta hajmdagi ma'lumotlar qayta ishlanmoqda. Tugagach sizga xabar beriladi.",
            parse_mode="HTML"
        )
        
        asyncio.create_task(run_excel_import_task(message, df, header_mode, resolved_cols, download_path, state))
        return

    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
        if os.path.exists(download_path):
            os.remove(download_path)
        await state.set_state(None)

async def run_excel_import_task(message: Message, df, header_mode, resolved_cols, download_path, state):
    from database import talaba_qosh, natija_qosh, maktablar_ol, ball_hisobla
    from student import bildirishnoma_yuborish
    from parent import _ota_onalarga_bildirish
    
    try:
        maktablar = maktablar_ol()
        count = 0
        skipped = 0
        
        for _, row in df.iterrows():
            try:
                if header_mode:
                    kod = str(row[resolved_cols["kod"]]).strip().upper()
                    ismlar = str(row[resolved_cols["ismlar"]]).strip()
                    sinf = str(row[resolved_cols["sinf"]]).strip()
                    yonalish = str(row[resolved_cols["yonalish"]]).strip()
                    maktab_raw = row[resolved_cols["maktab"]]
                    majburiy = int(row[resolved_cols["majburiy"]])
                    asosiy1 = int(row[resolved_cols["asosiy1"]])
                    asosiy2 = int(row[resolved_cols["asosiy2"]])
                else:
                    kod = str(row.iloc[0]).strip().upper()
                    ismlar = str(row.iloc[1]).strip()
                    sinf = str(row.iloc[2]).strip()
                    yonalish = str(row.iloc[3]).strip()
                    maktab_raw = row.iloc[4]
                    majburiy = int(row.iloc[5])
                    asosiy1 = int(row.iloc[6])
                    asosiy2 = int(row.iloc[7])

                if not kod or kod.lower() == "nan":
                    continue

                maktab = _resolve_maktab(maktab_raw, maktablar)
                if not maktab:
                    skipped += 1
                    continue

                if " - " not in sinf:
                    sinf = f"{sinf} - {maktab['nomi']}"

                ball = ball_hisobla(majburiy, asosiy1, asosiy2)

                talaba_qosh(kod, yonalish, sinf, ismlar, maktab["id"])
                natija_qosh(kod, majburiy, asosiy1, asosiy2)

                talaba = {
                    "kod": kod,
                    "ismlar": ismlar,
                    "yonalish": yonalish,
                    "sinf": sinf,
                    "maktab": maktab["nomi"],
                }
                natija = {
                    "majburiy": majburiy,
                    "asosiy_1": asosiy1,
                    "asosiy_2": asosiy2,
                    "umumiy_ball": ball,
                }
                
                asyncio.create_task(bildirishnoma_yuborish(message.bot, kod, natija, talaba))
                asyncio.create_task(_ota_onalarga_bildirish(message.bot, kod, ball))

                count += 1
                if count % 20 == 0:
                    await asyncio.sleep(0.01)
                    
            except Exception as e:
                logging.error(f"Import row error: {e}")
                skipped += 1
                continue

        await message.answer(
            f"✅ <b>Import yakunlandi!</b>\n\n"
            f"👤 Import qilinganlar: <b>{count}</b>\n"
            f"⚠️ O'tkazib yuborilganlar: <b>{skipped}</b>",
            parse_mode="HTML",
            reply_markup=admin_menu_keyboard(),
        )
    except Exception as e:
        await message.answer(f"❌ Import jarayonida xatolik: {e}")
    finally:
        if os.path.exists(download_path):
            os.remove(download_path)
        await state.clear()


# ─────────────────────────────────────────
# Statistika va Sozlamalar
# ─────────────────────────────────────────


@router.message(F.text == "📊 Statistika")
async def stats_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    
    # Get all schools for selection
    maktablar = maktablar_ol()
    if not maktablar:
        await message.answer("⚠️ Hozircha maktablar mavjud emas.")
        return
    
    await message.answer(
        "📊 <b>Statistika uchun maktabni tanlang:</b>\n\n"
        "Qaysi maktabning statistikasini ko'rmoqchisiz?",
        parse_mode="HTML",
        reply_markup=maktab_tanlash_keyboard(maktablar, "stats_maktab")
    )


@router.callback_query(F.data.startswith("stats_maktab:"))
async def stats_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """School selection for statistics callback handler."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    try:
        maktab_id = int(callback.data.split(":")[1])
        maktablar = maktablar_ol()
        maktab = next((m for m in maktablar if m['id'] == maktab_id), None)
        
        if not maktab:
            await callback.answer("❌ Maktab topilmadi", show_alert=True)
            return
        
        # Store selected school in state
        await state.update_data(selected_maktab_id=maktab_id, selected_maktab_nomi=maktab['nomi'])
        
        # Get statistics for this school
        from database import statistika_by_maktab
        res = statistika_by_maktab(maktab_id)
        
        if not res:
            text = f"📊 <b>{maktab['nomi']} statistikasi:</b>\n\n⚠️ Ma'lumotlar yo'q."
            await callback.message.edit_text(text, parse_mode="HTML")
        else:
            text = (
                f"📊 <b>{maktab['nomi']} statistikasi:</b>\n\n"
                f"👥 Jami o'quvchilar: <b>{res['jami']} ta</b>\n"
                f"📈 O'rtacha ball: <b>{res['ortacha']} ball</b>\n"
                f"🏆 Eng yuqori ball: <b>{res['eng_yuqori']} ball</b>\n"
                f"📉 Eng past ball: <b>{res['eng_past']} ball</b>\n\n"
                f"🔍 <b>Batafsil statistika uchun:</b>"
            )
            await callback.message.edit_text(
                text, parse_mode="HTML", reply_markup=stats_keyboard()
            )
        
    except (ValueError, IndexError):
        await callback.answer("❌ Noto'g'ri maktab tanlovi", show_alert=True)
    
    await callback.answer()


@router.callback_query(F.data.startswith("stats:"))
async def stats_callback_handler(callback: CallbackQuery, state: FSMContext):
    """Statistika callback handlerlari"""
    if not await admin_tekshir(state, callback.from_user.id):
        await callback.answer("❌ Sizda ruxsat yo'q!", show_alert=True)
        return

    action = callback.data.split(":")[1]

    if action == "direction":
        # Yo'nalishlar bo'yicha o'rtacha ballar
        from database import get_avg_score_by_direction, get_avg_score_by_direction_by_maktab
        
        data = await state.get_data()
        maktab_id = data.get("selected_maktab_id")
        maktab_nomi = data.get("selected_maktab_nomi", "Noma'lum maktab")
        
        if maktab_id:
            stats = get_avg_score_by_direction_by_maktab(maktab_id)
            title = f"🎯 <b>{maktab_nomi} - Yo'nalishlar bo'yicha o'rtacha ballar:</b>\n\n"
        else:
            stats = get_avg_score_by_direction()
            title = "🎯 <b>Yo'nalishlar bo'yicha o'rtacha ballar:</b>\n\n"

        if not stats:
            text = title + "⚠️ Ma'lumotlar yo'q."
            await callback.message.edit_text(text, parse_mode="HTML")
        else:
            # Grafik yasash
            chart_file = create_direction_chart(stats)

            # Matnli ma'lumot
            text = title
            for i, stat in enumerate(stats, 1):
                yonalish = stat["yonalish"] or "Umumiy"
                avg = stat["avg_score"]
                text += f"{i}. <b>{yonalish}</b>: {avg:.1f} ball\n"

            # Grafikni jo'natish
            if chart_file and os.path.exists(chart_file):
                await callback.message.answer_photo(
                    photo=FSInputFile(chart_file),
                    caption=text,
                    parse_mode="HTML",
                )
                # Faylni o'chirish
                os.remove(chart_file)
            else:
                await callback.message.edit_text(text, parse_mode="HTML")

    elif action == "class":
        # Sinflar bo'yicha o'rtacha ballar
        from database import get_class_comparison, get_class_comparison_by_maktab
        
        data = await state.get_data()
        maktab_id = data.get("selected_maktab_id")
        maktab_nomi = data.get("selected_maktab_nomi", "Noma'lum maktab")
        
        if maktab_id:
            stats = get_class_comparison_by_maktab(maktab_id)
            title = f"🏫 <b>{maktab_nomi} - Sinflar bo'yicha o'rtacha ballar:</b>\n\n"
        else:
            stats = get_class_comparison()
            title = "🏫 <b>Sinflar bo'yicha o'rtacha ballar:</b>\n\n"

        if not stats:
            text = title + "⚠️ Ma'lumotlar yo'q."
            await callback.message.edit_text(text, parse_mode="HTML")
        else:
            # Grafik yasash
            chart_file = create_class_chart(stats)

            # Matnli ma'lumot
            text = title
            for i, stat in enumerate(stats, 1):
                sinf = stat["sinf"]
                avg = stat["avg_score"]
                text += f"{i}. <b>{sinf}</b>: {avg:.1f} ball\n"

            # Grafikni jo'natish
            if chart_file and os.path.exists(chart_file):
                await callback.message.answer_photo(
                    photo=FSInputFile(chart_file),
                    caption=text,
                    parse_mode="HTML",
                )
                # Faylni o'chirish
                os.remove(chart_file)
            else:
                await callback.message.edit_text(text, parse_mode="HTML")

    elif action == "improved":
        # Eng ko'p o'sganlar
        from database import get_most_improved_students, get_most_improved_students_by_maktab
        
        data = await state.get_data()
        maktab_id = data.get("selected_maktab_id")
        maktab_nomi = data.get("selected_maktab_nomi", "Noma'lum maktab")
        
        if maktab_id:
            students = get_most_improved_students_by_maktab(maktab_id)
            title = f"🚀 <b>{maktab_nomi} - Eng ko'p o'sgan o'quvchilar:</b>\n\n"
        else:
            students = get_most_improved_students()
            title = "🚀 <b>Eng ko'p o'sgan o'quvchilar:</b>\n\n"

        if not students:
            text = title + "⚠️ Ma'lumotlar yo'q."
        else:
            text = title
            for i, student in enumerate(students[:10], 1):  # Top 10
                name = student["ismlar"]
                improvement = student["diff"]
                text += f"{i}. <b>{name}</b>: +{improvement:.1f} ball\n"

        await callback.message.edit_text(text, parse_mode="HTML")

    elif action == "declined":
        from database import get_most_declined_students, get_most_declined_students_by_maktab
        
        data = await state.get_data()
        maktab_id = data.get("selected_maktab_id")
        maktab_nomi = data.get("selected_maktab_nomi", "Noma'lum maktab")
        
        if maktab_id:
            students = get_most_declined_students_by_maktab(maktab_id)
            title = f"📉 <b>{maktab_nomi} - Eng ko'p pasaygan o'quvchilar:</b>\n\n"
        else:
            students = get_most_declined_students()
            title = "📉 <b>Eng ko'p pasaygan o'quvchilar:</b>\n\n"

        if not students:
            text = title + "⚠️ Ma'lumotlar yo'q."
        else:
            text = title + "(oxirgi test avvalgisiga nisbatan)\n\n"
            for i, student in enumerate(students[:10], 1):
                name = student["ismlar"]
                change = student["diff"]
                text += f"{i}. <b>{name}</b>: {change:.1f} ball\n"

        await callback.message.edit_text(text, parse_mode="HTML")

    await callback.answer()


@router.message(F.text == "⚙️ Sozlamalar")
async def settings_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    ranking_enabled = get_setting("ranking_enabled", "True")
    stats_enabled = get_setting("stats_enabled", "True")
    chatbot_enabled = get_setting("chatbot_enabled", "True")
    mock_enabled = get_setting("mock_enabled", "True")
    quiz_enabled = get_setting("quiz_enabled", "True")
    mini_test_enabled = get_setting("mini_test_enabled", "True")

    text = (
        "⚙️ <b>Bot sozlamalari:</b>\n\n"
        f"🏆 Reyting (Top-50): <b>{'✅ Yoqilgan' if ranking_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"📈 Statistika: <b>{'✅ Yoqilgan' if stats_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"🤖 AI Chatbot: <b>{'✅ Yoqilgan' if chatbot_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"🧪 Mock natijalarim (User menyu): <b>{'✅ Yoqilgan' if mock_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"📝 Mashq qilish (Quiz): <b>{'✅ Yoqilgan' if quiz_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"📦 Mini-testlar: <b>{'✅ Yoqilgan' if mini_test_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n\n"
        "O'zgartirish uchun tugmalarni bosing:"
    )
    await message.answer(
        text,
        parse_mode="HTML",
        reply_markup=settings_keyboard(ranking_enabled, stats_enabled, chatbot_enabled, mock_enabled, quiz_enabled, mini_test_enabled),
    )


@router.callback_query(F.data.startswith("toggle_setting:"))
async def toggle_setting_handler(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    key = callback.data.split(":")[1]
    current = get_setting(key, "True")
    new_value = "False" if current == "True" else "True"
    set_setting(key, new_value)

    ranking_enabled = get_setting("ranking_enabled", "True")
    stats_enabled = get_setting("stats_enabled", "True")
    chatbot_enabled = get_setting("chatbot_enabled", "True")
    mock_enabled = get_setting("mock_enabled", "True")
    quiz_enabled = get_setting("quiz_enabled", "True")
    mini_test_enabled = get_setting("mini_test_enabled", "True")

    text = (
        "⚙️ <b>Bot sozlamalari:</b>\n\n"
        f"🏆 Reyting (Top-50): <b>{'✅ Yoqilgan' if ranking_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"📈 Statistika: <b>{'✅ Yoqilgan' if stats_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"🤖 AI Chatbot: <b>{'✅ Yoqilgan' if chatbot_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"🧪 Mock natijalarim (User menyu): <b>{'✅ Yoqilgan' if mock_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"📝 Mashq qilish (Quiz): <b>{'✅ Yoqilgan' if quiz_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n"
        f"📦 Mini-testlar: <b>{'✅ Yoqilgan' if mini_test_enabled == 'True' else '❌ O' + chr(39) + 'chirilgan'}</b>\n\n"
        "O'zgartirish uchun tugmalarni bosing:"
    )
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=settings_keyboard(ranking_enabled, stats_enabled, chatbot_enabled, mock_enabled, quiz_enabled, mini_test_enabled),
    )
    await callback.answer("✅ Sozlama yangilandi")


@router.callback_query(F.data == "open_cert_sozlama")
async def open_cert_sozlama_cb(callback: CallbackQuery, state: FSMContext):
    """Sozlamalar menyusidagi '📜 Sertifikat sozlamalari' tugmasi."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    # cert_admin.py'dagi menyu funksiyasini chaqiramiz
    from cert_admin import _get_all_settings, _menu_text, _menu_keyboard
    s = _get_all_settings()
    await callback.message.edit_text(
        _menu_text(s), parse_mode="HTML", reply_markup=_menu_keyboard()
    )
    await callback.answer()


@router.message(F.text == "🔔 So'rovlar")
async def requests_list(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    requests = get_pending_requests()
    if not requests:
        await message.answer("✅ Hozircha yangi so'rovlar yo'q.")
        return

    for req in requests:
        text = (
            f"🔔 <b>Yangi kirish so'rovi:</b>\n\n"
            f"👤 Ism: <b>{req['ismlar']}</b>\n"
            f"🏫 Sinf: <b>{req['sinf']}</b>\n"
            f"🎯 Yo'nalish: <b>{req.get('yonalish', 'Noma' + chr(39) + 'lum')}</b>\n"
            f"🆔 Kod: <code>{req['talaba_kod']}</code>\n"
            f"👤 Telegram ID: <code>{req['user_id']}</code>"
        )
        await message.answer(
            text,
            parse_mode="HTML",
            reply_markup=request_actions_keyboard(req["id"], req["user_id"]),
        )


@router.callback_query(F.data.startswith("request_action:"))
async def request_action_handler(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    data_parts = callback.data.split(":")
    action = data_parts[1]
    target_id = int(data_parts[2])  # request_id yoki user_id bo'lishi mumkin

    from database import (
        get_connection,
        release_connection,
        revoke_access_by_user,
    )
    from datetime import datetime, timedelta, timezone

    if action == "revoke":
        revoke_access_by_user(target_id)
        await callback.answer("🚫 Ruxsat qaytarib olindi")
        await callback.message.edit_text(
            f"🚫 Foydalanuvchining (ID: {target_id}) barcha ruxsatlari bekor qilindi."
        )
        return

    request_id = target_id
    if action.startswith("approve"):
        expires_at = None
        if action == "approve_5m":
            expires_at = datetime.now(timezone.utc) + timedelta(minutes=5)
        elif action == "approve_30m":
            expires_at = datetime.now(timezone.utc) + timedelta(minutes=30)
        elif action == "approve_1h":
            expires_at = datetime.now(timezone.utc) + timedelta(hours=1)

        update_request_status(request_id, "approved", expires_at)
        # User_id ni topish
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT user_id FROM access_requests WHERE id = %s", (request_id,)
        )
        user_id = cur.fetchone()[0]
        cur.close()
        release_connection(conn)

        try:
            await callback.bot.send_message(
                user_id,
                "✅ Sizning botdan foydalanish so'rovingiz admin tomonidan tasdiqlandi! Endi natijalaringizni ko'rishingiz mumkin.",
            )
        except Exception:
            pass
        await callback.message.edit_text("✅ So'rov tasdiqlandi.")

    elif action == "reject":
        update_request_status(request_id, "rad_etildi")
        await callback.message.edit_text("❌ So'rov rad etildi.")

    await callback.answer()


# ─────────────────────────────────────────
# O'quvchilar ro'yxati va Excelga yuklash
# ─────────────────────────────────────────


@router.message(F.text == "📋 O'quvchilar ro'yxati")
async def students_list(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    
    # Get all schools for selection
    maktablar = maktablar_ol()
    if not maktablar:
        await message.answer("⚠️ Hozircha maktablar mavjud emas.")
        return
    
    await message.answer(
        "📋 <b>O'quvchilar ro'yxati uchun maktabni tanlang:</b>\n\n"
        "Qaysi maktabning o'quvchilar ro'yxatini ko'rmoqchisiz?",
        parse_mode="HTML",
        reply_markup=maktab_tanlash_keyboard(maktablar, "students_maktab")
    )


@router.callback_query(F.data.startswith("students_maktab:"))
async def students_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """School selection for students list callback handler."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    try:
        maktab_id = int(callback.data.split(":")[1])
        maktablar = maktablar_ol()
        maktab = next((m for m in maktablar if m['id'] == maktab_id), None)
        
        if not maktab:
            await callback.answer("❌ Maktab topilmadi", show_alert=True)
            return
        
        # Store selected school in state
        await state.update_data(students_selected_maktab_id=maktab_id, students_selected_maktab_nomi=maktab['nomi'])
        
        # Show filtering options for this school
        await callback.message.edit_text(
            f"📋 <b>{maktab['nomi']} - O'quvchilar ro'yxatini ko'rish usulini tanlang:</b>",
            reply_markup=oquvchilar_filtrlash_keyboard()
        )
        
    except (ValueError, IndexError):
        await callback.answer("❌ Noto'g'ri maktab tanlovi", show_alert=True)
    
    await callback.answer()


@router.callback_query(F.data.startswith("filter_type:"))
async def filter_type_process(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    # Get selected school from state
    data = await state.get_data()
    maktab_id = data.get("students_selected_maktab_id")
    maktab_nomi = data.get("students_selected_maktab_nomi", "Noma'lum maktab")
    
    ftype = callback.data.split(":")[1]
    if ftype in ("all", "hammasi"):
        if maktab_id:
            talabalar = talaba_hammasi_by_maktab(maktab_id)
            title = f"📋 <b>{maktab_nomi} - Barcha o'quvchilar</b>"
        else:
            talabalar = talaba_hammasi()
            title = "📋 <b>Barcha o'quvchilar</b>"
            
        if not talabalar:
            await callback.answer("⚠️ O'quvchilar yo'q.", show_alert=True)
            return
        text, page, total_pages = _build_students_page(
            talabalar, title, 1
        )
        await safe_edit(
            callback,
            text,
            parse_mode="HTML",
            reply_markup=filter_actions_keyboard(
                "all", "all", page, total_pages
            ),
        )
    elif ftype == "sinf":
        await safe_edit(
            callback,
            "🏫 Sinfni tanlang:",
            reply_markup=sinf_tanlash_keyboard(),
        )
    elif ftype == "yonalish":
        await safe_edit(
            callback,
            "🎯 Yo'nalishni tanlang:",
            reply_markup=yonalish_tanlash_keyboard(),
        )
    elif ftype in ("back", "orqaga"):
        await safe_edit(
            callback,
            "📋 O'quvchilar ro'yxatini ko'rish usulini tanlang:",
            reply_markup=oquvchilar_filtrlash_keyboard(),
        )
    await callback.answer()


@router.callback_query(F.data.startswith("filter_val:"))
async def filter_val_process(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    # Get selected school from state
    data = await state.get_data()
    maktab_id = data.get("students_selected_maktab_id")
    maktab_nomi = data.get("students_selected_maktab_nomi", "Noma'lum maktab")
    
    parts = callback.data.split(":")
    ftype = parts[1]
    fval = parts[2]

    if ftype == "sinf":
        if maktab_id:
            talabalar = talaba_filtrlangan_by_maktab(maktab_id, sinf=fval)
            title = f"📋 <b>{maktab_nomi} - {fval} sinf o'quvchilari</b>"
        else:
            talabalar = talaba_filtrlangan(sinf=fval)
            title = f"📋 <b>{fval} sinf o'quvchilari</b>"
    else:
        if maktab_id:
            talabalar = talaba_filtrlangan_by_maktab(maktab_id, yonalish=fval)
            title = f"📋 <b>{maktab_nomi} - {fval} yo'nalishi o'quvchilari</b>"
        else:
            talabalar = talaba_filtrlangan(yonalish=fval)
            title = f"📋 <b>{fval} yo'nalishi o'quvchilari</b>"

    if not talabalar:
        await callback.answer(
            "⚠️ Bu filtr bo'yicha o'quvchilar topilmadi.", show_alert=True
        )
        return

    text, page, total_pages = _build_students_page(
        talabalar, title, 1
    )
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=filter_actions_keyboard(ftype, fval, page, total_pages),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("filter_page:"))
async def filter_page_process(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    parts = callback.data.split(":")
    ftype, fval, page_raw = parts[1], parts[2], parts[3]
    if ftype == "noop":
        await callback.answer()
        return

    try:
        page = int(page_raw)
    except ValueError:
        page = 1

    if ftype == "all":
        talabalar = talaba_hammasi()
        title = "📋 <b>Barcha o'quvchilar</b>"
    elif ftype == "sinf":
        talabalar = talaba_filtrlangan(sinf=fval)
        title = "📋 <b>Filtrlangan ro'yxat</b>"
    else:
        talabalar = talaba_filtrlangan(yonalish=fval)
        title = "📋 <b>Filtrlangan ro'yxat</b>"

    if not talabalar:
        await callback.answer("⚠️ Ma'lumot topilmadi.", show_alert=True)
        return

    text, page, total_pages = _build_students_page(talabalar, title, page)
    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=filter_actions_keyboard(ftype, fval, page, total_pages),
    )
    await callback.answer()


@router.callback_query(F.data == "talaba_ochir_start")
async def talaba_ochir_start(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    await state.set_state(TalabaOchir.kod_kutish)
    await callback.message.answer(
        "🗑 O'chirmoqchi bo'lgan o'quvchingizning 🆔 kodini kiriting:"
    )
    await callback.answer()


@router.message(TalabaOchir.kod_kutish)
async def talaba_ochir_kod(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)
    if not talaba:
        await message.answer(
            f"❌ Kod {kod} bo'yicha o'quvchi topilmadi. Qaytadan kiriting yoki /cancel deb yozing:"
        )
        return

    await state.update_data(ochirish_kod=kod)
    await state.set_state(TalabaOchir.tasdiq_kutish)

    text = (
        f"❓ Haqiqatan ham ushbu o'quvchini o'chirmoqchimisiz?\n\n"
        f"👤 Ism: <b>{talaba['ismlar']}</b>\n"
        f"🆔 Kod: <b>{talaba['kod']}</b>\n"
        f"🏫 Sinf: <b>{talaba['sinf']}</b>\n"
        f"🎯 Yo'nalish: <b>{talaba['yonalish']}</b>\n\n"
        f"⚠️ Barcha natijalari ham o'chib ketadi!"
    )

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✅ HA, o'chirilsin",
                    callback_data="talaba_ochir_tasdiq:ha",
                ),
                InlineKeyboardButton(
                    text="❌ YO'Q, bekor qilinsin",
                    callback_data="talaba_ochir_tasdiq:yoq",
                ),
            ]
        ]
    )
    await message.answer(text, parse_mode="HTML", reply_markup=keyboard)


@router.callback_query(F.data.startswith("talaba_ochir_tasdiq:"))
async def talaba_ochir_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "ha":
        data = await state.get_data()
        kod = data.get("ochirish_kod")
        if talaba_ochir(kod):
            await callback.message.edit_text(
                f"✅ O'quvchi (Kod: {kod}) muvaffaqiyatli o'chirildi."
            )
        else:
            await callback.message.edit_text(
                "❌ O'chirishda xatolik yuz berdi."
            )
    else:
        await callback.message.edit_text("❌ O'chirish bekor qilindi.")

    await state.clear()
    await callback.message.answer(
        "Asosiy menyu:", reply_markup=admin_menu_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("filter_excel:"))
async def filter_excel_process(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    parts = callback.data.split(":")
    ftype = parts[1]
    fval = parts[2]

    if ftype == "all":
        data = get_all_students_for_excel()
        filename_suffix = "barchasi"
    elif ftype == "sinf":
        data = talaba_filtrlangan(sinf=fval)
        filename_suffix = f"sinf_{fval}"
    else:
        data = talaba_filtrlangan(yonalish=fval)
        filename_suffix = f"yonalish_{fval}"

    if not data:
        await callback.answer(
            "⚠️ Yuklash uchun ma'lumot topilmadi.", show_alert=True
        )
        return

    path = f"filtr_{filename_suffix}_{callback.from_user.id}.xlsx"
    pd.DataFrame(data).to_excel(path, index=False)
    try:
        await callback.message.answer_document(
            FSInputFile(path), caption="📊 Filtrlangan ro'yxat (Excel)"
        )
    finally:
        if os.path.exists(path):
            os.remove(path)
    await callback.answer("✅ Excel yuborildi")


@router.callback_query(F.data == "admin_menu")
async def back_to_admin_menu(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.message.answer(
        "Asosiy menyu:", reply_markup=admin_menu_keyboard()
    )
    await callback.answer()


@router.message(F.text == "📥 Excelga yuklash")
async def export_excel(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = get_all_students_for_excel()
    if not data:
        await message.answer("⚠️ Ma'lumotlar yo'q.")
        return

    path = f"oquvchilar_{message.from_user.id}.xlsx"
    df = pd.DataFrame(data)
    df.to_excel(path, index=False)

    await message.answer_document(
        FSInputFile(path),
        caption="📊 Barcha o'quvchilar va natijalar ro'yxati (Excel)",
    )
    if os.path.exists(path):
        os.remove(path)


@router.message(F.text == "🏆 Reyting Excel")
async def export_ranking_excel(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    
    # Get all schools for selection
    maktablar = maktablar_ol()
    if not maktablar:
        await message.answer("⚠️ Hozircha maktablar mavjud emas.")
        return
    
    await message.answer(
        "🏆 <b>Reyting Excel uchun maktabni tanlang:</b>\n\n"
        "Qaysi maktabning reytingini Excelga yuklamoqchisiz?",
        parse_mode="HTML",
        reply_markup=maktab_tanlash_keyboard(maktablar, "ranking_maktab")
    )


@router.callback_query(F.data.startswith("ranking_maktab:"))
async def ranking_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """School selection for ranking Excel callback handler."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    try:
        maktab_id = int(callback.data.split(":")[1])
        maktablar = maktablar_ol()
        maktab = next((m for m in maktablar if m['id'] == maktab_id), None)
        
        if not maktab:
            await callback.answer("❌ Maktab topilmadi", show_alert=True)
            return
        
        # Get school-specific data for Excel
        from database import get_sorted_students_for_excel_by_maktab
        data = get_sorted_students_for_excel_by_maktab(maktab_id)
        
        if not data:
            await callback.answer("⚠️ Bu maktabda ma'lumotlar yo'q (Natijalar hali mavjud emas).", show_alert=True)
            return

        path = f"reyting_{callback.from_user.id}_{maktab_id}.xlsx"
        df = pd.DataFrame(data)
        # Ustunlarni chiroyli qilish
        df.columns = [
            "Kod",
            "Sinf",
            "Yo'nalish",
            "Ism-sharif",
            "Majburiy",
            "1-Asosiy",
            "2-Asosiy",
            "Umumiy ball",
            "Sana",
        ]
        df.to_excel(path, index=False)

        await callback.message.answer_document(
            FSInputFile(path),
            caption=f"🏆 {maktab['nomi']} - Ballar bo'yicha saralangan reyting ro'yxati (Excel)",
        )

        # Faylni o'chirib tashlash
        os.remove(path)
        
    except (ValueError, IndexError):
        await callback.answer("❌ Noto'g'ri maktab tanlovi", show_alert=True)
    
    await callback.answer()


@router.message(F.text == "👨‍🏫 O'qituvchilarni boshqarish")
async def oqituvchi_menu(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "👨‍🏫 O'qituvchilarni boshqarish bo'limi:",
        reply_markup=oqituvchi_boshqarish_keyboard(),
    )


@router.callback_query(F.data.startswith("oqituvchi_boshqar:"))
async def oqituvchi_boshqar_actions(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    action = callback.data.split(":")[1]

    if action == "ro'yxat":
        oqituvchilar = oqituvchilar_hammasi()
        if not oqituvchilar:
            await callback.answer(
                "⚠️ O'qituvchilar hali qo'shilmagan.", show_alert=True
            )
            return
        text = "📋 <b>O'qituvchilar ro'yxati:</b>\n\n"
        for o in oqituvchilar:
            text += f"👤 {o['ismlar']} | 🏫 {o['sinf']} | ID: <code>{o['user_id']}</code>\n"
        await callback.message.edit_text(
            text,
            parse_mode="HTML",
            reply_markup=oqituvchi_boshqarish_keyboard(),
        )

    elif action == "qosh":
        await state.set_state(OqituvchiQosh.user_id_kutish)
        await callback.message.answer(
            "🆔 O'qituvchining Telegram ID raqamini kiriting:"
        )
        await callback.answer()

    elif action == "ochir":
        await callback.message.edit_text(
            "❌ O'chirmoqchi bo'lgan o'qituvchini tanlang:",
            reply_markup=oqituvchi_ochirish_keyboard(),
        )

    elif action == "orqaga":
        await callback.message.delete()
    await callback.answer()


@router.callback_query(F.data.startswith("oqituvchi_ochir:"))
async def oqituvchi_ochir_process(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    user_id = int(callback.data.split(":")[1])
    if oqituvchi_ochir(user_id):
        await callback.answer("✅ O'qituvchi o'chirildi")
    else:
        await callback.answer("❌ Xatolik yuz berdi")
    await callback.message.edit_text(
        "❌ O'chirmoqchi bo'lgan o'qituvchini tanlang:",
        reply_markup=oqituvchi_ochirish_keyboard(),
    )


@router.message(OqituvchiQosh.user_id_kutish)
async def oqituvchi_id_get(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    try:
        user_id = int(message.text)
        await state.update_data(teacher_id=user_id)
        await state.set_state(OqituvchiQosh.ismlar_kutish)
        await message.answer("👤 O'qituvchining ism-familiyasini kiriting:")
    except ValueError:
        await message.answer("❌ Faqat raqam kiriting!")


@router.message(OqituvchiQosh.ismlar_kutish)
async def oqituvchi_ism_get(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.update_data(teacher_name=message.text)
    await state.set_state(OqituvchiQosh.sinf_kutish)
    await message.answer(
        "🏫 O'qituvchi mas'ul bo'lgan sinfni kiriting (masalan: 11-A):",
        reply_markup=sinf_keyboard(),
    )


@router.callback_query(
    F.data.startswith("sinf_id:"), OqituvchiQosh.sinf_kutish
)
async def oqituvchi_sinf_get(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    sinf_id = int(callback.data.split(":")[1])
    sinf = sinf_id_dan_full_nomi(sinf_id)
    if not sinf:
        await callback.answer("❌ Sinf topilmadi.", show_alert=True)
        return
    data = await state.get_data()
    if oqituvchi_qosh(data["teacher_id"], data["teacher_name"], sinf):
        await callback.message.answer(
            f"✅ O'qituvchi {data['teacher_name']} ({sinf}) muvaffaqiyatli qo'shildi!"
        )
    else:
        await callback.message.answer("❌ Xatolik yuz berdi.")
    await state.clear()
    await state.update_data(admin=True)
    await callback.answer()


# O'qituvchi uchun handlerlar
@router.message(F.text == "📊 Mening sinfim statistikasi")
async def teacher_stats(message: Message):
    teacher = oqituvchi_ol(message.from_user.id)
    if not teacher:
        return
    talabalar = talaba_filtrlangan(sinf=teacher["sinf"])
    if not talabalar:
        await message.answer("⚠️ Sinfingizda o'quvchilar yo'q.")
        return

    ballar = [
        t["umumiy_ball"] for t in talabalar if t["umumiy_ball"] is not None
    ]
    avg = round(sum(ballar) / len(ballar), 2) if ballar else 0
    text = (
        f"🏫 <b>{teacher['sinf']} sinf statistikasi:</b>\n\n"
        f"👥 O'quvchilar soni: <b>{len(talabalar)} ta</b>\n"
        f"📈 O'rtacha ball: <b>{avg} ball</b>\n"
        f"🏆 Eng yuqori ball: <b>{max(ballar) if ballar else 0} ball</b>"
    )
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "🏆 Mening sinfim reytingi")
async def teacher_ranking(message: Message):
    teacher = oqituvchi_ol(message.from_user.id)
    if not teacher:
        return
    talabalar = get_all_in_class(teacher["sinf"])
    if not talabalar:
        await message.answer("⚠️ Natijalar yo'q.")
        return
    text = f"🏆 <b>{teacher['sinf']} sinf reytingi:</b>\n\n"
    for i, t in enumerate(talabalar, 1):
        text += f"{i}. {t['ismlar']} — <b>{t['umumiy_ball']}</b>\n"
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "📋 Sinfim o'quvchilari")
async def teacher_students(message: Message):
    teacher = oqituvchi_ol(message.from_user.id)
    if not teacher:
        return
    talabalar = talaba_filtrlangan(sinf=teacher["sinf"])
    text = f"📋 <b>{teacher['sinf']} sinf o'quvchilari:</b>\n\n"
    for i, t in enumerate(talabalar, 1):
        text += f"{i}. {t['kod']} | {t['ismlar']}\n"
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "📥 Sinfim natijalari (Excel)")
async def teacher_excel(message: Message):
    teacher = oqituvchi_ol(message.from_user.id)
    if not teacher:
        return
    data = talaba_filtrlangan(sinf=teacher["sinf"])
    if not data:
        await message.answer("⚠️ Ma'lumotlar yo'q.")
        return
    path = f"sinf_{teacher['sinf']}.xlsx"
    pd.DataFrame(data).to_excel(path, index=False)
    await message.answer_document(
        FSInputFile(path), caption=f"📊 {teacher['sinf']} sinf natijalari"
    )
    if os.path.exists(path):
        os.remove(path)


# ─────────────────────────────────────────
# Reyting va Xabar yuborish
# ─────────────────────────────────────────


@router.message(F.text == "🏆 Reyting")
async def ranking_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "🏆 Reyting bo'limi:", reply_markup=ranking_keyboard(is_admin=True)
    )





@router.callback_query(
    F.data.startswith("ranking:"), F.from_user.id.in_(ADMIN_IDS)
)
async def ranking_admin_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    action = callback.data.split(":")[1]
    if action == "select_class":
        await callback.message.edit_text(
            "🏫 Reytingni ko'rish uchun sinfni tanlang:",
            reply_markup=sinf_tanlash_ranking_keyboard(),
        )
    elif action == "top_by_classes":
        await callback.message.edit_text(
            "📊 Qaysi parallel sinflar bo'yicha ko'rmoqchisiz?",
            reply_markup=sinf_prefix_tanlash_keyboard(),
        )
    elif action == "view_class":
        sinf = callback.data.split(":", 2)[2]
        talabalar = get_all_in_class(sinf)
        if not talabalar:
            await callback.answer(
                "⚠️ Bu sinfda natijalar yo'q.", show_alert=True
            )
            return
        text = f"🏆 <b>{sinf} sinf reytingi:</b>\n\n"
        for i, t in enumerate(talabalar, 1):
            text += f"{i}. {t['ismlar']} — <b>{t['umumiy_ball']}</b>\n"
        await callback.message.edit_text(
            text,
            parse_mode="HTML",
            reply_markup=sinf_tanlash_ranking_keyboard(),
        )
    elif action == "top_prefix":
        sinf_prefix = callback.data.split(":", 2)[2]
        top50 = get_overall_ranking(sinf_prefix=sinf_prefix)
        if not top50:
            await callback.answer(
                "⚠️ Hozircha natijalar yo'q.", show_alert=True
            )
            return
        text = f"🌍 <b>{sinf_prefix}-sinflar orasidagi Top 50:</b>\n\n"
        for i, t in enumerate(top50, 1):
            emoji = (
                "🥇"
                if i == 1
                else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            )
            text += f"{emoji} {t['ismlar']} ({t['sinf']}) — <b>{t['umumiy_ball']}</b> ball\n"
        await callback.message.edit_text(
            text,
            parse_mode="HTML",
            reply_markup=sinf_prefix_tanlash_keyboard(),
        )
    elif action == "back_admin":
        await callback.message.edit_text(
            "🏆 Reyting bo'limi:", reply_markup=ranking_keyboard(is_admin=True)
        )

    await callback.answer()


@router.message(F.text == "📢 Xabar yuborish")
async def broadcast_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.set_state(Broadcast.maktab_tanlash)
    await message.answer(
        "📢 <b>Xabar yuborish</b>\n\nKimga yubormoqchisiz?",
        parse_mode="HTML",
        reply_markup=broadcast_target_keyboard(),
    )


@router.callback_query(Broadcast.maktab_tanlash, F.data == "bcast_target:cancel")
async def broadcast_target_cancel(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    await state.clear()
    await callback.message.edit_text("❌ Xabar yuborish bekor qilindi.")
    await callback.message.answer("🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard())
    await callback.answer()


@router.callback_query(Broadcast.maktab_tanlash, F.data == "bcast_target:all")
async def broadcast_target_all(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    await state.update_data(broadcast_target="all", broadcast_maktab_id=None)
    await state.set_state(Broadcast.xabar_kutish)
    await callback.message.edit_text("✅ Barcha foydalanuvchilar tanlandi.")
    await callback.message.answer(
        "📝 Yubormoqchi bo'lgan xabaringizni yozing:",
        reply_markup=broadcast_cancel_keyboard(),
    )
    await callback.answer()


@router.callback_query(Broadcast.maktab_tanlash, F.data == "bcast_target:school")
async def broadcast_target_school(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    maktablar = maktablar_ol()
    if not maktablar:
        await callback.answer("❌ Hozircha maktablar yo'q.", show_alert=True)
        return
    await callback.message.edit_text(
        "🏫 Qaysi maktabga yubormoqchisiz?",
        reply_markup=broadcast_maktab_tanlash_keyboard(maktablar),
    )
    await callback.answer()


@router.callback_query(Broadcast.maktab_tanlash, F.data.startswith("bcast_school:"))
async def broadcast_school_selected(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        await callback.answer()
        return
    # split faqat birinchi ":" dan keyin — maktab nomida ":" bo'lmasligi uchun
    parts = callback.data.split(":", 1)
    raw = parts[1] if len(parts) > 1 else ""
    if raw == "cancel":
        await state.clear()
        try:
            await callback.message.edit_text("❌ Xabar yuborish bekor qilindi.")
        except Exception:
            pass
        await callback.message.answer("🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard())
        await callback.answer()
        return
    try:
        maktab_id = int(raw)
    except ValueError:
        await callback.answer("❌ Noto'g'ri tanlov.", show_alert=True)
        return
    maktablar = maktablar_ol()
    maktab_nomi = next((m["nomi"] for m in maktablar if m["id"] == maktab_id), "Noma'lum")
    # Nechta foydalanuvchi borligi
    from database import count_users_by_maktab
    user_count = count_users_by_maktab(maktab_id)
    await state.update_data(broadcast_target="school", broadcast_maktab_id=maktab_id, broadcast_maktab_nomi=maktab_nomi)
    await state.set_state(Broadcast.xabar_kutish)
    # Inline klaviaturani o'chirish va tasdiqlash
    try:
        await callback.message.edit_text(
            f"✅ <b>{maktab_nomi}</b> tanlandi.\n"
            f"👥 Botga ulangan o'quvchilar: <b>{user_count} ta</b>",
            parse_mode="HTML",
        )
    except Exception:
        pass
    if user_count == 0:
        await callback.message.answer(
            f"⚠️ <b>{maktab_nomi}</b> maktabidan hali hech kim botga ulanmagan.\n"
            "Boshqa maktab tanlang yoki barcha foydalanuvchilarga yuboring.",
            parse_mode="HTML",
            reply_markup=admin_menu_keyboard(),
        )
        await state.clear()
        await callback.answer()
        return
    await callback.message.answer(
        f"📝 <b>{maktab_nomi}</b> ga yubormoqchi bo'lgan xabaringizni yozing:",
        parse_mode="HTML",
        reply_markup=broadcast_cancel_keyboard(),
    )
    await callback.answer()


@router.message(
    Broadcast.xabar_kutish, F.text == "❌ Xabar yuborishni bekor qilish"
)
async def broadcast_cancel_during_input(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.clear()
    await message.answer(
        "❌ Xabar yuborish bekor qilindi.", reply_markup=admin_menu_keyboard()
    )


@router.message(Broadcast.xabar_kutish)
async def broadcast_process(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    preview = (
        message.text
        or message.caption
        or f"{message.content_type} formatdagi xabar"
    )
    await state.update_data(
        broadcast_source_chat_id=message.chat.id,
        broadcast_source_message_id=message.message_id,
    )
    await state.set_state(Broadcast.tasdiq_kutish)
    await message.answer(
        f"📨 Xabar tayyor:\n\n{preview}\n\nJo'natishni tasdiqlaysizmi?",
        reply_markup=broadcast_confirm_keyboard(),
    )


@router.callback_query(Broadcast.tasdiq_kutish, F.data == "broadcast:cancel")
async def broadcast_cancel_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    await state.clear()
    await callback.message.edit_text("❌ Xabar yuborish bekor qilindi.")
    await callback.message.answer(
        "🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard()
    )
    await callback.answer()


@router.callback_query(Broadcast.tasdiq_kutish, F.data == "broadcast:confirm")
async def broadcast_confirm_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    data = await state.get_data()
    src_chat_id = data.get("broadcast_source_chat_id")
    src_message_id = data.get("broadcast_source_message_id")
    broadcast_target = data.get("broadcast_target", "all")
    broadcast_maktab_id = data.get("broadcast_maktab_id")

    if not src_chat_id or not src_message_id:
        await state.clear()
        await callback.message.edit_text(
            "❌ Xabar topilmadi. Qaytadan urinib ko'ring."
        )
        await callback.message.answer(
            "🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )
        await callback.answer()
        return

    if broadcast_target == "school" and broadcast_maktab_id:
        user_ids = get_user_ids_by_maktab(broadcast_maktab_id)
        maktab_nomi = data.get("broadcast_maktab_nomi", "Tanlangan maktab")
    else:
        user_ids = get_all_user_ids()
        maktab_nomi = "Barcha foydalanuvchilar"
    total = len(user_ids)

    if total == 0:
        await state.clear()
        await callback.message.edit_text(
            f"⚠️ <b>{maktab_nomi}</b> uchun botga ulangan foydalanuvchi topilmadi.",
            parse_mode="HTML",
        )
        await callback.message.answer("🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard())
        await callback.answer()
        return

    sent_count = 0
    fail_count = 0
    processed = 0
    progress_step = 25

    await callback.message.edit_text(
        f"📤 <b>{maktab_nomi}</b> ga xabar yuborish boshlandi...\n"
        f"0/{total} ta foydalanuvchi ko'rib chiqildi.",
        parse_mode="HTML",
    )

    for uid in user_ids:
        if not is_notification_enabled(uid, "notify_admin_messages"):
            processed += 1
            continue
        try:
            await callback.bot.copy_message(
                chat_id=uid,
                from_chat_id=src_chat_id,
                message_id=src_message_id,
            )
            sent_count += 1
            await asyncio.sleep(0.05)
        except Exception:
            fail_count += 1
        finally:
            processed += 1

        if processed % progress_step == 0 or processed == total:
            await callback.message.edit_text(
                f"📤 <b>{maktab_nomi}</b> ga yuborilmoqda...\n"
                f"{processed}/{total} ta foydalanuvchi ko'rib chiqildi.\n"
                f"✅ Yuborildi: {sent_count} | ❌ Xatolik: {fail_count}",
                parse_mode="HTML",
            )

    await state.clear()
    await callback.message.edit_text(
        f"✅ Xabar yuborish yakunlandi.\n"
        f"🏫 Maqsad: <b>{maktab_nomi}</b>\n"
        f"📊 Jami: {total}\n"
        f"✅ Yuborildi: {sent_count}\n"
        f"❌ Xatolik: {fail_count}",
        parse_mode="HTML",
    )
    await callback.message.answer(
        "🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard()
    )
    await callback.answer()


# ─────────────────────────────────────────────
# ✉️ SHAXSIY XABAR YUBORISH
# ─────────────────────────────────────────────

@router.message(F.text == "✉️ Shaxsiy xabar yuborish")
async def shaxsiy_xabar_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.set_state(ShaxsiyXabar.kod_kutish)
    await message.answer(
        "✉️ <b>Shaxsiy xabar yuborish</b>\n\n"
        "O'quvchining shaxsiy kodini kiriting:\n"
        "<i>Masalan: A-007 yoki 52B</i>",
        parse_mode="HTML",
        reply_markup=broadcast_cancel_keyboard(),
    )


@router.message(ShaxsiyXabar.kod_kutish, F.text == "❌ Xabar yuborishni bekor qilish")
async def shaxsiy_xabar_cancel_kod(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_menu_keyboard())


@router.message(ShaxsiyXabar.kod_kutish)
async def shaxsiy_xabar_kod_qabul(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)
    if not talaba:
        await message.answer(
            f"❌ <b>{kod}</b> kodi bo'yicha o'quvchi topilmadi.\n"
            "Qaytadan kiriting yoki bekor qiling.",
            parse_mode="HTML",
        )
        return
    if not talaba.get("user_id"):
        await message.answer(
            f"⚠️ <b>{talaba['ismlar']}</b> ({kod}) hali botga ulanmagan.\n"
            "Unga xabar yuborib bo'lmaydi.",
            parse_mode="HTML",
        )
        return
    await state.update_data(pm_talaba_kod=kod, pm_talaba_user_id=talaba["user_id"], pm_talaba_ismi=talaba["ismlar"])
    await state.set_state(ShaxsiyXabar.xabar_kutish)
    await message.answer(
        f"✅ O'quvchi topildi:\n"
        f"👤 <b>{talaba['ismlar']}</b>\n"
        f"🏫 Sinf: {talaba.get('sinf', '—')}\n"
        f"🆔 Kod: <code>{kod}</code>\n\n"
        f"📝 Yubormoqchi bo'lgan xabaringizni yozing:",
        parse_mode="HTML",
        reply_markup=broadcast_cancel_keyboard(),
    )


@router.message(ShaxsiyXabar.xabar_kutish, F.text == "❌ Xabar yuborishni bekor qilish")
async def shaxsiy_xabar_cancel_xabar(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_menu_keyboard())


@router.message(ShaxsiyXabar.xabar_kutish)
async def shaxsiy_xabar_xabar_qabul(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = await state.get_data()
    ismi = data.get("pm_talaba_ismi", "")
    user_id = data.get("pm_talaba_user_id")
    preview = message.text or message.caption or f"{message.content_type} formatdagi xabar"
    await state.update_data(
        pm_src_chat_id=message.chat.id,
        pm_src_message_id=message.message_id,
    )
    await state.set_state(ShaxsiyXabar.tasdiq_kutish)
    await message.answer(
        f"📨 <b>{ismi}</b> ga yubormoqchi bo'lgan xabar:\n\n"
        f"{preview}\n\n"
        f"Tasdiqlaysizmi?",
        parse_mode="HTML",
        reply_markup=shaxsiy_xabar_confirm_keyboard(user_id),
    )


@router.callback_query(ShaxsiyXabar.tasdiq_kutish, F.data == "pmsend:cancel")
async def shaxsiy_xabar_tasdiq_cancel(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    await state.clear()
    await callback.message.edit_text("❌ Xabar yuborish bekor qilindi.")
    await callback.message.answer("🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard())
    await callback.answer()


@router.callback_query(ShaxsiyXabar.tasdiq_kutish, F.data.startswith("pmsend:confirm:"))
async def shaxsiy_xabar_tasdiq_confirm(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    data = await state.get_data()
    target_user_id = int(callback.data.split(":")[2])
    src_chat_id = data.get("pm_src_chat_id")
    src_message_id = data.get("pm_src_message_id")
    ismi = data.get("pm_talaba_ismi", "")
    if not is_notification_enabled(target_user_id, "notify_admin_messages"):
        await callback.message.edit_text(
            f"⚠️ <b>{ismi}</b> admin xabarlarini o'chirib qo'ygan.",
            parse_mode="HTML",
        )
        await state.clear()
        await callback.answer()
        return
    try:
        await callback.bot.copy_message(
            chat_id=target_user_id,
            from_chat_id=src_chat_id,
            message_id=src_message_id,
        )
        await callback.message.edit_text(
            f"✅ Xabar <b>{ismi}</b> ga muvaffaqiyatli yuborildi!",
            parse_mode="HTML",
        )
    except Exception as e:
        await callback.message.edit_text(
            f"❌ Xabar yuborishda xatolik: {e}"
        )
    await state.clear()
    await callback.message.answer("🏠 Asosiy menyu:", reply_markup=admin_menu_keyboard())
    await callback.answer()


@router.message(F.text == "⚖️ Apellyatsiyalar")
async def appeals_list_view(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    appeals = appeals_ol("kutilmoqda")
    if not appeals:
        await message.answer("✅ Hozircha yangi apellyatsiyalar yo'q.")
        return
    await message.answer(
        "⚖️ <b>Yangi apellyatsiyalar:</b>",
        parse_mode="HTML",
        reply_markup=appeals_keyboard(appeals),
    )


@router.callback_query(F.data.startswith("appeal_view:"))
async def appeal_view_handler(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    appeal_id = int(callback.data.split(":")[1])
    appeal = appeal_ol_id(appeal_id)
    if not appeal:
        await callback.answer("❌ Apellyatsiya topilmadi.", show_alert=True)
        return
    text = (
        f"⚖️ <b>Apellyatsiya</b>\n\n"
        f"🆔 ID: <code>{appeal['id']}</code>\n"
        f"👤 User ID: <code>{appeal['user_id']}</code>\n"
        f"🎓 Kod: <code>{appeal['talaba_kod']}</code>\n"
        f"📌 Status: <b>{appeal['status']}</b>\n\n"
        f"📝 Xabar:\n<i>{appeal['xabar']}</i>"
    )
    await callback.message.edit_text(
        text, parse_mode="HTML", reply_markup=appeal_action_keyboard(appeal_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("appeal_reply:"))
async def appeal_reply_start(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    appeal_id = int(callback.data.split(":")[1])
    await state.update_data(appeal_id=appeal_id)
    await state.set_state(AppealReply.javob_kutish)
    await callback.message.answer("✍️ Apellyatsiyaga javob matnini yuboring:")
    await callback.answer()


@router.message(AppealReply.javob_kutish)
async def appeal_reply_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = await state.get_data()
    appeal_id = data.get("appeal_id")
    if not appeal_id:
        await message.answer("❌ Apellyatsiya ID topilmadi.")
        await state.set_state(None)
        return
    appeal = appeal_ol_id(appeal_id)
    if not appeal:
        await message.answer("❌ Apellyatsiya topilmadi.")
        await state.set_state(None)
        return
    appeal_javob_ber(appeal_id, message.text)
    try:
        await message.bot.send_message(
            appeal["user_id"],
            f"📩 <b>Apellyatsiyangizga javob:</b>\n\n{message.html_text}",
            parse_mode="HTML",
        )
    except Exception:
        pass
    await message.answer(
        "✅ Apellyatsiyaga javob yuborildi.",
        reply_markup=admin_menu_keyboard(),
    )
    await state.set_state(None)


@router.callback_query(F.data == "appeal_list")
async def appeal_list_back(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    appeals = appeals_ol("kutilmoqda")
    if not appeals:
        await callback.message.edit_text(
            "✅ Hozircha yangi apellyatsiyalar yo'q."
        )
    else:
        await callback.message.edit_text(
            "⚖️ <b>Yangi apellyatsiyalar:</b>",
            parse_mode="HTML",
            reply_markup=appeals_keyboard(appeals),
        )
    await callback.answer()


# ─────────────────────────────────────────
# Bazani tozalash (MUAMMO TUZATILGAN JOY)
# ─────────────────────────────────────────


@router.message(F.text == "🧹 Bazani tozalash")
async def clear_db_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "⚠️ <b>DIQQAT!</b> Barcha o'quvchilar va natijalar o'chib ketadi. Tasdiqlaysizmi?",
        reply_markup=baza_tozalash_keyboard(),
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("baza_tozalash:"))
async def clear_db_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        await callback.answer("Siz admin emassiz!", show_alert=True)
        return

    action = callback.data.split(":")[1]

    if action == "ha":
        try:
            delete_all_data()
            await callback.message.edit_text(
                "✅ Barcha ma'lumotlar o'chirildi."
            )
        except Exception as e:
            await callback.message.answer(f"❌ Xatolik yuz berdi: {str(e)}")
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )
    else:
        try:
            await callback.message.edit_text("❌ Bekor qilindi.")
        except Exception:
            await callback.message.answer("❌ Bekor qilindi.")
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )

    await callback.answer()
    await state.set_state(None)


# ─────────────────────────────────────────
# Boshqa callback handlerlar
# ─────────────────────────────────────────


@router.callback_query(F.data.startswith("murojaat_javob:"))
async def murojaat_javob_start(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    user_id = callback.data.split(":")[1]
    await state.update_data(reply_to=user_id)
    await state.set_state(MurojaatJavob.javob_kutish)
    await callback.message.answer(
        f"📝 Foydalanuvchiga (ID: {user_id}) javobingizni yozing:"
    )
    await callback.answer()


@router.message(MurojaatJavob.javob_kutish)
async def murojaat_javob_send(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    data = await state.get_data()
    user_id = data.get("reply_to")
    try:
        await message.bot.send_message(
            user_id,
            f"📩 <b>Admin javobi:</b>\n\n{message.html_text}",
            parse_mode="HTML",
        )
        await message.answer("✅ Javob yuborildi.")
    except Exception as e:
        await message.answer(f"❌ Yuborishda xato: {e}")
    await state.set_state(None)


# ─────────────────────────────────────────
# Eslatma tizimi handlerlari
# ─────────────────────────────────────────


@router.message(F.text == "⏰ Eslatmalar")
async def reminder_menu(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "⏰ Eslatma tizimi bo'limi:",
        reply_markup=reminder_boshqarish_keyboard(),
    )


@router.callback_query(F.data == "reminder:qosh")
async def reminder_qosh_start(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    await state.set_state(ReminderAdd.xabar_kutish)
    await call.message.edit_text("📝 Eslatma xabarini kiriting:")
    await call.answer()


@router.callback_query(F.data == "reminder:ro'yxat")
async def reminder_list_view(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    reminders = kutilayotgan_reminders_ol()
    if not reminders:
        await call.answer("📋 Kutilayotgan eslatmalar yo'q.", show_alert=True)
        return
    await call.message.edit_text(
        "📋 <b>Kutilayotgan eslatmalar:</b>\n\nO'chirish uchun ustiga bosing:",
        parse_mode="HTML",
        reply_markup=reminder_list_keyboard(reminders),
    )
    await call.answer()


@router.callback_query(F.data.startswith("reminder_del:"))
async def reminder_delete_process(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    rid = int(call.data.split(":")[1])
    reminder_ochir(rid)
    await call.answer("✅ Eslatma o'chirildi")
    reminders = kutilayotgan_reminders_ol()
    if reminders:
        await call.message.edit_reply_markup(
            reply_markup=reminder_list_keyboard(reminders)
        )
    else:
        await call.message.edit_text(
            "⏰ Eslatma tizimi bo'limi:",
            reply_markup=reminder_boshqarish_keyboard(),
        )


@router.callback_query(F.data == "reminder:orqaga")
async def reminder_back(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    await call.message.edit_text(
        "⏰ Eslatma tizimi bo'limi:",
        reply_markup=reminder_boshqarish_keyboard(),
    )
    await call.answer()


@router.message(ReminderAdd.xabar_kutish)
async def reminder_xabar_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.update_data(reminder_xabar=message.text)
    await state.set_state(ReminderAdd.vaqt_kutish)
    await message.answer(
        "📅 Yuborish vaqtini kiriting (YYYY-MM-DD HH:MM formatida):\nMasalan: 2024-05-20 09:00"
    )


@router.message(ReminderAdd.vaqt_kutish)
async def reminder_vaqt_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    try:
        from datetime import datetime

        datetime.strptime(message.text, "%Y-%m-%d %H:%M")
        data = await state.get_data()
        reminder_qosh(data["reminder_xabar"], message.text)
        await state.clear()
        await state.update_data(admin=True)
        await message.answer(
            "✅ Eslatma muvaffaqiyatli saqlandi!",
            reply_markup=admin_menu_keyboard(),
        )
    except ValueError:
        await message.answer(
            "❌ Vaqt formati noto'g'ri. Qaytadan kiriting (YYYY-MM-DD HH:MM):"
        )


# ─────────────────────────────────────────
# Maktab va Guruh boshqaruvi
# ─────────────────────────────────────────


@router.message(F.text == "🏫 Maktablarni boshqarish")
async def maktab_menu(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "🏫 Maktablarni boshqarish bo'limi:",
        reply_markup=maktab_boshqarish_keyboard(),
    )


@router.callback_query(F.data == "maktab:qosh")
async def maktab_qosh_start(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    await state.set_state(MaktabAdd.nomi_kutish)
    await call.message.edit_text("🏫 Yangi maktab nomini kiriting:")
    await call.answer()


@router.callback_query(F.data == "maktab:ro'yxat")
async def maktab_list_view(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    maktablar = maktablar_ol()
    if not maktablar:
        await call.answer("⚠️ Maktablar ro'yxati bo'sh.", show_alert=True)
        return
    await call.message.edit_text(
        "🏫 <b>Maktablar ro'yxati:</b>",
        parse_mode="HTML",
        reply_markup=maktab_list_keyboard(maktablar),
    )
    await call.answer()


@router.callback_query(F.data == "maktab:orqaga")
async def maktab_back(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    await call.message.edit_text(
        "🏫 Maktablarni boshqarish bo'limi:",
        reply_markup=maktab_boshqarish_keyboard(),
    )
    await call.answer()


@router.callback_query(F.data.startswith("maktab_view:"))
async def maktab_view_detail(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    maktab_id = int(call.data.split(":")[1])
    maktablar = maktablar_ol()
    maktab = next((m for m in maktablar if m["id"] == maktab_id), None)
    if not maktab:
        await call.answer("❌ Maktab topilmadi.", show_alert=True)
        return
    sinflar = maktab_sinflari_ol(maktab_id)
    sinf_text = (
        "\n".join([f"• {s}" for s in sinflar])
        if sinflar
        else "Hali bog'lanmagan"
    )
    text = f"🏫 <b>{maktab['nomi']}</b>\n\n" f"📚 Sinflar:\n{sinf_text}"
    await call.message.edit_text(
        text, parse_mode="HTML", reply_markup=maktab_detail_keyboard(maktab_id)
    )
    await call.answer()


@router.callback_query(F.data.startswith("maktab_sinf_add:"))
async def maktab_sinf_add_start(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    maktab_id = int(call.data.split(":")[1])
    await state.update_data(maktab_id=maktab_id)
    await state.set_state(MaktabSinfAdd.sinf_nomi)
    await call.message.edit_text(
        "🏫 Qaysi sinfni bog'lamoqchisiz?", reply_markup=sinf_keyboard()
    )
    await call.answer()


@router.callback_query(MaktabSinfAdd.sinf_nomi, F.data.startswith("sinf_id:"))
async def maktab_sinf_add_save(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    sinf_id = int(call.data.split(":")[1])
    sinf_full_nomi = sinf_id_dan_full_nomi(sinf_id)
    data = await state.get_data()
    maktab_id = data.get("maktab_id")
    if not maktab_id:
        await call.answer("❌ Maktab ID topilmadi.", show_alert=True)
        return
    if sinf_full_nomi:
        sinf_maktabga_bogla(sinf_full_nomi, maktab_id)
    await state.set_state(None)
    maktablar = maktablar_ol()
    maktab = next((m for m in maktablar if m["id"] == maktab_id), None)
    sinflar = maktab_sinflari_ol(maktab_id)
    sinf_text = (
        "\n".join([f"• {s}" for s in sinflar])
        if sinflar
        else "Hali bog'lanmagan"
    )
    text = (
        f"🏫 <b>{maktab['nomi'] if maktab else 'Maktab'}</b>\n\n"
        f"📚 Sinflar:\n{sinf_text}"
    )
    await call.message.edit_text(
        text, parse_mode="HTML", reply_markup=maktab_detail_keyboard(maktab_id)
    )
    await call.answer("✅ Sinf maktabga bog'landi")


@router.callback_query(F.data.startswith("maktab_del:"))
async def maktab_delete(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    maktab_id = int(call.data.split(":")[1])
    muvaffaqiyat = maktab_ochir(maktab_id)
    if muvaffaqiyat:
        await call.answer("✅ Maktab o'chirildi", show_alert=True)
        await call.message.edit_text(
            "🏫 Maktablarni boshqarish bo'limi:",
            reply_markup=maktab_boshqarish_keyboard(),
        )
    else:
        await call.answer(
            "❌ Bu maktabda sinflar mavjud!\nAvval sinflarni o'chiring yoki boshqa maktabga ko'chiring.",
            show_alert=True,
        )


@router.message(MaktabAdd.nomi_kutish)
async def maktab_nomi_save(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    nomi = message.text.strip()
    from database import maktab_qosh

    if maktab_qosh(nomi):
        await message.answer(
            f"✅ Maktab qo'shildi: {nomi}", reply_markup=admin_menu_keyboard()
        )
    else:
        await message.answer("❌ Xato yuz berdi.")
    await state.clear()
    await state.update_data(admin=True)


@router.message(F.text == "📢 Guruhlarni boshqarish")
async def guruh_menu(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "📢 Guruhlarni boshqarish bo'limi:",
        reply_markup=guruh_boshqarish_keyboard(),
    )


@router.callback_query(F.data.startswith("guruh:"))
async def guruh_actions(call: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, call.from_user.id):
        return
    action = call.data.split(":")[1]
    guruhlar = guruhlar_ol()

    if action == "ro'yxat":
        if not guruhlar:
            await call.answer("⚠️ Hali guruhlar mavjud emas.", show_alert=True)
            return
        text = "📋 <b>Ulangan guruhlar:</b>\n\n"
        for i, g in enumerate(guruhlar, 1):
            text += f"{i}. {g.get('nomi') or 'Noma' + chr(39) + 'lum'} — <code>{g['chat_id']}</code>\n"
        await call.message.edit_text(
            text, parse_mode="HTML", reply_markup=guruh_boshqarish_keyboard()
        )
        await call.answer()
        return

    elif action == "ranking":
        if not guruhlar:
            await call.answer("⚠️ Guruhlar topilmadi.", show_alert=True)
            return
        sinflar = sinf_ol()
        text = "🏆 <b>Top-10 reyting (sinflar kesimida):</b>\n\n"
        for s in sinflar:
            talabalar = get_all_in_class(s)
            if talabalar:
                text += f"📍 <b>{s}:</b>\n"
                for i, t in enumerate(talabalar[:10], 1):
                    text += f"{i}. {t['ismlar']} — {t['umumiy_ball']}\n"
                text += "\n"
        sent = 0
        for g in guruhlar:
            try:
                await call.bot.send_message(
                    g["chat_id"], text, parse_mode="HTML"
                )
                sent += 1
            except Exception:
                pass
        await call.answer(f"✅ {sent} ta guruhga yuborildi.", show_alert=True)
        return

    elif action == "backup":
        data = get_all_students_for_excel()
        if not data:
            await call.answer(
                "⚠️ Backup uchun ma'lumot yo'q.", show_alert=True
            )
            return
        filename = f"group_backup_{call.from_user.id}.xlsx"
        pd.DataFrame(data).to_excel(filename, index=False)
        sent = 0
        try:
            for g in guruhlar:
                try:
                    await call.bot.send_document(
                        g["chat_id"],
                        FSInputFile(filename),
                        caption="💾 Guruh uchun backup fayl",
                    )
                    sent += 1
                except Exception:
                    pass
        finally:
            if os.path.exists(filename):
                os.remove(filename)
        await call.answer(
            f"✅ Backup {sent} ta guruhga yuborildi.", show_alert=True
        )
        return


@router.message(F.text == "📱 Ro'yxatdan o'tganlar")
async def registered_users_list(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    users = get_all_registered_users()
    if not users:
        await message.answer(
            "⚠️ Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q."
        )
        return

    text = "📱 <b>Ro'yxatdan o'tgan foydalanuvchilar:</b>\n\n"
    for i, user in enumerate(users, 1):
        username = f"@{user['username']}" if user["username"] else "yo'q"
        phone = (
            user["phone_number"] if user["phone_number"] else "kiritilmagan"
        )
        text += (
            f"{i}. <b>{user['first_name']} {user['last_name'] or ''}</b>\n"
            f"   🆔 ID: <code>{user['user_id']}</code>\n"
            f"   🔗 Username: {username}\n"
            f"   📞 Tel: <code>{phone}</code>\n"
            f"   📅 Sana: {user['registered_at'].strftime('%d.%m.%Y %H:%M')}\n\n"
        )

    # Matn uzun bo'lsa bo'lib yuborish
    chunks = _split_long_text(text)
    for chunk in chunks:
        await message.answer(chunk, parse_mode="HTML")


@router.message(F.text == "🔍 Kod bo'yicha qidirish")
async def kod_qidirish_handler(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.set_state(KodQidirish.kod_kutish)
    await message.answer(
        "🔍 Qidirish uchun o'quvchi <b>kodini</b> kiriting:", parse_mode="HTML"
    )


@router.message(KodQidirish.kod_kutish)
async def kod_qidirish_process(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)
    if not talaba:
        await message.answer("O'quvchi topilmadi.")
    else:
        last_result = talaba_songi_natija(kod)
        text = (
            f"O'quvchi ma'lumotlari:\n\n"
            f"Kod: {talaba['kod']}\n"
            f"Ism: {talaba['ismlar']}\n"
            f"Sinf: {talaba['sinf']}\n"
            f"Yo'nalish: {talaba['yonalish']}\n"
            f"Ball: {last_result['umumiy_ball'] if last_result else 'Mavjud emas'}"
        )
        await message.answer(text, parse_mode="HTML")
    await state.set_state(None)


# Yangi FSM state handlerlar
@router.message(BittaNatijaOchirish.kod_kutish)
async def bitta_natija_ochirish_process(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return

    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)

    if not talaba:
        await message.answer("O'quvchi topilmadi. Qaytadan urinib ko'ring:")
        return

    natijalar = talaba_natijalari(kod)

    if not natijalar:
        await message.answer(f"{talaba['ismlar']} uchun natijalar topilmadi.")
        await state.set_state(None)
        return

    text = f"O'chirish uchun natijani tanlang ({talaba['ismlar']}):"
    await message.answer(text, reply_markup=natijalar_ochirish_keyboard(kod))
    await state.set_state(None)


@router.message(TalabaTahrirlash.ism_kutish)
async def talaba_ism_tahrirlash_process(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return

    data = await state.get_data()
    kod = data.get("talaba_kod")

    if not kod:
        await message.answer("Xatolik yuz berdi. Qaytadan urinib ko'ring.")
        await state.set_state(None)
        return

    yangi_ism = message.text.strip()

    if talaba_tahrirlash(kod, ismlar=yangi_ism):
        await message.answer(
            f"O'quvchi ismi muvaffaqiyatli o'zgartirildi: {yangi_ism}"
        )
    else:
        await message.answer("Xatolik yuz berdi.")

    await state.set_state(None)


@router.callback_query(TalabaTahrirlash.sinf_kutish, F.data.startswith("sinf_id:"))
async def talaba_sinf_tahrirlash_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    data = await state.get_data()
    kod = data.get("talaba_kod")
    if not kod:
        await callback.message.edit_text(
            "Xatolik yuz berdi. Qaytadan urinib ko'ring."
        )
        await state.clear()
        await callback.answer()
        return

    try:
        sinf_id = int(callback.data.split(":")[1])
        yangi_sinf = sinf_id_dan_full_nomi(sinf_id)
        if not yangi_sinf:
            await callback.message.edit_text("❌ Sinf topilmadi.")
            await state.clear()
            await callback.answer()
            return
    except (ValueError, IndexError):
        await callback.message.edit_text("❌ Sinf topilmadi.")
        await state.clear()
        await callback.answer()
        return

    if talaba_tahrirlash(kod, sinf=yangi_sinf):
        await callback.message.edit_text(
            f"✅ O'quvchi sinfi muvaffaqiyatli o'zgartirildi: {yangi_sinf}"
        )
    else:
        await callback.message.edit_text("❌ Xatolik yuz berdi.")

    await state.clear()
    await callback.answer()


@router.message(TalabaTahrirlash.sinf_kutish)
async def talaba_sinf_tahrirlash_process(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return

    data = await state.get_data()
    kod = data.get("talaba_kod")

    if not kod:
        await message.answer("Xatolik yuz berdi. Qaytadan urinib ko'ring.")
        await state.set_state(None)
        return

    yangi_sinf = message.text.strip()

    if talaba_tahrirlash(kod, sinf=yangi_sinf):
        await message.answer(
            f"O'quvchi sinfi muvaffaqiyatli o'zgartirildi: {yangi_sinf}"
        )
    else:
        await message.answer("Xatolik yuz berdi.")

    await state.set_state(None)


@router.message(TalabaTahrirlash.yonalish_kutish)
async def talaba_yonalish_tahrirlash_process(
    message: Message, state: FSMContext
):
    if not await admin_tekshir(state, message.from_user.id):
        return

    data = await state.get_data()
    kod = data.get("talaba_kod")

    if not kod:
        await message.answer("Xatolik yuz berdi. Qaytadan urinib ko'ring.")
        await state.set_state(None)
        return

    yangi_yonalish = message.text.strip()

    if talaba_tahrirlash(kod, yonalish=yangi_yonalish):
        await message.answer(
            f"O'quvchi yo'nalishi muvaffaqiyatli o'zgartirildi: {yangi_yonalish}"
        )
    else:
        await message.answer("Xatolik yuz berdi.")

    await state.set_state(None)


@router.callback_query(
    TalabaTahrirlash.yonalish_kutish, F.data.startswith("yonalish_idx:")
)
async def talaba_yonalish_tahrirlash_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    data = await state.get_data()
    kod = data.get("talaba_kod")
    if not kod:
        await callback.message.edit_text(
            "Xatolik yuz berdi. Qaytadan urinib ko'ring."
        )
        await state.clear()
        await callback.answer()
        return

    try:
        idx = int(callback.data.split(":", 1)[1])
        yonalishlar = yonalish_ol()
        yangi_yonalish = yonalishlar[idx]
    except (ValueError, IndexError):
        await callback.message.edit_text("❌ Yo'nalish topilmadi.")
        await state.clear()
        await callback.answer()
        return

    if talaba_tahrirlash(kod, yonalish=yangi_yonalish):
        await callback.message.edit_text(
            f"✅ O'quvchi yo'nalishi muvaffaqiyatli o'zgartirildi: <b>{yangi_yonalish}</b>",
            parse_mode="HTML",
        )
    else:
        await callback.message.edit_text("❌ Xatolik yuz berdi.")

    await state.clear()
    await callback.answer()


# =====================================================================
# TUZATILGAN: Ishlamayotgan 6 ta tugma uchun F.text handlerlari
# =====================================================================


@router.message(F.text == "✏️ O'quvchi ma'lumotlarini tahrirlash")
async def talaba_malumot_tahrirlash_start(message: Message, state: FSMContext):
    """O'quvchi ma'lumotlarini tahrirlash — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return

    talabalar = get_all_students()
    if not talabalar:
        await message.answer("❌ Hozircha o'quvchilar mavjud emas.")
        return

    await message.answer(
        "✏️ <b>Tahrirlash uchun o'quvchini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=talaba_tahrirlash_keyboard(talabalar, page=1),
    )


@router.message(F.text == "🗑️ Bitta natijani o'chirish")
async def bitta_natija_ochirish_start(message: Message, state: FSMContext):
    """Bitta natijani o'chirish — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return

    await state.set_state(BittaNatijaOchirish.kod_kutish)
    await message.answer(
        "🗑️ <b>Natijani o'chirmoqchi bo'lgan o'quvchining kodini kiriting:</b>",
        parse_mode="HTML",
    )


@router.message(F.text == "🔄 Sinf transferi")
async def sinf_transferi_start(message: Message, state: FSMContext):
    """Sinf transferi — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return

    await message.answer(
        "🔄 <b>Sinf transferi</b>\n\nQaysi amalni bajarmoqchisiz?",
        parse_mode="HTML",
        reply_markup=sinf_transferi_keyboard(),
    )


@router.message(F.text == "📦 Bitiruvchilarni arxivlash")
async def bitiruvchilar_arxivlash_start(message: Message, state: FSMContext):
    """Bitiruvchilarni arxivlash — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return

    await message.answer(
        "📦 <b>Bitiruvchilarni arxivlash</b>\n\nQaysi sinfni arxivlamoqchisiz?",
        parse_mode="HTML",
        reply_markup=bitiruvchilar_arxivlash_keyboard(),
    )


@router.message(F.text == "🔍 Dublikatlarni topish")
async def dublikatlarni_topish_start(message: Message, state: FSMContext):
    """Dublikatlarni topish — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return

    dublikatlar = dublikatlarni_topish()
    if not dublikatlar:
        await message.answer("✅ Dublikat o'quvchilar topilmadi.")
        return

    await message.answer(
        f"🔍 <b>{len(dublikatlar)} ta dublikat topildi.</b>\n\nBirlashtirmoqchi bo'lgan o'quvchini tanlang:",
        parse_mode="HTML",
        reply_markup=dublikatlar_keyboard(dublikatlar),
    )


@router.message(F.text == "📄 PDF Hisobot")
async def pdf_hisobot_start(message: Message, state: FSMContext):
    """PDF Hisobot — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return

    await message.answer(
        "📄 <b>PDF Hisobot</b>\n\nQanday hisobot yaratmoqchisiz?",
        parse_mode="HTML",
        reply_markup=pdf_export_keyboard(),
    )


# =====================================================================
# TUZATILGAN: Maktab statistikasi va dublikatlar:back handlerlari
# =====================================================================


@router.message(F.text == "🏫 Maktab statistikasi")
async def maktab_statistikasi_start(message: Message, state: FSMContext):
    """Maktab statistikasi — tugma handleri"""
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "🏫 <b>Maktab statistikasi</b>\n\nQaysi ma'lumotni ko'rmoqchisiz?",
        parse_mode="HTML",
        reply_markup=maktab_statistikasi_keyboard(),
    )


@router.callback_query(F.data == "dublikatlar:back")
async def dublikatlar_back_callback(
    callback: CallbackQuery, state: FSMContext
):
    """Dublikatlar ro'yxatiga qaytish"""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    dublikatlar = dublikatlarni_topish()
    if not dublikatlar:
        await callback.message.edit_text("✅ Dublikat o'quvchilar topilmadi.")
    else:
        await callback.message.edit_text(
            f"🔍 <b>{len(dublikatlar)} ta dublikat topildi.</b>\n\n"
            "Birlashtirmoqchi bo'lgan o'quvchini tanlang:",
            parse_mode="HTML",
            reply_markup=dublikatlar_keyboard(dublikatlar),
        )
    await callback.answer()


# =====================================================================


@router.message(F.text == "/cancel")
async def cancel_handler(message: Message, state: FSMContext):
    """Barcha faol FSM holatlarini bekor qilish — foydalanuvchi roliga qarab to'g'ri menyu ko'rsatiladi"""
    current_state = await state.get_state()

    if current_state is None:
        await message.answer(
            "❌ Hech qanday amal bekor qilinmadi. Siz hozircha hech qanday jarayonda emassiz."
        )
        return

    await state.clear()

    # ✅ FIX: Admin bo'lsa admin menyusi, bo'lmasa o'quvchi menyusi ko'rsatiladi
    if is_admin_id(message.from_user.id):
        await message.answer(
            "✅ <b>Barcha amallar bekor qilindi!</b>\n\n🏠 Admin menyusi:",
            parse_mode="HTML",
            reply_markup=admin_menu_keyboard(),
        )
    else:
        from database import get_setting
        from keyboards import user_menu_keyboard
        await message.answer(
            "✅ <b>Barcha amallar bekor qilindi!</b>\n\n🏠 Asosiy menyu:",
            parse_mode="HTML",
            reply_markup=user_menu_keyboard(
                get_setting("ranking_enabled", "True"),
                get_setting("stats_enabled", "True"),
                get_setting("chatbot_enabled", "True"),
                get_setting("mock_enabled", "True"),
                get_setting("quiz_enabled", "True"),
                get_setting("mini_test_enabled", "True"),
            ),
        )


@router.callback_query(F.data.startswith("cancel:"))
async def cancel_callback_handler(callback: CallbackQuery, state: FSMContext):
    """Callback orqali bekor qilish"""
    current_state = await state.get_state()

    # FSM holatini tozalash
    await state.clear()

    # Callback data ga qarab qayerga qaytishni aniqlash
    target = callback.data.split(":", 1)[1]

    if target == "admin_menu":
        await callback.message.edit_text(
            "✅ <b>Barcha amallar bekor qilindi!</b>\n\n" "🏠 Asosiy menyu:",
            parse_mode="HTML",
            reply_markup=None,
        )
        await callback.message.answer(
            "Asosiy menyu:", reply_markup=admin_menu_keyboard()
        )
    else:
        await callback.message.edit_text("✅ Amal bekor qilindi.")

    await callback.answer()


# ─────────────────────────────────────────
# Yangi funktsiyalar callback handlerlari
# ─────────────────────────────────────────


# O'quvchi tahrirlash callbacklari
@router.callback_query(F.data.startswith("talaba_edit:"))
async def talaba_edit_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    kod = callback.data.split(":")[1]
    talaba = talaba_topish(kod)

    if not talaba:
        await callback.answer("❌ O'quvchi topilmadi.", show_alert=True)
        return

    text = (
        f"✏️ <b>O'quvchi ma'lumotlari:</b>\n\n"
        f"🆔 Kod: <b>{talaba['kod']}</b>\n"
        f"👤 Ism: <b>{talaba['ismlar']}</b>\n"
        f"🏫 Sinf: <b>{talaba['sinf']}</b>\n"
        f"🎯 Yo'nalish: <b>{talaba['yonalish']}</b>"
    )

    await callback.message.edit_text(
        text, parse_mode="HTML", reply_markup=talaba_edit_options_keyboard(kod)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("talaba_edit_ism:"))
async def talaba_edit_ism_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    kod = callback.data.split(":")[1]
    await state.update_data(talaba_kod=kod)
    await state.set_state(TalabaTahrirlash.ism_kutish)
    await callback.message.edit_text(
        "✏️ <b>Yangi ismni kiriting:</b>", parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data.startswith("talaba_edit_sinf:"))
async def talaba_edit_sinf_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    kod = callback.data.split(":")[1]
    await state.update_data(talaba_kod=kod)
    await state.set_state(TalabaTahrirlash.sinf_kutish)
    await callback.message.edit_text(
        "🏫 <b>Yangi sinfni tanlang:</b>",
        parse_mode="HTML",
        reply_markup=sinf_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("talaba_edit_yonalish:"))
async def talaba_edit_yonalish_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    kod = callback.data.split(":")[1]
    await state.update_data(talaba_kod=kod)
    await state.set_state(TalabaTahrirlash.yonalish_kutish)
    await callback.message.edit_text(
        "🎯 <b>Yangi yo'nalishni tanlang:</b>",
        parse_mode="HTML",
        reply_markup=yonalish_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "talaba_tahrirlash:back")
async def talaba_tahrirlash_back_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    talabalar = get_all_students()
    if not talabalar:
        await callback.message.edit_text("❌ O'quvchilar topilmadi.")
        return

    await callback.message.edit_text(
        "✏️ <b>Tahrirlash uchun o'quvchini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=talaba_tahrirlash_keyboard(talabalar, page=1),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("talaba_edit_page:"))
async def talaba_tahrirlash_page_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    page_value = callback.data.split(":", 1)[1]
    if page_value == "noop":
        await callback.answer()
        return

    talabalar = get_all_students()
    if not talabalar:
        await callback.message.edit_text("❌ O'quvchilar topilmadi.")
        await callback.answer()
        return

    await callback.message.edit_text(
        "✏️ <b>Tahrirlash uchun o'quvchini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=talaba_tahrirlash_keyboard(
            talabalar, page=int(page_value)
        ),
    )
    await callback.answer()


# Maktab statistikasi callbacklari
@router.callback_query(F.data == "maktab_stat:barchasi")
async def maktab_stat_barchasi_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    stats = maktab_statistikasi()

    if not stats:
        await callback.answer(
            "📊 Statistika ma'lumotlari topilmadi.", show_alert=True
        )
        return

    text = "📊 <b>Barcha maktablar statistikasi:</b>\n\n"
    for stat in stats:
        text += (
            f"🏫 <b>{stat['maktab_nomi']}</b> - {stat['sinf']}\n"
            f"👥 O'quvchilar: {stat['oquvchilar_soni']} ta\n"
            f"📈 O'rtacha ball: {stat['ortacha_ball']:.1f}\n"
            f"🏆 Eng yuqori: {stat['eng_yuqori_ball']:.1f}\n"
            f"📉 Eng past: {stat['eng_past_ball']:.1f}\n\n"
        )

    await callback.message.edit_text(text, parse_mode="HTML")
    await callback.answer()


@router.callback_query(F.data == "maktab_stat:solishtirish")
async def maktab_stat_solishtirish_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    await callback.message.edit_text(
        "📈 <b>Solishtirish uchun birinchi maktabni tanlang:</b>",
        parse_mode="HTML",
        reply_markup=maktab_solishtirish_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("maktab_stat:"))
async def maktab_stat_ayrim_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    stats = maktab_statistikasi(maktab_id)

    if not stats:
        await callback.answer(
            "📊 Bu maktab statistikasi topilmadi.", show_alert=True
        )
        return

    text = f"📊 <b>{stats[0]['maktab_nomi']} statistikasi:</b>\n\n"
    for stat in stats:
        text += (
            f"🏫 {stat['sinf']}\n"
            f"👥 O'quvchilar: {stat['oquvchilar_soni']} ta\n"
            f"📈 O'rtacha ball: {stat['ortacha_ball']:.1f}\n"
            f"🏆 Eng yuqori: {stat['eng_yuqori_ball']:.1f}\n"
            f"📉 Eng past: {stat['eng_past_ball']:.1f}\n\n"
        )

    await callback.message.edit_text(text, parse_mode="HTML")
    await callback.answer()


# Natijalarni o'chirish callbacklari
@router.callback_query(F.data.startswith("natija_ochir:"))
async def natija_ochir_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    parts = callback.data.split(":")
    talaba_kod = parts[1]
    natija_id = int(parts[2])

    if bitta_natija_ochir(talaba_kod, natija_id):
        await callback.answer(
            "Natija muvaffaqiyatli o'chirildi.", show_alert=True
        )

        # Yangilangan ro'yxatni ko'rsatish
        natijalar = talaba_natijalari(talaba_kod)
        if natijalar:
            await callback.message.edit_text(
                f"Natija o'chirildi. Qolgan natijalar:",
                reply_markup=natijalar_ochirish_keyboard(talaba_kod),
            )
        else:
            await callback.message.edit_text("Barcha natijalar o'chirildi.")
    else:
        await callback.answer("Xatolik yuz berdi.", show_alert=True)


# Sinf transferi callbacklari
@router.callback_query(F.data == "sinf_transfer:barchasi")
async def sinf_transfer_barchasi_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    await callback.message.edit_text(
        "Barcha sinflarni yuqoriga ko'chirish amali bajarilmoqda...",
        reply_markup=None,
    )

    # 9->10, 10->11 sinflarga ko'chirish
    transfer_count = 0
    transfer_count += sinf_transferi("9-A", "10-A")
    transfer_count += sinf_transferi("9-B", "10-B")
    transfer_count += sinf_transferi("10-A", "11-A")
    transfer_count += sinf_transferi("10-B", "11-B")

    await callback.message.edit_text(
        f"Jami {transfer_count} ta o'quvchi yuqori sinfga ko'chirildi."
    )
    await callback.answer()


@router.callback_query(F.data == "sinf_transfer:tanlash")
async def sinf_transfer_tanlash_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    await callback.message.edit_text(
        "Ko'chirish uchun eski sinfni tanlang:",
        reply_markup=sinf_tanlash_transfer_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "sinf_transfer:menu")
async def sinf_transfer_menu_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    await callback.message.edit_text(
        "🔄 <b>Sinf transferi</b>\n\nQaysi amalni bajarmoqchisiz?",
        parse_mode="HTML",
        reply_markup=sinf_transferi_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("sinf_transfer_eski:"))
async def sinf_transfer_eski_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    eski_sinf_id = int(callback.data.split(":")[1])
    eski_sinf_nomi = sinf_nomi_ol_by_id(eski_sinf_id) or str(eski_sinf_id)
    await callback.message.edit_text(
        f"Yangi sinfni tanlang ({eski_sinf_nomi} -> ?):",
        reply_markup=yangi_sinf_tanlash_keyboard(eski_sinf_id),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("sinf_transfer_yangi:"))
async def sinf_transfer_yangi_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    parts = callback.data.split(":")
    eski_sinf_id = int(parts[1])
    yangi_sinf_id = int(parts[2])

    count = sinf_transferi_by_id(eski_sinf_id, yangi_sinf_id)

    eski_nomi = sinf_nomi_ol_by_id(eski_sinf_id) or str(eski_sinf_id)
    yangi_nomi = sinf_nomi_ol_by_id(yangi_sinf_id) or str(yangi_sinf_id)

    await callback.message.edit_text(
        f"✅ {eski_nomi} sinfidan {yangi_nomi} sinfiga {count} ta o'quvchi ko'chirildi."
    )
    await callback.answer()


# Arxivlash callbacklari
@router.callback_query(F.data == "arxivlash:barcha_11")
async def arxivlash_barcha_11_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    count = bitiruvchilarni_arxivlash()
    await callback.message.edit_text(
        f"Jami {count} ta 11-sinf o'quvchisi arxivlandi."
    )
    await callback.answer()


@router.callback_query(F.data == "arxivlash:tanlash")
async def arxivlash_tanlash_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    await callback.message.edit_text(
        "Arxivlash uchun sinfni tanlang:",
        reply_markup=sinf_arxivlash_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "arxivlash:chiqarish")
async def arxivdan_chiqarish_tanlash_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    await callback.message.edit_text(
        "📤 Arxivdan chiqarish uchun sinfni tanlang:",
        reply_markup=sinf_arxivdan_chiqarish_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("arxivlash_sinf:"))
async def arxivlash_sinf_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    sinf = callback.data.split(":")[1]
    count = bitiruvchilarni_arxivlash(sinf)

    await callback.message.edit_text(
        f"{sinf} sinfidan {count} ta o'quvchi arxivlandi."
    )
    await callback.answer()


@router.callback_query(F.data.startswith("arxivdan_chiqarish_sinf:"))
async def arxivdan_chiqarish_sinf_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    sinf = callback.data.split(":")[1]
    count = bitiruvchilarni_arxivdan_chiqarish(sinf)

    await callback.message.edit_text(
        f"{sinf} sinfidan {count} ta o'quvchi arxivdan chiqarildi."
    )
    await callback.answer()


# Dublikatlar callbacklari
@router.callback_query(F.data.startswith("dublikat_tanlash:"))
async def dublikat_tanlash_callback(
    callback: CallbackQuery, state: FSMContext
):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    ism = callback.data.split(":", 1)[1]

    from database import dublikatlarni_topish

    dublikatlar = dublikatlarni_topish()
    dublikat = next((d for d in dublikatlar if d["ismlar"] == ism), None)

    if not dublikat:
        await callback.answer("Dublikat topilmadi.", show_alert=True)
        return

    await callback.message.edit_text(
        f"Qaysi o'quvchini asosiy qilib saqlashni xohlaysiz?\n\n"
        f"ism: {dublikat['ismlar']}\n"
        f"kodlar: {dublikat['kodlar']}",
        reply_markup=dublikat_birlashtirish_keyboard(ism, dublikat["kodlar"]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("dublikat_asosiy:"))
async def dublikat_asosiy_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    parts = callback.data.split(":")
    ism = parts[1]
    asosiy_kod = parts[2]

    from database import dublikatlarni_topish

    dublikatlar = dublikatlarni_topish()
    dublikat = next((d for d in dublikatlar if d["ismlar"] == ism), None)

    if not dublikat:
        await callback.answer("Dublikat topilmadi.", show_alert=True)
        return

    # Asosiy kodni olib tashlab, qolganlarini qo'shimcha qilish
    barcha_kodlar = dublikat["kodlar"].split(", ")
    qoshimcha_kodlar = [kod for kod in barcha_kodlar if kod != asosiy_kod]

    if dublikatlarni_birlashtir(asosiy_kod, qoshimcha_kodlar):
        await callback.message.edit_text(
            f"O'quvchilar muvaffaqiyatli birlashtirildi!\n"
            f"Asosiy kod: {asosiy_kod}\n"
            f"Birlashtirilgan: {len(qoshimcha_kodlar)} ta"
        )
        await callback.answer("Birlashtirish muvaffaqiyatli!", show_alert=True)
    else:
        await callback.answer("Xatolik yuz berdi.", show_alert=True)


# Maktablar solishtirish callbacklari
@router.callback_query(F.data.startswith("maktab_comp1:"))
async def maktab_comp1_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab1_id = int(callback.data.split(":")[1])
    await callback.message.edit_text(
        "2-maktabni tanlang:", reply_markup=maktab_comp2_keyboard(maktab1_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("maktab_comp2:"))
async def maktab_comp2_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    parts = callback.data.split(":")
    maktab1_id = int(parts[1])
    maktab2_id = int(parts[2])

    # Ikkala maktab statistikasini olish
    stats1 = maktab_statistikasi(maktab1_id)
    stats2 = maktab_statistikasi(maktab2_id)

    if not stats1 or not stats2:
        await callback.answer(
            "Maktablar statistikasi topilmadi.", show_alert=True
        )
        return

    # Maktab nomlarini olish
    from database import maktablar_ol

    maktablar = maktablar_ol()
    maktab1 = next((m for m in maktablar if m["id"] == maktab1_id), None)
    maktab2 = next((m for m in maktablar if m["id"] == maktab2_id), None)

    text = f"📊 <b>Maktablar solishtirish:</b>\n\n"
    text += f"🏫 {maktab1['nomi'] if maktab1 else 'Maktab 1'}\n"
    text += f"🏫 {maktab2['nomi'] if maktab2 else 'Maktab 2'}\n\n"

    # Umumiy statistikani solishtirish
    total1 = sum(s["oquvchilar_soni"] for s in stats1)
    total2 = sum(s["oquvchilar_soni"] for s in stats2)
    avg1 = (
        sum(s["ortacha_ball"] * s["oquvchilar_soni"] for s in stats1) / total1
        if total1 > 0
        else 0
    )
    avg2 = (
        sum(s["ortacha_ball"] * s["oquvchilar_soni"] for s in stats2) / total2
        if total2 > 0
        else 0
    )

    text += f"👥 O'quvchilar soni: {total1} vs {total2}\n"
    text += f"📈 O'rtacha ball: {avg1:.1f} vs {avg2:.1f}\n\n"

    # Sinf bo'yicha solishtirish
    sinflar1 = {s["sinf"]: s for s in stats1}
    sinflar2 = {s["sinf"]: s for s in stats2}

    common_sinflar = set(sinflar1.keys()) & set(sinflar2.keys())

    if common_sinflar:
        text += "📋 <b>Sinf bo'yicha solishtirish:</b>\n"
        for sinf in sorted(common_sinflar):
            s1 = sinflar1[sinf]
            s2 = sinflar2[sinf]
            text += f"📌 {sinf}: {s1['ortacha_ball']:.1f} vs {s2['ortacha_ball']:.1f}\n"

    await callback.message.edit_text(text, parse_mode="HTML")
    await callback.answer()


# PDF export callback handlerlari
@router.callback_query(F.data.startswith("pdf:"))
async def pdf_export_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    action = callback.data.split(":")[1]

    if action == "student":
        await state.set_state(PDFStudentKod.kod_kutish)
        await callback.message.edit_text(
            "📄 PDF hisobotini yaratmoqchi bo'lgan o'quvchi kodini kiriting:"
        )

    elif action == "maktab_stat":
        await callback.message.edit_text(
            "🏫 Maktab statistikasi uchun maktabni tanlang:",
            reply_markup=maktab_tanlash_keyboard(maktablar_ol(), "pdf_maktab"),
        )

    elif action == "sinf_reyting":
        await callback.message.edit_text(
            "🏆 Sinf reytingi uchun sinfni tanlang:",
            reply_markup=sinf_tanlash_pdf_keyboard(),
        )

    elif action == "menu":
        await callback.message.edit_text(
            "📄 <b>PDF Hisobot:</b>\n\n" "Qaysi turdagi hisobotni yaratasiz?",
            reply_markup=pdf_export_keyboard(),
        )

    await callback.answer()


@router.callback_query(F.data.startswith("pdf_maktab:"))
async def pdf_maktab_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])

    wait_msg = await callback.message.answer("⏳ PDF yaratilmoqda...")

    try:
        from pdf_export import PDFExporter

        exporter = PDFExporter()
        pdf_path = exporter.create_maktab_statistika_pdf(maktab_id)

        if pdf_path and os.path.exists(pdf_path):
            await callback.bot.send_document(
                callback.from_user.id,
                FSInputFile(pdf_path),
                caption="📄 Maktab statistikasi PDF hisoboti",
            )
            await wait_msg.delete()
            await callback.answer(
                "✅ PDF muvaffaqiyatli yaratildi!", show_alert=True
            )
        else:
            await wait_msg.delete()
            await callback.answer(
                "❌ PDF yaratishda xatolik!", show_alert=True
            )

    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


@router.callback_query(F.data == "pdf:menu")
async def pdf_menu_callback(callback: CallbackQuery, state: FSMContext):
    """PDF menu callback handler."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    await callback.message.edit_text(
        "📄 <b>PDF Hisobot:</b>\n\n" "Qaysi turdagi hisobotni yaratasiz?",
        reply_markup=pdf_export_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("lang:"))
async def lang_callback(callback: CallbackQuery, state: FSMContext):
    """Language selection callback handler."""
    action = callback.data.split(":")[1]
    
    if action == "select":
        # Show language selection menu
        from keyboards import language_selection_keyboard
        await callback.message.edit_text(
            "🌐 Tilni tanlang:",
            reply_markup=language_selection_keyboard()
        )
    else:
        # Language code selected
        try:
            from i18n import i18n
            lang_code = action
            if lang_code in i18n.get_available_languages():
                # Set user language preference
                await state.update_data(language=lang_code)
                await callback.answer("✅ Til o'zgartirildi!", show_alert=True)
            else:
                await callback.answer("❌ Noto'g'ri til tanlovi", show_alert=True)
        except Exception:
            await callback.answer("❌ Til o'zgartirishda xatolik", show_alert=True)
    
    await callback.answer()


@router.callback_query(F.data == "menu:main")
async def menu_main_callback(callback: CallbackQuery, state: FSMContext):
    """Main menu callback handler."""
    await state.clear()
    await callback.message.edit_text(
        "🏠 Asosiy menyu:",
        reply_markup=admin_menu_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("select_maktab:"))
async def select_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """Maktab tanlash callback handler."""
    if not await admin_tekshir(state, callback.from_user.id):
        return
    
    try:
        maktab_id = int(callback.data.split(":")[1])
        # This handler can be used for different purposes
        # For now, just acknowledge the selection
        await callback.answer(f"🏫 Maktab {maktab_id} tanlandi", show_alert=True)
    except (ValueError, IndexError):
        await callback.answer("❌ Noto'g'ri maktab tanlovi", show_alert=True)


@router.callback_query(F.data.startswith("pdf_sinf:"))
async def pdf_sinf_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    sinf = callback.data.split(":")[1]

    wait_msg = await callback.message.answer("⏳ PDF yaratilmoqda...")

    try:
        from pdf_export import PDFExporter

        exporter = PDFExporter()
        pdf_path = exporter.create_sinf_reyting_pdf(
            sinf if sinf != "all" else None
        )

        if pdf_path and os.path.exists(pdf_path):
            await callback.bot.send_document(
                callback.from_user.id,
                FSInputFile(pdf_path),
                caption="📄 Sinf reytingi PDF hisoboti",
            )
            await wait_msg.delete()
            await callback.answer(
                "✅ PDF muvaffaqiyatli yaratildi!", show_alert=True
            )
        else:
            await wait_msg.delete()
            await callback.answer(
                "❌ PDF yaratishda xatolik!", show_alert=True
            )

    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


@router.callback_query(F.data == "pdf_sinf_by_maktab")
async def pdf_sinf_by_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """Maktab tanlash — keyin shu maktabdagi sinflar ko'rsatiladi."""
    if not await admin_tekshir(state, callback.from_user.id):
        return

    from database import maktablar_ol

    maktablar = maktablar_ol()
    await callback.message.edit_text(
        "🏫 Maktabni tanlang:",
        reply_markup=maktab_tanlash_keyboard(maktablar, prefix="pdf_reyting_maktab"),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("pdf_reyting_maktab:"))
async def pdf_reyting_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """Tanlangan maktabdagi sinflar ro'yxatini ko'rsatish."""
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    await callback.message.edit_text(
        "📋 Sinfni tanlang yoki barcha sinflarni bir PDF da oling:",
        reply_markup=sinf_tanlash_pdf_maktab_keyboard(maktab_id),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("pdf_sinf_maktab:"))
async def pdf_sinf_maktab_callback(callback: CallbackQuery, state: FSMContext):
    """Tanlangan maktabdagi BARCHA sinflar reytingini PDF ga chiqarish."""
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    wait_msg = await callback.message.answer("⏳ PDF yaratilmoqda...")

    try:
        from pdf_export import PDFExporter

        exporter = PDFExporter()
        pdf_path = exporter.create_sinf_reyting_pdf(maktab_id=maktab_id)

        if pdf_path and os.path.exists(pdf_path):
            await callback.bot.send_document(
                callback.from_user.id,
                FSInputFile(pdf_path),
                caption="📄 Maktab sinflar reytingi PDF hisoboti",
            )
            await wait_msg.delete()
            await callback.answer("✅ PDF muvaffaqiyatli yaratildi!", show_alert=True)
        else:
            await wait_msg.delete()
            await callback.answer("❌ PDF yaratishda xatolik!", show_alert=True)

    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


# PDF student kod uchun handler
class PDFStudentKod(StatesGroup):
    kod_kutish = State()


@router.message(PDFStudentKod.kod_kutish)
async def pdf_student_kod_handler(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return

    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)

    if not talaba:
        await message.answer("❌ O'quvchi topilmadi. Qaytadan urinib ko'ring:")
        return

    wait_msg = await message.answer("⏳ PDF yaratilmoqda...")

    try:
        from pdf_export import PDFExporter

        exporter = PDFExporter()
        pdf_path = exporter.create_student_pdf(kod)

        if pdf_path and os.path.exists(pdf_path):
            await message.bot.send_document(
                message.from_user.id,
                FSInputFile(pdf_path),
                caption=f"📄 {talaba['ismlar']} uchun PDF hisoboti",
            )
            await wait_msg.delete()
            await message.answer("✅ PDF muvaffaqiyatli yaratildi!")
        else:
            await wait_msg.delete()
            await message.answer("❌ PDF yaratishda xatolik!")

    except Exception as e:
        await wait_msg.delete()
        await message.answer(f"❌ Xatolik: {str(e)}")

    await state.set_state(None)


# ─────────────────────────────────────────
# CHATBOT — Admin ko'rish paneli
# ─────────────────────────────────────────

@router.message(F.text == "🤖 Chatbot foydalanuvchilar")
async def chatbot_foydalanuvchilar_handler(message: Message, state: FSMContext):
    """Admin oxirgi chatbot foydalanuvchilarini ko'radi."""
    if not await admin_tekshir(state, message.from_user.id):
        return

    bugun = chatbot_bugungi_son()
    foydalanuvchilar = chatbot_oxirgi_foydalanuvchilar(limit=20)

    if not foydalanuvchilar:
        await message.answer(
            f"🤖 <b>Chatbot statistikasi</b>\n\n"
            f"📅 Bugun: <b>{bugun} ta so'rov</b>\n\n"
            "Hali hech kim chatbotdan foydalanmagan.",
            parse_mode="HTML",
        )
        return

    text = (
        f"🤖 <b>Chatbot statistikasi</b>\n"
        f"📅 Bugun: <b>{bugun} ta so'rov</b>\n"
        f"👥 So'nggi foydalanuvchilar:\n"
        "─────────────────────\n"
    )

    for i, u in enumerate(foydalanuvchilar, 1):
        ism = u.get("talaba_ism") or "Noma'lum"
        kod = u.get("talaba_kod") or "—"
        vaqt = u.get("created_at")
        vaqt_str = vaqt.strftime("%d.%m %H:%M") if vaqt else "—"
        savol = u.get("savol", "")
        savol_qisqa = (savol[:40] + "...") if len(savol) > 40 else savol

        text += (
            f"\n{i}. <b>{ism}</b> ({kod})\n"
            f"   🕐 {vaqt_str}\n"
            f"   💬 <i>{savol_qisqa}</i>\n"
        )

    await message.answer(text, parse_mode="HTML")


# =====================================================================
# EXCEL HISOBOT — barcha handlerlar
# =====================================================================

class ExcelStudentKod(StatesGroup):
    kod_kutish = State()


@router.message(F.text == "📊 Excel Hisobot")
async def excel_hisobot_start(message: Message, state: FSMContext):
    """Excel Hisobot — asosiy menyu."""
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "📊 <b>Excel Hisobot</b>\n\nQanday hisobot yaratmoqchisiz?",
        parse_mode="HTML",
        reply_markup=excel_export_keyboard(),
    )


@router.callback_query(F.data.startswith("excel:"))
async def excel_export_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    action = callback.data.split(":")[1]

    if action == "student":
        await state.set_state(ExcelStudentKod.kod_kutish)
        await callback.message.edit_text(
            "📊 Excel hisobotini yaratmoqchi bo'lgan o'quvchi kodini kiriting:"
        )

    elif action == "maktab_stat":
        await callback.message.edit_text(
            "🏫 Maktab statistikasi uchun maktabni tanlang:",
            reply_markup=maktab_tanlash_keyboard(maktablar_ol(), "excel_maktab"),
        )

    elif action == "sinf_reyting":
        await callback.message.edit_text(
            "🏆 Sinf reytingi uchun sinfni tanlang:",
            reply_markup=sinf_tanlash_excel_keyboard(),
        )

    elif action == "menu":
        await callback.message.edit_text(
            "📊 <b>Excel Hisobot:</b>\n\nQanday hisobot yaratmoqchisiz?",
            parse_mode="HTML",
            reply_markup=excel_export_keyboard(),
        )

    await callback.answer()


@router.message(ExcelStudentKod.kod_kutish)
async def excel_student_kod_handler(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return

    kod = message.text.strip().upper()
    talaba = talaba_topish(kod)

    if not talaba:
        await message.answer(
            f"❌ <code>{kod}</code> kodli o'quvchi topilmadi.",
            parse_mode="HTML",
        )
        await state.clear()
        return

    wait_msg = await message.answer("⏳ Excel yaratilmoqda...")

    try:
        from excel_export import ExcelExporter

        exporter = ExcelExporter()
        path = exporter.create_student_excel(kod)

        if path and os.path.exists(path):
            await message.bot.send_document(
                message.from_user.id,
                FSInputFile(path),
                caption=f"📊 {talaba['ismlar']} — Excel hisoboti",
            )
            await wait_msg.delete()
        else:
            await wait_msg.delete()
            await message.answer("❌ Excel yaratishda xatolik yuz berdi.")
    except Exception as e:
        await wait_msg.delete()
        await message.answer(f"❌ Xatolik: {str(e)}")

    await state.clear()


@router.callback_query(F.data.startswith("excel_maktab:"))
async def excel_maktab_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    wait_msg = await callback.message.answer("⏳ Excel yaratilmoqda...")

    try:
        from excel_export import ExcelExporter

        exporter = ExcelExporter()
        path = exporter.create_maktab_statistika_excel(maktab_id)

        if path and os.path.exists(path):
            await callback.bot.send_document(
                callback.from_user.id,
                FSInputFile(path),
                caption="📊 Maktab statistikasi — Excel hisoboti",
            )
            await wait_msg.delete()
            await callback.answer("✅ Excel muvaffaqiyatli yaratildi!", show_alert=True)
        else:
            await wait_msg.delete()
            await callback.answer("❌ Excel yaratishda xatolik!", show_alert=True)
    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


@router.callback_query(F.data.startswith("excel_sinf:"))
async def excel_sinf_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    sinf = callback.data.split(":")[1]
    wait_msg = await callback.message.answer("⏳ Excel yaratilmoqda...")

    try:
        from excel_export import ExcelExporter

        exporter = ExcelExporter()
        path = exporter.create_sinf_reyting_excel(
            sinf if sinf != "all" else None
        )

        if path and os.path.exists(path):
            await callback.bot.send_document(
                callback.from_user.id,
                FSInputFile(path),
                caption="📊 Sinf reytingi — Excel hisoboti",
            )
            await wait_msg.delete()
            await callback.answer("✅ Excel muvaffaqiyatli yaratildi!", show_alert=True)
        else:
            await wait_msg.delete()
            await callback.answer("❌ Excel yaratishda xatolik!", show_alert=True)
    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


@router.callback_query(F.data == "excel_sinf_by_maktab")
async def excel_sinf_by_maktab_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktablar = maktablar_ol()
    await callback.message.edit_text(
        "🏫 Maktabni tanlang:",
        reply_markup=maktab_tanlash_keyboard(maktablar, prefix="excel_reyting_maktab"),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("excel_reyting_maktab:"))
async def excel_reyting_maktab_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    await callback.message.edit_text(
        "📋 Sinfni tanlang yoki barcha sinflarni bir Excel da oling:",
        reply_markup=sinf_tanlash_excel_maktab_keyboard(maktab_id),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("excel_sinf_maktab:"))
async def excel_sinf_maktab_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    wait_msg = await callback.message.answer("⏳ Excel yaratilmoqda...")

    try:
        from excel_export import ExcelExporter

        exporter = ExcelExporter()
        path = exporter.create_sinf_reyting_excel(maktab_id=maktab_id)

        if path and os.path.exists(path):
            await callback.bot.send_document(
                callback.from_user.id,
                FSInputFile(path),
                caption="📊 Maktab sinflar reytingi — Excel hisoboti",
            )
            await wait_msg.delete()
            await callback.answer("✅ Excel muvaffaqiyatli yaratildi!", show_alert=True)
        else:
            await wait_msg.delete()
            await callback.answer("❌ Excel yaratishda xatolik!", show_alert=True)
    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


# =====================================================================
# SINF TAQQOSLASH — to'liq handler bloki
# =====================================================================

class SinfTaqqoslash(StatesGroup):
    sinf_a_kutish = State()
    sinf_b_kutish = State()


def _taqq_sinf_matn(s: dict) -> str:
    """Bir sinf statistikasini chiroyli matn sifatida qaytaradi."""
    top3_text = ""
    for i, t in enumerate(s.get("top3", []), 1):
        medal = ["🥇", "🥈", "🥉"][i - 1]
        top3_text += f"  {medal} {t['ismlar']} — {t['umumiy_ball']} ball\n"

    return (
        f"👥 O'quvchilar: <b>{s['oquvchilar_soni']} ta</b>\n"
        f"📊 O'rtacha: <b>{s['avg_umumiy']} ball</b>\n"
        f"   • Majburiy: {s['avg_majburiy']}\n"
        f"   • Asosiy 1: {s['avg_asosiy1']}\n"
        f"   • Asosiy 2: {s['avg_asosiy2']}\n"
        f"🏅 Eng yuqori: <b>{s['eng_yuqori']} ball</b>\n"
        f"📉 Eng past: <b>{s['eng_past']} ball</b>\n"
        f"\n📊 Ball tarqalishi:\n"
        f"  ≥160: {s['yuqori_soni']} ta  |  130–159: {s['orta_soni']} ta\n"
        f"  100–129: {s['past_soni']} ta  |  &lt;100: {s['juda_past_soni']} ta\n"
        f"\n🏆 Top-3:\n{top3_text}"
    )


def _taqq_farq_emoji(a_val, b_val) -> str:
    """Ikkita son farqini emoji bilan qaytaradi."""
    if a_val is None or b_val is None:
        return "—"
    diff = float(a_val) - float(b_val)
    if diff > 5:
        return f"▲ +{diff:.1f}"
    elif diff < -5:
        return f"▼ {diff:.1f}"
    else:
        return f"≈ {diff:+.1f}"


def _create_comparison_chart(a: dict, b: dict) -> str | None:
    """Ikkita sinf uchun yonma-yon bar chart yaratadi."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np

        categories = ["Majburiy", "Asosiy 1", "Asosiy 2", "Umumiy"]
        vals_a = [
            float(a["avg_majburiy"] or 0),
            float(a["avg_asosiy1"] or 0),
            float(a["avg_asosiy2"] or 0),
            float(a["avg_umumiy"] or 0),
        ]
        vals_b = [
            float(b["avg_majburiy"] or 0),
            float(b["avg_asosiy1"] or 0),
            float(b["avg_asosiy2"] or 0),
            float(b["avg_umumiy"] or 0),
        ]

        x = np.arange(len(categories))
        width = 0.35

        fig, ax = plt.subplots(figsize=(10, 6))
        bars_a = ax.bar(x - width / 2, vals_a, width, label=a["sinf"],
                        color="#4A90D9", alpha=0.85)
        bars_b = ax.bar(x + width / 2, vals_b, width, label=b["sinf"],
                        color="#F5A623", alpha=0.85)

        ax.set_title(
            f"Sinf taqqoslash: {a['sinf']} vs {b['sinf']}",
            fontsize=15, fontweight="bold", pad=16
        )
        ax.set_xticks(x)
        ax.set_xticklabels(categories, fontsize=11)
        ax.set_ylabel("O'rtacha ball", fontsize=11)
        ax.legend(fontsize=11)
        ax.grid(axis="y", alpha=0.3)

        max_val = max(max(vals_a), max(vals_b))
        ax.set_ylim(0, max_val * 1.15)

        for bar in bars_a:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h + 1,
                    f"{h:.1f}", ha="center", va="bottom",
                    fontsize=9, fontweight="bold", color="#2C5F8A")

        for bar in bars_b:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h + 1,
                    f"{h:.1f}", ha="center", va="bottom",
                    fontsize=9, fontweight="bold", color="#8B5E00")

        plt.tight_layout()
        filename = f"taqq_{a['sinf']}_{b['sinf']}.png".replace(" ", "_")
        plt.savefig(filename, dpi=110, bbox_inches="tight")
        plt.close()
        return filename
    except Exception:
        return None


@router.message(F.text == "⚖️ Sinf taqqoslash")
async def sinf_taqq_start(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await state.set_state(SinfTaqqoslash.sinf_a_kutish)
    await message.answer(
        "⚖️ <b>Sinf taqqoslash</b>\n\n"
        "1-sinfni tanlang:",
        parse_mode="HTML",
        reply_markup=sinf_taqqoslash_birinchi_keyboard(),
    )


@router.callback_query(F.data == "taqq_bekor")
async def taqq_bekor(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    try:
        await callback.message.edit_text("❌ Taqqoslash bekor qilindi.")
    except Exception:
        try:
            await callback.message.delete()
        except Exception:
            pass
        await callback.message.answer("❌ Taqqoslash bekor qilindi.")
    await callback.answer()


@router.callback_query(F.data == "taqq_qayta")
async def taqq_qayta(callback: CallbackQuery, state: FSMContext):
    await state.set_state(SinfTaqqoslash.sinf_a_kutish)
    try:
        await callback.message.edit_text(
            "⚖️ <b>Sinf taqqoslash</b>\n\n1-sinfni tanlang:",
            parse_mode="HTML",
            reply_markup=sinf_taqqoslash_birinchi_keyboard(),
        )
    except Exception:
        try:
            await callback.message.delete()
        except Exception:
            pass
        await callback.message.answer(
            "⚖️ <b>Sinf taqqoslash</b>\n\n1-sinfni tanlang:",
            parse_mode="HTML",
            reply_markup=sinf_taqqoslash_birinchi_keyboard(),
        )
    await callback.answer()


@router.callback_query(SinfTaqqoslash.sinf_a_kutish, F.data.startswith("taqq_a:"))
async def taqq_sinf_a_tanlandi(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    sinf_a = callback.data.split(":", 1)[1]
    await state.update_data(sinf_a=sinf_a)
    await state.set_state(SinfTaqqoslash.sinf_b_kutish)
    await callback.message.edit_text(
        f"✅ 1-sinf: <b>{sinf_a}</b>\n\n"
        "2-sinfni tanlang:",
        parse_mode="HTML",
        reply_markup=sinf_taqqoslash_ikkinchi_keyboard(sinf_a),
    )
    await callback.answer()


@router.callback_query(SinfTaqqoslash.sinf_b_kutish, F.data.startswith("taqq_b:"))
async def taqq_sinf_b_tanlandi(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    sinf_b = callback.data.split(":", 1)[1]
    data = await state.get_data()
    sinf_a = data.get("sinf_a")
    await state.clear()

    wait_msg = await callback.message.edit_text("⏳ Taqqoslanmoqda...")

    result = get_two_sinf_comparison(sinf_a, sinf_b)
    if not result:
        await wait_msg.edit_text(
            f"❌ <b>{sinf_a}</b> yoki <b>{sinf_b}</b> sinfi uchun ma'lumot topilmadi.\n"
            "O'quvchilar test topshirgan bo'lishi kerak.",
            parse_mode="HTML",
        )
        await callback.answer()
        return

    a = result["a"]
    b = result["b"]

    # Asosiy taqqoslash xabari
    farq_umumiy  = _taqq_farq_emoji(a["avg_umumiy"],   b["avg_umumiy"])
    farq_majb    = _taqq_farq_emoji(a["avg_majburiy"],  b["avg_majburiy"])
    farq_asosiy1 = _taqq_farq_emoji(a["avg_asosiy1"],   b["avg_asosiy1"])
    farq_asosiy2 = _taqq_farq_emoji(a["avg_asosiy2"],   b["avg_asosiy2"])

    text = (
        f"⚖️ <b>Sinf taqqoslash natijasi</b>\n\n"
        f"{'─' * 28}\n"
        f"📚 <b>{a['sinf']}</b>\n"
        f"{_taqq_sinf_matn(a)}\n"
        f"{'─' * 28}\n"
        f"📚 <b>{b['sinf']}</b>\n"
        f"{_taqq_sinf_matn(b)}\n"
        f"{'─' * 28}\n"
        f"📊 <b>Farq ({a['sinf']} → {b['sinf']})</b>\n"
        f"  Umumiy:   {farq_umumiy}\n"
        f"  Majburiy: {farq_majb}\n"
        f"  Asosiy 1: {farq_asosiy1}\n"
        f"  Asosiy 2: {farq_asosiy2}\n"
    )

    # Grafik yaratish
    chart_path = await asyncio.get_event_loop().run_in_executor(
        None, _create_comparison_chart, a, b
    )

    if chart_path and os.path.exists(chart_path):
        await callback.bot.send_photo(
            callback.from_user.id,
            photo=FSInputFile(chart_path),
            caption=text,
            parse_mode="HTML",
            reply_markup=sinf_taqqoslash_natija_keyboard(sinf_a, sinf_b),
        )
        await wait_msg.delete()
        try:
            os.remove(chart_path)
        except Exception:
            pass
    else:
        await wait_msg.edit_text(
            text,
            parse_mode="HTML",
            reply_markup=sinf_taqqoslash_natija_keyboard(sinf_a, sinf_b),
        )

    await callback.answer()


@router.callback_query(F.data.startswith("taqq_excel:"))
async def taqq_excel_eksport(callback: CallbackQuery, state: FSMContext):
    """Taqqoslash natijasini Excel faylga eksport qilish."""
    if not await admin_tekshir(state, callback.from_user.id):
        return

    parts = callback.data.split(":", 1)[1]
    sinf_a, sinf_b = parts.split("|")

    wait = await callback.message.answer("⏳ Excel yaratilmoqda...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os as _os

        result = get_two_sinf_comparison(sinf_a, sinf_b)
        if not result:
            await wait.delete()
            await callback.answer("❌ Ma'lumot topilmadi", show_alert=True)
            return

        a = result["a"]
        b = result["b"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Taqqoslash"

        # Sarlavha
        ws.merge_cells("A1:C1")
        tc = ws["A1"]
        tc.value = f"Sinf taqqoslash: {a['sinf']} vs {b['sinf']}"
        tc.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hfill = PatternFill("solid", start_color="1F4E79")
        hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")

        headers = ["Ko'rsatkich", a["sinf"], b["sinf"]]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = hfont; cell.fill = hfill
            cell.alignment = center; cell.border = border
        ws.row_dimensions[2].height = 22

        rows_data = [
            ("O'quvchilar soni",      a["oquvchilar_soni"],  b["oquvchilar_soni"]),
            ("O'rtacha ball",         a["avg_umumiy"],       b["avg_umumiy"]),
            ("O'rtacha (Majburiy)",   a["avg_majburiy"],     b["avg_majburiy"]),
            ("O'rtacha (Asosiy 1)",   a["avg_asosiy1"],      b["avg_asosiy1"]),
            ("O'rtacha (Asosiy 2)",   a["avg_asosiy2"],      b["avg_asosiy2"]),
            ("Eng yuqori ball",       a["eng_yuqori"],       b["eng_yuqori"]),
            ("Eng past ball",         a["eng_past"],         b["eng_past"]),
            ("≥160 ball (yuqori)",    a["yuqori_soni"],      b["yuqori_soni"]),
            ("130–159 ball (o'rta)",  a["orta_soni"],        b["orta_soni"]),
            ("100–129 ball (past)",   a["past_soni"],        b["past_soni"]),
            ("<100 ball",             a["juda_past_soni"],   b["juda_past_soni"]),
        ]

        alt_fill = PatternFill("solid", start_color="D6E4F0")
        normal_font = Font(name="Arial", size=10)
        for i, (label, val_a, val_b) in enumerate(rows_data, 3):
            is_alt = (i % 2 == 0)
            fill = alt_fill if is_alt else PatternFill()
            for col, val in enumerate([label, val_a, val_b], 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = normal_font; cell.fill = fill
                cell.alignment = center; cell.border = border

        # Top-3 har bir sinf uchun
        top_row = len(rows_data) + 4
        for sinf_key, stat in [(a["sinf"], a), (b["sinf"], b)]:
            ws.merge_cells(
                start_row=top_row, start_column=1,
                end_row=top_row,   end_column=3
            )
            th = ws.cell(row=top_row, column=1,
                         value=f"🏆 {sinf_key} — Top-3")
            th.font = Font(name="Arial", bold=True, size=11, color="1F4E79")
            th.alignment = Alignment(horizontal="left", vertical="center")
            top_row += 1
            for rank, t in enumerate(stat.get("top3", []), 1):
                ws.cell(row=top_row, column=1,
                        value=f"{rank}. {t['ismlar']}").font = normal_font
                ws.cell(row=top_row, column=2,
                        value=t["umumiy_ball"]).font = normal_font
                top_row += 1
            top_row += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        fname = f"taqq_{sinf_a}_{sinf_b}.xlsx".replace(" ", "_")
        wb.save(fname)

        await callback.bot.send_document(
            callback.from_user.id,
            FSInputFile(fname),
            caption=f"📊 {sinf_a} vs {sinf_b} — taqqoslash Excel hisoboti",
        )
        await wait.delete()
        if _os.path.exists(fname):
            _os.remove(fname)
        await callback.answer("✅ Excel yuborildi!", show_alert=True)

    except Exception as e:
        await wait.delete()
        await callback.answer(f"❌ Xatolik: {e}", show_alert=True)


# =====================================================================
# GOOGLE SHEETS EKSPORT — handler bloki
# =====================================================================

class SheetsStudentKod(StatesGroup):
    kod_kutish = State()


@router.message(F.text == "📗 Google Sheets Eksport")
async def sheets_hisobot_start(message: Message, state: FSMContext):
    """Google Sheets Eksport — asosiy menyu."""
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "📗 <b>Google Sheets Eksport</b>\n\nQaysi ma'lumotni eksport qilmoqchisiz?",
        parse_mode="HTML",
        reply_markup=sheets_export_keyboard(),
    )


@router.callback_query(F.data.startswith("sheets:"))
async def sheets_export_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    action = callback.data.split(":")[1]

    if action == "student":
        await state.set_state(SheetsStudentKod.kod_kutish)
        await callback.message.edit_text(
            "📗 Google Sheets ga yuklash uchun o'quvchi kodini kiriting:"
        )

    elif action == "all_students":
        wait_msg = await callback.message.answer("⏳ Google Sheets ga yuklanmoqda...")
        try:
            from sheets_export import export_all_students
            url = export_all_students()
            await wait_msg.delete()
            if url.startswith("❌"):
                await callback.message.answer(url)
            else:
                await callback.message.answer(
                    f"✅ <b>Barcha o'quvchilar Google Sheets ga yuklandi!</b>\n\n"
                    f"🔗 <a href='{url}'>Sheets ni ochish</a>",
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                )
        except Exception as e:
            await wait_msg.delete()
            await callback.message.answer(f"❌ Xatolik: {e}")

    elif action == "maktab_stat":
        from keyboards import maktab_tanlash_keyboard
        await callback.message.edit_text(
            "🏫 Maktab statistikasi uchun maktabni tanlang:",
            reply_markup=maktab_tanlash_keyboard(maktablar_ol(), "sheets_maktab"),
        )

    elif action == "sinf_reyting":
        from keyboards import sinf_tanlash_excel_keyboard as _sinf_kb
        # Sheets uchun alohida sinf callback ishlatamiz
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        from database import sinf_ol as _sinf_ol
        sinflar = _sinf_ol()
        buttons = [
            [InlineKeyboardButton(text=s, callback_data=f"sheets_sinf:{s}")]
            for s in sinflar
        ]
        buttons.append(
            [InlineKeyboardButton(text="📋 Barcha sinflar", callback_data="sheets_sinf:all")]
        )
        buttons.append(
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="sheets:menu")]
        )
        await callback.message.edit_text(
            "🏆 Sinf reytingi uchun sinfni tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
        )

    elif action == "menu":
        await callback.message.edit_text(
            "📗 <b>Google Sheets Eksport</b>\n\nQaysi ma'lumotni eksport qilmoqchisiz?",
            parse_mode="HTML",
            reply_markup=sheets_export_keyboard(),
        )

    await callback.answer()


@router.message(SheetsStudentKod.kod_kutish)
async def sheets_student_kod_handler(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return

    kod = message.text.strip().upper()
    from database import talaba_topish as _talaba_topish
    talaba = _talaba_topish(kod)

    if not talaba:
        await message.answer(
            f"❌ <code>{kod}</code> kodli o'quvchi topilmadi.",
            parse_mode="HTML",
        )
        await state.clear()
        return

    wait_msg = await message.answer("⏳ Google Sheets ga yuklanmoqda...")
    try:
        from sheets_export import export_student_history
        url = export_student_history(kod)
        await wait_msg.delete()
        if url.startswith("❌"):
            await message.answer(url)
        else:
            await message.answer(
                f"✅ <b>{talaba['ismlar']}</b> tarixi Google Sheets ga yuklandi!\n\n"
                f"🔗 <a href='{url}'>Sheets ni ochish</a>",
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
    except Exception as e:
        await wait_msg.delete()
        await message.answer(f"❌ Xatolik: {e}")

    await state.clear()


@router.callback_query(F.data.startswith("sheets_maktab:"))
async def sheets_maktab_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    maktab_id = int(callback.data.split(":")[1])
    wait_msg = await callback.message.answer("⏳ Google Sheets ga yuklanmoqda...")
    try:
        from sheets_export import export_maktab_statistika
        url = export_maktab_statistika(maktab_id)
        await wait_msg.delete()
        if url.startswith("❌"):
            await callback.message.answer(url)
        else:
            await callback.message.answer(
                f"✅ <b>Maktab statistikasi</b> Google Sheets ga yuklandi!\n\n"
                f"🔗 <a href='{url}'>Sheets ni ochish</a>",
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        await callback.answer("✅ Yuklandi!", show_alert=False)
    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {e}", show_alert=True)


@router.callback_query(F.data.startswith("sheets_sinf:"))
async def sheets_sinf_callback(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    sinf = callback.data.split(":")[1]
    wait_msg = await callback.message.answer("⏳ Google Sheets ga yuklanmoqda...")
    try:
        from sheets_export import export_sinf_reyting
        url = export_sinf_reyting(sinf=sinf if sinf != "all" else None)
        await wait_msg.delete()
        if url.startswith("❌"):
            await callback.message.answer(url)
        else:
            await callback.message.answer(
                f"✅ <b>Sinf reytingi</b> Google Sheets ga yuklandi!\n\n"
                f"🔗 <a href='{url}'>Sheets ni ochish</a>",
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
        await callback.answer("✅ Yuklandi!", show_alert=False)
    except Exception as e:
        await wait_msg.delete()
        await callback.answer(f"❌ Xatolik: {e}", show_alert=True)
