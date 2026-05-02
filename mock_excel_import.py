"""
Mock imtihon natijalari — Excel import moduli.

Qo'llab-quvvatlanadigan varaqlar:
  DTM_MOCK     — Majburiy, Asosiy 1, Asosiy 2
  IELTS        — Listening, Reading, Writing, Speaking
  CEFR         — Listening, Reading, Writing, Speaking + Daraja
  SAT          — Reading&Writing, Math
  MILLIY_SERT  — Fan nomi + Ball
  BOSHQA       — Ixtiyoriy imtihon (Bo'limlar nomi + qiymati)

Foydalanish:
  from mock_excel_import import router
  dp.include_router(router)
"""

import io
import asyncio
import logging
from aiogram import Router, F
from aiogram.types import Message, Document
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

logger = logging.getLogger(__name__)
router = Router()


class MockExcelImport(StatesGroup):
    fayl_kutish = State()


# ── Ustun xaritalari ─────────────────────────────────────────────────────────

SHEET_CONFIGS = {
    "DTM_MOCK": {
        "exam_key": "DTM_MOCK",
        "required_cols": ["O'quvchi Kodi"],
        "section_cols": [
            ("Majburiy Fanlar", "majburiy", 30.0),
            ("1-Asosiy Fan",    "asosiy_1", 30.0),
            ("2-Asosiy Fan",    "asosiy_2", 30.0),
        ],
        "level_col":   None,
        "subject_col": None,
        "notes_col":   "Izoh",
    },
    "IELTS": {
        "exam_key": "IELTS",
        "required_cols": ["O'quvchi Kodi"],
        "section_cols": [
            ("Listening", "listening", 9.0),
            ("Reading",   "reading",   9.0),
            ("Writing",   "writing",   9.0),
            ("Speaking",  "speaking",  9.0),
        ],
        "level_col":   "Daraja",
        "subject_col": None,
        "notes_col":   "Izoh",
    },
    "CEFR": {
        "exam_key": "CEFR",
        "required_cols": ["O'quvchi Kodi"],
        "section_cols": [
            ("Listening", "listening", 75.0),
            ("Reading",   "reading",   75.0),
            ("Writing",   "writing",   75.0),
            ("Speaking",  "speaking",  75.0),
        ],
        "level_col":   "Daraja",
        "subject_col": None,
        "notes_col":   "Izoh",
    },
    "SAT": {
        "exam_key": "SAT",
        "required_cols": ["O'quvchi Kodi"],
        "section_cols": [
            ("Reading & Writing", "reading_writing", 800.0),
            ("Math",               "math",            800.0),
        ],
        "level_col":   None,
        "subject_col": None,
        "notes_col":   "Izoh",
    },
    "MILLIY_SERT": {
        "exam_key": "MILLIY_SERT",
        "required_cols": ["O'quvchi Kodi", "Fan nomi"],
        "section_cols": [
            ("Ball", "ball", 100.0),
        ],
        "level_col":   None,
        "subject_col": "Fan nomi",
        "notes_col":   "Izoh",
    },
    "BOSHQA": {
        "exam_key": None,   # Excel dagi "Imtihon Kaliti" ustunidan olinadi
        "required_cols": ["O'quvchi Kodi", "Imtihon Kaliti", "Bo'lim 1 nomi", "Bo'lim 1 qiymati"],
        "section_cols":  None,   # Dinamik
        "level_col":     "Daraja",
        "subject_col":   "Fan nomi",
        "notes_col":     "Izoh",
    },
}

# Ustun sarlavhalarini normalize qilish uchun (qisman moslik)
def _find_col(headers: list, keyword: str) -> int | None:
    """Sarlavhalar ichida kalit so'z bo'yicha ustun indeksini topadi."""
    kw = keyword.lower().strip()
    for i, h in enumerate(headers):
        if h and kw in str(h).lower().strip():
            return i
    return None


def _safe_float(val) -> float | None:
    if val is None or str(val).strip() in ("", "-", "—"):
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


# ── Excel o'qish va natijalarni import qilish ────────────────────────────────

def _parse_sheet(sheet, sheet_name: str) -> tuple[list[dict], list[str]]:
    """
    Varaqni o'qib, (muvaffaqiyatli_qatorlar, xatolar) qaytaradi.
    Har bir muvaffaqiyatli qator:
      {talaba_kod, exam_key, sections, level, subject_name, notes}
    """
    from database import talaba_topish
    from mock_database import exam_type_ol

    cfg = SHEET_CONFIGS.get(sheet_name)
    if cfg is None:
        return [], []   # Noma'lum varaq — o'tkazib yuboriladi

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], [f"[{sheet_name}] Yetarli ma'lumot yo'q (sarlavha + kamida 1 qator kerak)"]

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    results, errors = [], []

    # Kod ustunini topish
    kod_idx = _find_col(headers, "o'quvchi kodi") 
    if kod_idx is None:
        kod_idx = _find_col(headers, "kod")
    if kod_idx is None:
        errors.append(f"[{sheet_name}] \"O'quvchi Kodi\" ustuni topilmadi")
        return [], errors

    for row_num, row in enumerate(rows[1:], start=2):
        # Bo'sh qatorni o'tkazib yuborish
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        kod_raw = row[kod_idx] if kod_idx < len(row) else None
        if not kod_raw or str(kod_raw).strip() == "":
            continue

        kod = str(kod_raw).strip().upper()
        talaba = talaba_topish(kod)
        if not talaba:
            errors.append(f"[{sheet_name}] Qator {row_num}: '{kod}' kodi topilmadi")
            continue

        # exam_key
        if cfg["exam_key"]:
            exam_key = cfg["exam_key"]
        else:
            # BOSHQA varaqda "Imtihon Kaliti" ustunidan olish
            ek_idx = _find_col(headers, "imtihon kaliti")
            if ek_idx is None:
                errors.append(f"[{sheet_name}] Qator {row_num}: 'Imtihon Kaliti' ustuni topilmadi")
                continue
            exam_key = str(row[ek_idx]).strip().upper() if ek_idx < len(row) and row[ek_idx] else ""
            if not exam_key:
                errors.append(f"[{sheet_name}] Qator {row_num}: Imtihon kaliti bo'sh")
                continue

        # Sections
        if cfg["section_cols"] is not None:
            sections = {}
            skip = False
            for (col_name, sec_key, max_val) in cfg["section_cols"]:
                col_idx = _find_col(headers, col_name.split("(")[0].strip())
                if col_idx is None:
                    errors.append(f"[{sheet_name}] Qator {row_num}: '{col_name}' ustuni topilmadi")
                    skip = True
                    break
                val = _safe_float(row[col_idx] if col_idx < len(row) else None)
                if val is None:
                    errors.append(f"[{sheet_name}] Qator {row_num}: '{col_name}' bo'sh yoki noto'g'ri ({row[col_idx] if col_idx < len(row) else '?'})")
                    skip = True
                    break
                if max_val is not None:
                    # DTM_MOCK uchun 30 dan katta qiymatga ruxsat (tayyor ball)
                    if exam_key != "DTM_MOCK" and val > max_val:
                        errors.append(f"[{sheet_name}] Qator {row_num}: '{col_name}' = {val}, max {max_val}")
                        skip = True
                        break
                sections[sec_key] = {"label": col_name, "value": val, "max": max_val}
            if skip:
                continue
        else:
            # BOSHQA — dinamik bo'limlar (Bo'lim 1 nomi / qiymati, Bo'lim 2 nomi / qiymati, ...)
            sections = {}
            for pair_num in range(1, 6):   # maksimal 5 juft
                name_col = f"bo'lim {pair_num} nomi"
                val_col  = f"bo'lim {pair_num} qiymati"
                n_idx = _find_col(headers, name_col)
                v_idx = _find_col(headers, val_col)
                if n_idx is None or v_idx is None:
                    break
                sec_name = row[n_idx] if n_idx < len(row) else None
                sec_val  = _safe_float(row[v_idx] if v_idx < len(row) else None)
                if not sec_name or str(sec_name).strip() == "":
                    break
                if sec_val is None:
                    break
                import re
                sec_key = re.sub(r"[^a-z0-9_]", "_", str(sec_name).lower())[:30]
                sections[sec_key] = {"label": str(sec_name).strip(), "value": sec_val, "max": None}

            if not sections:
                errors.append(f"[{sheet_name}] Qator {row_num}: Bo'lim ma'lumotlari topilmadi")
                continue

        # Level
        level = None
        if cfg["level_col"]:
            lv_idx = _find_col(headers, cfg["level_col"].lower())
            if lv_idx is not None and lv_idx < len(row):
                lv = row[lv_idx]
                if lv and str(lv).strip() not in ("", "-", "—"):
                    level = str(lv).strip()

        # Subject
        subject_name = None
        if cfg["subject_col"]:
            sub_idx = _find_col(headers, cfg["subject_col"].lower().split("(")[0].strip())
            if sub_idx is not None and sub_idx < len(row):
                sv = row[sub_idx]
                if sv and str(sv).strip() not in ("", "-", "—"):
                    subject_name = str(sv).strip()

        # Notes
        notes = None
        if cfg["notes_col"]:
            nt_idx = _find_col(headers, cfg["notes_col"].lower())
            if nt_idx is not None and nt_idx < len(row):
                nv = row[nt_idx]
                if nv and str(nv).strip() not in ("", "-", "—"):
                    notes = str(nv).strip()

        results.append({
            "talaba_kod":   kod,
            "ismlar":       talaba.get("ismlar", kod),
            "exam_key":     exam_key,
            "sections":     sections,
            "level":        level,
            "subject_name": subject_name,
            "notes":        notes,
        })

    return results, errors


def import_excel(file_bytes: bytes, admin_id: int) -> dict:
    """
    Excel faylni o'qib mock natijalarini bazaga saqlaydi.
    Qaytaradi: {
        "saved": int,
        "skipped": int,
        "errors": [str],
        "details": [(talaba_kod, ismlar, exam_key), ...]
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"saved": 0, "skipped": 0, "errors": ["openpyxl o'rnatilmagan"], "details": []}

    from mock_database import mock_natija_qosh

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"saved": 0, "skipped": 0, "errors": [f"Excel faylni o'qib bo'lmadi: {e}"], "details": []}

    all_results, all_errors = [], []

    for sheet_name in wb.sheetnames:
        clean_name = sheet_name.strip().lstrip("📋").lstrip("📥").strip()
        # "Ko'rsatmalar" varaqni o'tkazib yuborish
        if "ko'rsatma" in clean_name.lower() or "instruction" in clean_name.lower():
            continue
        ws = wb[sheet_name]
        res, errs = _parse_sheet(ws, clean_name)
        all_results.extend(res)
        all_errors.extend(errs)

    saved, skipped, details = 0, 0, []
    for item in all_results:
        try:
            new_id = mock_natija_qosh(
                talaba_kod=item["talaba_kod"],
                exam_key=item["exam_key"],
                sections=item["sections"],
                level_label=item.get("level"),
                subject_name=item.get("subject_name"),
                notes=item.get("notes"),
                admin_id=admin_id,
            )
            if new_id:
                saved += 1
                details.append((item["talaba_kod"], item["ismlar"], item["exam_key"]))
            else:
                skipped += 1
                all_errors.append(f"{item['talaba_kod']} — saqlashda xato")
        except Exception as e:
            skipped += 1
            all_errors.append(f"{item['talaba_kod']} — {e}")

    return {
        "saved":   saved,
        "skipped": skipped,
        "errors":  all_errors,
        "details": details,
    }


# ── Telegram handlerlari ─────────────────────────────────────────────────────

@router.message(F.text == "📥 Excel import")
async def excel_import_start(message: Message, state: FSMContext):
    from admin import is_admin_id
    if not is_admin_id(message.from_user.id):
        return
    await state.set_state(MockExcelImport.fayl_kutish)
    await message.answer(
        "📥 <b>Mock natijalar — Excel import</b>\n\n"
        "Excel faylni yuboring (.xlsx).\n\n"
        "Shablon olish uchun: <b>📄 Shablon yuklab olish</b> tugmasini bosing.",
        parse_mode="HTML",
    )


@router.message(F.text == "📄 Shablon yuklab olish")
async def shablon_yuborish(message: Message, state: FSMContext):
    from admin import is_admin_id
    if not is_admin_id(message.from_user.id):
        return
    import os
    shablon_path = os.path.join(os.path.dirname(__file__), "mock_natijalar_shablon.xlsx")
    if not os.path.exists(shablon_path):
        await message.answer("❌ Shablon fayl topilmadi. Admin bilan bog'laning.")
        return
    from aiogram.types import FSInputFile
    await message.answer_document(
        FSInputFile(shablon_path),
        caption=(
            "📄 <b>Mock natijalar shablon fayli</b>\n\n"
            "1. Kerakli varaqni to'ldiring\n"
            "2. .xlsx formatida saqlang\n"
            "3. <b>📥 Excel import</b> tugmasini bosib faylni yuboring"
        ),
        parse_mode="HTML",
    )


@router.message(MockExcelImport.fayl_kutish, F.document)
async def excel_import_fayl(message: Message, state: FSMContext):
    from admin import is_admin_id
    if not is_admin_id(message.from_user.id):
        return

    doc: Document = message.document
    if not doc.file_name or not doc.file_name.endswith(".xlsx"):
        await message.answer("❌ Faqat <b>.xlsx</b> formatdagi fayl qabul qilinadi.", parse_mode="HTML")
        return

    wait_msg = await message.answer("⏳ Fayl o'qilmoqda...")

    try:
        file = await message.bot.get_file(doc.file_id)
        buf = io.BytesIO()
        await message.bot.download_file(file.file_path, buf)
        file_bytes = buf.getvalue()
    except Exception as e:
        await wait_msg.edit_text(f"❌ Faylni yuklab olishda xato: {e}")
        await state.clear()
        return

    # Import (CPU-bound — thread poolda ishlatish)
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None, import_excel, file_bytes, message.from_user.id
    )

    # Natija xabari
    saved   = result["saved"]
    skipped = result["skipped"]
    errors  = result["errors"]
    details = result["details"]

    lines = [
        f"📊 <b>Import yakunlandi</b>",
        f"",
        f"✅ Muvaffaqiyatli saqlandi: <b>{saved} ta</b>",
        f"⚠️ O'tkazib yuborildi:       <b>{skipped} ta</b>",
    ]

    if details:
        lines.append("\n<b>Saqlangan natijalar:</b>")
        for (kod, ism, ek) in details[:20]:
            lines.append(f"  • {ism} ({kod}) — {ek}")
        if len(details) > 20:
            lines.append(f"  <i>...va yana {len(details) - 20} ta</i>")

    if errors:
        lines.append(f"\n<b>⚠️ Xatolar ({len(errors)} ta):</b>")
        for err in errors[:15]:
            lines.append(f"  ❌ {err}")
        if len(errors) > 15:
            lines.append(f"  <i>...va yana {len(errors) - 15} ta xato</i>")

    await wait_msg.edit_text("\n".join(lines), parse_mode="HTML")
    await state.clear()

    # O'quvchilarga bildirishnoma yuborish (fon rejimda)
    if details:
        asyncio.create_task(_xabar_barchasiga(message.bot, [d[0] for d in details]))


async def _xabar_barchasiga(bot, talaba_kodlar: list):
    """Import qilingan barcha o'quvchilarga bildirishnoma yuboradi."""
    from database import talaba_topish, is_notification_enabled
    from mock_database import mock_natijalari_ol, format_mock_natija_matn
    import asyncio

    for kod in talaba_kodlar:
        try:
            talaba = talaba_topish(kod)
            if not talaba or not talaba.get("user_id"):
                continue
            if not is_notification_enabled(talaba["user_id"], "notify_mock_results"):
                continue
            natijalar = mock_natijalari_ol(kod, limit=1)
            if not natijalar:
                continue
            matn = "✅ <b>Yangi mock natija qo'shildi!</b>\n\n" + format_mock_natija_matn(natijalar[0])
            await bot.send_message(talaba["user_id"], matn, parse_mode="HTML")
            await asyncio.sleep(0.05)   # flood limit
        except Exception:
            pass
