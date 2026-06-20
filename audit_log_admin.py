"""
audit_log_admin.py — Audit jurnalini Telegram admin panelida ko'rsatish.

Menyu: "🗂 Audit jurnali" (admin_menu_keyboard ga qo'shilgan)
  ├── 🔐 Admin amallari      — faqat admin/o'qituvchi/maktab-admin harakatlari
  ├── 👥 Foydalanuvchi amallari — o'quvchi/ota-ona harakatlari
  ├── 📋 Hammasi              — ikkalasi birga, xronologik
  └── 📤 Excel eksport        — joriy filtr bo'yicha to'liq .xlsx fayl

Super admin (ADMIN_IDS) — butun tizim bo'yicha ko'radi.
Maktab admin (franchise) — faqat o'z maktabiga tegishli yozuvlarni ko'radi.
"""

import logging
import os
from datetime import datetime

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext

from admin import admin_tekshir, is_admin_id
from audit_log import (
    get_audit_logs,
    get_audit_log_count,
    format_log_entry,
    ROLE_LABELS,
    ACTION_LABELS,
)

logger = logging.getLogger(__name__)
router = Router()

PAGE_SIZE = 10


def _maktab_filter_for(user_id: int) -> int | None:
    """Super admin uchun None (cheklovsiz), maktab admin uchun maktab_id."""
    if is_admin_id(user_id):
        return None
    try:
        from payment import get_maktab_by_admin
        maktab = get_maktab_by_admin(user_id)
        return maktab["id"] if maktab else None
    except Exception:
        return None


def _audit_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔐 Admin amallari", callback_data="audit_view:admin:0")],
            [InlineKeyboardButton(text="👥 Foydalanuvchi amallari", callback_data="audit_view:user:0")],
            [InlineKeyboardButton(text="📋 Hammasi", callback_data="audit_view:all:0")],
            [InlineKeyboardButton(text="📤 Excel eksport", callback_data="audit_export:all")],
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="audit:orqaga")],
        ]
    )


def _pagination_keyboard(scope: str, page: int, total: int) -> InlineKeyboardMarkup:
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton(text="⬅️ Oldingi", callback_data=f"audit_view:{scope}:{page-1}"))
    nav_row.append(InlineKeyboardButton(text=f"{page+1}/{total_pages}", callback_data="no_action"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton(text="Keyingi ➡️", callback_data=f"audit_view:{scope}:{page+1}"))

    rows = [nav_row] if nav_row else []
    rows.append([InlineKeyboardButton(text="📤 Shu bo'limni Excelga", callback_data=f"audit_export:{scope}")])
    rows.append([InlineKeyboardButton(text="🔙 Audit menyusi", callback_data="audit:menu")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _scope_to_group(scope: str) -> str | None:
    return {"admin": "admin", "user": "user", "all": None}.get(scope)


@router.message(F.text == "🗂 Audit jurnali")
async def audit_menu_open(message: Message, state: FSMContext):
    if not await admin_tekshir(state, message.from_user.id):
        return
    await message.answer(
        "🗂 <b>Audit jurnali</b>\n\n"
        "Bu yerda tizimda bajarilgan barcha amallar — admin harakatlari va "
        "oddiy foydalanuvchilar faoliyati — xronologik tartibda saqlanadi.\n\n"
        "Bo'limni tanlang:",
        parse_mode="HTML",
        reply_markup=_audit_menu_keyboard(),
    )


@router.callback_query(F.data == "audit:menu")
async def audit_menu_back(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return
    await callback.message.edit_text(
        "🗂 <b>Audit jurnali</b>\n\nBo'limni tanlang:",
        parse_mode="HTML",
        reply_markup=_audit_menu_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "audit:orqaga")
async def audit_menu_close(callback: CallbackQuery, state: FSMContext):
    await callback.message.delete()
    await callback.answer()


@router.callback_query(F.data.startswith("audit_view:"))
async def audit_view_page(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    _, scope, page_str = callback.data.split(":")
    page = int(page_str)
    group = _scope_to_group(scope)
    maktab_id = _maktab_filter_for(callback.from_user.id)

    total = get_audit_log_count(action_group=group, maktab_id=maktab_id)
    entries = get_audit_logs(
        action_group=group,
        maktab_id=maktab_id,
        limit=PAGE_SIZE,
        offset=page * PAGE_SIZE,
    )

    if not entries:
        text = "📭 Bu bo'limda hozircha hech qanday yozuv yo'q."
    else:
        scope_title = {"admin": "🔐 Admin amallari", "user": "👥 Foydalanuvchi amallari", "all": "📋 Barcha amallar"}[scope]
        blocks = [format_log_entry(e) for e in entries]
        text = f"<b>{scope_title}</b> ({total} ta yozuv)\n\n" + "\n\n".join(blocks)

    try:
        await callback.message.edit_text(
            text, parse_mode="HTML", reply_markup=_pagination_keyboard(scope, page, total)
        )
    except Exception:
        # Matn juda uzun bo'lsa yoki o'zgarmagan bo'lsa
        await callback.message.answer(
            text, parse_mode="HTML", reply_markup=_pagination_keyboard(scope, page, total)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("audit_export:"))
async def audit_export_excel(callback: CallbackQuery, state: FSMContext):
    if not await admin_tekshir(state, callback.from_user.id):
        return

    scope = callback.data.split(":")[1]
    group = _scope_to_group(scope)
    maktab_id = _maktab_filter_for(callback.from_user.id)

    await callback.answer("⏳ Excel fayl tayyorlanmoqda...")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    total = get_audit_log_count(action_group=group, maktab_id=maktab_id)
    if total == 0:
        await callback.message.answer("📭 Eksport qilish uchun yozuv topilmadi.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit jurnali"

    headers = ["Sana/vaqt", "Rol", "Ism", "Telegram ID", "Amal", "Obyekt", "Tafsilot"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    CHUNK = 500
    offset = 0
    while offset < total:
        batch = get_audit_logs(
            action_group=group, maktab_id=maktab_id, limit=CHUNK, offset=offset
        )
        for e in batch:
            label, _ = ACTION_LABELS.get(e["action_type"], (e["action_type"], "user"))
            role_label = ROLE_LABELS.get(e["actor_role"], e["actor_role"])
            sana = e["created_at"].strftime("%d.%m.%Y %H:%M") if e.get("created_at") else ""
            ws.append([
                sana,
                role_label,
                e.get("actor_name") or "",
                e.get("actor_id") or "",
                label,
                e.get("target") or "",
                e.get("details") or "",
            ])
        offset += CHUNK

    for i, width in enumerate([16, 16, 22, 14, 32, 18, 36], start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    os.makedirs("exports", exist_ok=True)
    fname = f"exports/audit_log_{scope}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)

    await callback.message.answer_document(
        FSInputFile(fname),
        caption=f"📤 Audit jurnali eksporti ({total} ta yozuv)",
    )

    try:
        os.remove(fname)
    except Exception:
        pass
