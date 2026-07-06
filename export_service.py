"""
services/export_service.py
--------------------------
O'quvchining barcha natijalari uchun:
  - Har bir mock uchun PDF hisobot
  - Umumiy Excel jadval
  - Hammasini ZIP arxivga jamlash
"""

import io
import zipfile
from datetime import datetime
from typing import Optional

import asyncpg
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────
#  DB dan ma'lumot olish
# ─────────────────────────────────────────────────

async def fetch_student_data(pool: asyncpg.Pool, telegram_id: int) -> dict:
    """O'quvchining barcha natijalarini DB dan tortib oladi."""
    async with pool.acquire() as conn:
        # Asosiy ma'lumot
        student = await conn.fetchrow(
            """
            SELECT id, full_name, sinf, yonalish, created_at
            FROM students
            WHERE telegram_id = $1
            """,
            telegram_id,
        )
        if not student:
            return {}

        # Mock imtihon natijalari
        mock_results = await conn.fetch(
            """
            SELECT
                exam_type,
                total_score,
                band_score,
                listening_score,
                reading_score,
                writing_score,
                speaking_score,
                created_at,
                certificate_code
            FROM mock_results
            WHERE student_id = $1
            ORDER BY created_at DESC
            """,
            student["id"],
        )

        # Mini-test natijalari (so'nggi 30 ta)
        mini_results = await conn.fetch(
            """
            SELECT subject, correct_count, total_count, percentage, created_at
            FROM test_results
            WHERE student_id = $1
            ORDER BY created_at DESC
            LIMIT 30
            """,
            student["id"],
        )

    return {
        "student":      dict(student),
        "mock_results": [dict(r) for r in mock_results],
        "mini_results": [dict(r) for r in mini_results],
    }


# ─────────────────────────────────────────────────
#  PDF — har bir mock uchun alohida hisobot
# ─────────────────────────────────────────────────

class ResultPDF(FPDF):
    """Loyihadagi fpdf2 uslubiga mos PDF sinfi."""

    HEADER_COLOR = (30, 58, 138)   # ko'k
    ACCENT_COLOR = (59, 130, 246)
    LIGHT_BG     = (239, 246, 255)
    TEXT_DARK    = (15, 23, 42)
    TEXT_MUTED   = (100, 116, 139)

    def header(self):
        self.set_fill_color(*self.HEADER_COLOR)
        self.rect(0, 0, 210, 28, "F")
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(255, 255, 255)
        self.set_xy(0, 8)
        self.cell(210, 10, "BUSTANLIK SS TESTING SYSTEM", align="C")
        self.set_font("Helvetica", "", 9)
        self.set_xy(0, 18)
        self.cell(210, 8, "Natija hisoboti / Result Report", align="C")
        self.set_text_color(*self.TEXT_DARK)
        self.ln(10)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(*self.TEXT_MUTED)
        self.cell(0, 10, f"Sahifa {self.page_no()} | bustanlik.school", align="C")

    def section_title(self, title: str):
        self.set_font("Helvetica", "B", 10)
        self.set_fill_color(*self.LIGHT_BG)
        self.set_text_color(*self.ACCENT_COLOR)
        self.cell(0, 8, f"  {title}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
        self.set_text_color(*self.TEXT_DARK)
        self.ln(2)

    def info_row(self, label: str, value: str, bold_value=False):
        self.set_font("Helvetica", "", 10)
        self.set_text_color(*self.TEXT_MUTED)
        self.cell(55, 7, label + ":")
        self.set_font("Helvetica", "B" if bold_value else "", 10)
        self.set_text_color(*self.TEXT_DARK)
        self.cell(0, 7, value, new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def _band_color(band: Optional[float]) -> tuple:
    """Band skori asosida rang qaytaradi."""
    if band is None:
        return (107, 114, 128)
    if band >= 7.0:
        return (22, 163, 74)    # yashil
    if band >= 5.5:
        return (234, 179, 8)    # sariq
    return (220, 38, 38)        # qizil


def generate_mock_pdf(result: dict, student: dict) -> bytes:
    """Bitta mock imtihon natijasi uchun PDF hosil qiladi."""
    pdf = ResultPDF()
    pdf.add_page()
    pdf.set_margins(15, 10, 15)
    pdf.set_auto_page_break(auto=True, margin=15)

    exam_date = result["created_at"].strftime("%d.%m.%Y %H:%M") if result.get("created_at") else "—"
    exam_type = result.get("exam_type", "Noma'lum").upper()

    # ── O'quvchi ma'lumotlari ─────────────────────
    pdf.section_title("O'quvchi ma'lumotlari")
    pdf.info_row("Ism Familiya", student.get("full_name", "—"))
    pdf.info_row("Sinf",         student.get("sinf", "—"))
    pdf.info_row("Yo'nalish",    student.get("yonalish", "—"))
    pdf.info_row("Sana",         exam_date)
    pdf.ln(4)

    # ── Umumiy ball ──────────────────────────────
    pdf.section_title(f"{exam_type} — Umumiy natija")

    band = result.get("band_score")
    total = result.get("total_score")

    pdf.set_fill_color(*_band_color(band))
    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(60, 18, str(band or total or "—"), align="C", fill=True,
             new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*ResultPDF.TEXT_MUTED)
    pdf.set_xy(pdf.get_x() + 4, pdf.get_y() + 4)
    pdf.cell(0, 8, "Band / Umumiy ball", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    # ── Sektsiyalar ───────────────────────────────
    sections = [
        ("Listening", result.get("listening_score")),
        ("Reading",   result.get("reading_score")),
        ("Writing",   result.get("writing_score")),
        ("Speaking",  result.get("speaking_score")),
    ]
    has_sections = any(v is not None for _, v in sections)

    if has_sections:
        pdf.section_title("Sektsiyalar bo'yicha natijalar")
        col_w = (210 - 30) / 4
        for label, score in sections:
            if score is None:
                continue
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(*ResultPDF.TEXT_MUTED)
            pdf.cell(col_w, 6, label, align="C")
        pdf.ln()

        for _, score in sections:
            if score is None:
                continue
            color = _band_color(score if isinstance(score, float) else None)
            pdf.set_fill_color(*color)
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(col_w, 14, str(score), align="C", fill=True)
        pdf.ln(18)

    # ── Sertifikat kodi ───────────────────────────
    code = result.get("certificate_code")
    if code:
        pdf.section_title("Sertifikat")
        pdf.info_row("Tekshirish kodi", code, bold_value=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*ResultPDF.TEXT_MUTED)
        pdf.cell(0, 6,
                 "Tekshirish: https://bustanlik.school/verify/" + code,
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


# ─────────────────────────────────────────────────
#  Excel — barcha natijalar jadvali
# ─────────────────────────────────────────────────

def _xl_header_style() -> tuple:
    font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fill  = PatternFill("solid", fgColor="1E3A8A")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _xl_border() -> Border:
    thin = Side(style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_excel(data: dict) -> bytes:
    """Barcha natijalar uchun Excel fayl hosil qiladi."""
    wb = Workbook()
    student = data["student"]
    h_font, h_fill, h_align = _xl_header_style()
    border = _xl_border()

    # ── 1-varaq: Mock natijalari ──────────────────
    ws1 = wb.active
    ws1.title = "Mock natijalari"

    headers1 = [
        "Sana", "Imtihon turi", "Umumiy ball", "Band",
        "Listening", "Reading", "Writing", "Speaking", "Sertifikat kodi"
    ]
    widths1 = [16, 16, 14, 8, 12, 12, 12, 12, 22]

    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = h_align
        cell.border = border
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 28

    for row_i, r in enumerate(data["mock_results"], 2):
        row_data = [
            r["created_at"].strftime("%d.%m.%Y") if r.get("created_at") else "",
            r.get("exam_type", ""),
            r.get("total_score", ""),
            r.get("band_score", ""),
            r.get("listening_score", ""),
            r.get("reading_score", ""),
            r.get("writing_score", ""),
            r.get("speaking_score", ""),
            r.get("certificate_code", ""),
        ]
        bg = "F0F9FF" if row_i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_i, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = border

    # ── 2-varaq: Mini-test natijalari ─────────────
    ws2 = wb.create_sheet("Mini-testlar")

    headers2 = ["Sana", "Fan", "To'g'ri", "Jami", "Foiz (%)"]
    widths2   = [16, 20, 10, 10, 12]

    for col, (h, w) in enumerate(zip(headers2, widths2), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = h_align
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 28

    for row_i, r in enumerate(data["mini_results"], 2):
        row_data = [
            r["created_at"].strftime("%d.%m.%Y") if r.get("created_at") else "",
            r.get("subject", ""),
            r.get("correct_count", ""),
            r.get("total_count", ""),
            round(r.get("percentage", 0), 1),
        ]
        bg = "F0FDF4" if row_i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_i, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = border

    # ── 3-varaq: O'quvchi ma'lumoti ───────────────
    ws3 = wb.create_sheet("Profil")
    info = [
        ("Ism Familiya",       student.get("full_name", "")),
        ("Sinf",               student.get("sinf", "")),
        ("Yo'nalish",          student.get("yonalish", "")),
        ("Ro'yxatdan o'tgan",  student.get("created_at", "").strftime("%d.%m.%Y")
                               if hasattr(student.get("created_at"), "strftime") else ""),
        ("Jami mock natijalar", len(data["mock_results"])),
        ("Jami mini-testlar",   len(data["mini_results"])),
        ("Hisobot sanasi",     datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 30

    for row_i, (label, value) in enumerate(info, 1):
        a = ws3.cell(row=row_i, column=1, value=label)
        b = ws3.cell(row=row_i, column=2, value=value)
        a.font  = Font(bold=True, color="1E3A8A")
        a.fill  = PatternFill("solid", fgColor="EFF6FF")
        a.border = border
        b.border = border
        a.alignment = Alignment(vertical="center")
        b.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────
#  ZIP — hammasini jamlash
# ─────────────────────────────────────────────────

def build_zip_archive(data: dict) -> io.BytesIO:
    """
    PDF + Excel fayllarini bitta ZIP arxivga yig'adi.
    Qaytaradi: BytesIO ob'ekti (Telegram uchun tayyor)
    """
    student    = data["student"]
    safe_name  = (student.get("full_name") or "student").replace(" ", "_")
    date_str   = datetime.now().strftime("%Y%m%d")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:

        # Excel jadval
        excel_bytes = generate_excel(data)
        zf.writestr(f"{safe_name}_natijalar_{date_str}.xlsx", excel_bytes)

        # Har bir mock uchun alohida PDF
        for i, result in enumerate(data["mock_results"], 1):
            exam_type = result.get("exam_type", "exam").replace(" ", "_")
            exam_date = (result["created_at"].strftime("%Y%m%d")
                         if result.get("created_at") else f"{i:02d}")
            pdf_bytes = generate_mock_pdf(result, student)
            zf.writestr(
                f"pdf/{exam_type}_{exam_date}.pdf",
                pdf_bytes
            )

        # README.txt
        readme = (
            f"Bustanlik SS Testing System — Natijalar arxivi\n"
            f"{'='*50}\n\n"
            f"O'quvchi : {student.get('full_name', '—')}\n"
            f"Sinf     : {student.get('sinf', '—')}\n"
            f"Yo'nalish: {student.get('yonalish', '—')}\n"
            f"Sana     : {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
            f"Tarkib:\n"
            f"  /{safe_name}_natijalar_{date_str}.xlsx  — barcha natijalar jadvali\n"
            f"  /pdf/  — har bir mock uchun alohida PDF hisobot\n\n"
            f"Sertifikatni tekshirish: https://bustanlik.school/verify/<kod>\n"
        )
        zf.writestr("README.txt", readme.encode("utf-8"))

    zip_buf.seek(0)
    return zip_buf
