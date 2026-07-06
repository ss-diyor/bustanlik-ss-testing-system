"""
Mock imtihon natijalari — Excel eksport moduli.

Eksport turlari:
  1. Fan bo'yicha   — bitta exam_key bo'yicha barcha natijalar
  2. O'quvchi bo'yicha — bitta talabaning barcha mock natijalari
  3. Umumiy (barcha) — barcha fanlar, bitta faylda har biri alohida varaq

Foydalanish (mock_admin.py ga ulash):
  from mock_excel_export import router as mock_export_router
  dp.include_router(mock_export_router)

Menyu tugmalari (mavjud mock menyusiga qo'shish):
  "📤 Mock eksport"
"""

import io
import os
import asyncio
import logging
from datetime import datetime

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
router = Router()


# ════════════════════════════════════════════════════════════
# STIL KONSTANTALAR
# ════════════════════════════════════════════════════════════

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_SUBHEAD_FILL = PatternFill("solid", fgColor="17375E")
_ALT_FILL     = PatternFill("solid", fgColor="DEEAF1")
_TOTAL_FILL   = PatternFill("solid", fgColor="BDD7EE")
_LEVEL_FILL   = PatternFill("solid", fgColor="E2EFDA")  # yashil — daraja ustuni

_WHITE_FONT   = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
_BOLD_FONT    = Font(name="Arial", bold=True,  color="000000", size=10)
_NORMAL_FONT  = Font(name="Arial", bold=False, color="000000", size=10)
_ITALIC_FONT  = Font(name="Arial", italic=True, color="595959", size=9)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hcell(ws, row, col, val, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _WHITE_FONT
    c.fill      = fill or _HEADER_FILL
    c.alignment = _CENTER
    c.border    = _thin()
    return c

def _dcell(ws, row, col, val, alt=False, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _NORMAL_FONT
    c.fill      = _ALT_FILL if alt else PatternFill()
    c.alignment = align or _CENTER
    c.border    = _thin()
    return c

def _footer(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _ITALIC_FONT
    c.alignment = _LEFT

def _sana_str(val):
    if val is None:
        return "—"
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    return str(val)[:10]

def _val(v):
    """sections dict ichidan qiymatni olish — {value:...} yoki to'g'ridan-to'g'ri."""
    if isinstance(v, dict):
        return v.get("value")
    return v

def _now_str():
    try:
        from tz_utils import now as tz_now
        return tz_now().strftime("%d.%m.%Y %H:%M")
    except Exception:
        return datetime.now().strftime("%d.%m.%Y %H:%M")


# ════════════════════════════════════════════════════════════
# EKSPORT FUNKSIYALARI
# ════════════════════════════════════════════════════════════

def export_by_exam(exam_key: str) -> bytes | None:
    """
    Bitta fan (exam_key) bo'yicha barcha natijalarni Excel ga eksport qiladi.
    Natijalar sana bo'yicha kamayish tartibida. Limit yo'q.
    """
    from mock_database import mock_hammasi_by_exam, exam_type_ol

    # Limitni olib tashlaymiz — barcha natijalarni olamiz
    rows = mock_hammasi_by_exam(exam_key, limit=9999, offset=0)
    if not rows:
        return None

    et    = exam_type_ol(exam_key) or {}
    label = et.get("label", exam_key)

    # Sections ustunlarini aniqlaymiz (birinchi natijadan)
    sec_keys = []
    for r in rows:
        secs = r.get("sections") or {}
        for k in secs:
            if k not in sec_keys:
                sec_keys.append(k)

    # Har bir section label ni topamiz
    sec_labels = {}
    for r in rows:
        secs = r.get("sections") or {}
        for k, v in secs.items():
            if k not in sec_labels:
                if isinstance(v, dict) and v.get("label"):
                    sec_labels[k] = v["label"]
                else:
                    sec_labels[k] = k.replace("_", " ").title()

    has_level   = any(r.get("level_label")  for r in rows)
    has_subject = any(r.get("subject_name") for r in rows)
    has_notes   = any(r.get("notes")        for r in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = label[:31]

    # Sarlavha
    ncols = 4 + len(sec_keys) + int(has_level) + int(has_subject) + int(has_notes)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=f"📊 {label} — Mock natijalari")
    tc.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    tc.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    # Statistika satri
    soni = len(rows)
    oquv = len({r["talaba_kod"] for r in rows})
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc = ws.cell(row=2, column=1,
                 value=f"Jami: {soni} ta natija  |  {oquv} ta o'quvchi  |  Eksport: {_now_str()}")
    sc.font      = _ITALIC_FONT
    sc.alignment = _LEFT
    ws.row_dimensions[2].height = 16

    # Header qatori
    HR  = 3
    col = 1
    col_map = {}

    _hcell(ws, HR, col, "№");               col_map["n"]      = col; col += 1
    _hcell(ws, HR, col, "O'quvchi");        col_map["ism"]    = col; col += 1
    _hcell(ws, HR, col, "Sinf");            col_map["sinf"]   = col; col += 1
    _hcell(ws, HR, col, "Sana");            col_map["sana"]   = col; col += 1

    for sk in sec_keys:
        _hcell(ws, HR, col, sec_labels.get(sk, sk), _SUBHEAD_FILL)
        col_map[f"sec_{sk}"] = col; col += 1

    _hcell(ws, HR, col, "Umumiy ball");     col_map["ball"]   = col; col += 1

    if has_level:
        _hcell(ws, HR, col, "Daraja", PatternFill("solid", fgColor="375623"))
        col_map["level"] = col; col += 1
    if has_subject:
        _hcell(ws, HR, col, "Fan nomi");    col_map["subj"]   = col; col += 1
    if has_notes:
        _hcell(ws, HR, col, "Izoh");        col_map["notes"]  = col; col += 1

    ws.row_dimensions[HR].height = 22

    # Ma'lumotlar
    for i, r in enumerate(rows, 1):
        dr  = HR + i
        alt = (i % 2 == 0)
        secs = r.get("sections") or {}

        _dcell(ws, dr, col_map["n"],    i,                                   alt)
        _dcell(ws, dr, col_map["ism"],  r.get("ismlar") or r["talaba_kod"],  alt, _LEFT)
        sinf_raw = r.get("sinf") or ""
        _dcell(ws, dr, col_map["sinf"], sinf_raw.split(" - ")[0] if sinf_raw else "—", alt)
        _dcell(ws, dr, col_map["sana"], _sana_str(r.get("test_sanasi")),     alt)

        for sk in sec_keys:
            v = _val(secs.get(sk))
            _dcell(ws, dr, col_map[f"sec_{sk}"],
                   round(v, 2) if isinstance(v, float) else v or "—", alt)

        ball = r.get("umumiy_ball")
        _dcell(ws, dr, col_map["ball"],
               round(ball, 2) if isinstance(ball, float) else (ball or "—"), alt)

        if has_level:
            lv = r.get("level_label") or "—"
            c  = ws.cell(row=dr, column=col_map["level"], value=lv)
            c.font      = _NORMAL_FONT
            c.fill      = _LEVEL_FILL if lv != "—" else (_ALT_FILL if alt else PatternFill())
            c.alignment = _CENTER
            c.border    = _thin()

        if has_subject:
            _dcell(ws, dr, col_map["subj"], r.get("subject_name") or "—", alt, _LEFT)
        if has_notes:
            _dcell(ws, dr, col_map["notes"], r.get("notes") or "—", alt, _LEFT)

    # Statistika bloki
    data_end = HR + soni
    stat_row = data_end + 2

    ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=ncols)
    ws.cell(row=stat_row, column=1, value="📊 Statistika").font = Font(
        name="Arial", bold=True, size=11, color="1F4E79")

    balls = [r["umumiy_ball"] for r in rows if r.get("umumiy_ball") is not None]
    if balls:
        stats = [
            ("Jami natijalar",  soni),
            ("O'quvchilar soni", oquv),
            ("O'rtacha ball",   f"{sum(balls)/len(balls):.2f}"),
            ("Eng yuqori",      max(balls)),
            ("Eng past",        min(balls)),
        ]
        for j, (lbl, val) in enumerate(stats, stat_row + 1):
            lc = ws.cell(row=j, column=1, value=lbl)
            lc.font = _BOLD_FONT; lc.fill = _TOTAL_FILL
            lc.alignment = _LEFT; lc.border = _thin()
            ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=ncols)
            vc = ws.cell(row=j, column=2, value=val)
            vc.font = _NORMAL_FONT; vc.alignment = _CENTER; vc.border = _thin()
        footer_row = stat_row + len(stats) + 2
    else:
        footer_row = stat_row + 2

    _footer(ws, footer_row, ncols, f"Hisobot sanasi: {_now_str()}")

    # Ustun kengliklari
    widths = {"n": 5, "ism": 28, "sinf": 10, "sana": 12, "ball": 13, "level": 10,
              "subj": 18, "notes": 24}
    for key, c in col_map.items():
        if key.startswith("sec_"):
            ws.column_dimensions[get_column_letter(c)].width = 13
        else:
            ws.column_dimensions[get_column_letter(c)].width = widths.get(key, 14)

    ws.freeze_panes = f"A{HR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_by_student(talaba_kod: str) -> bytes | None:
    """
    Bitta o'quvchining barcha mock natijalarini Excel ga eksport qiladi.
    Har bir fan — alohida varaq, birinchi varaqda umumiy ko'rinish.
    """
    from mock_database import mock_natijalari_ol, mock_natija_turlari, exam_type_ol
    from database import talaba_topish

    talaba = talaba_topish(talaba_kod)
    if not talaba:
        return None

    fanlar = mock_natija_turlari(talaba_kod)
    if not fanlar:
        return None

    wb = Workbook()

    # ── Umumiy varaq ───────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Umumiy"

    ncols_sum = 5
    ws_sum.merge_cells(f"A1:{get_column_letter(ncols_sum)}1")
    tc = ws_sum["A1"]
    tc.value     = f"Mock natijalari — {talaba['ismlar']} ({talaba_kod})"
    tc.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    tc.alignment = _CENTER
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells(f"A2:{get_column_letter(ncols_sum)}2")
    sc = ws_sum["A2"]
    sc.value     = f"Sinf: {talaba.get('sinf','—')}  |  Eksport: {_now_str()}"
    sc.font      = _ITALIC_FONT
    sc.alignment = _LEFT

    sum_headers = ["Fan", "Daraja", "Umumiy ball", "Soni", "Oxirgi sana"]
    for ci, h in enumerate(sum_headers, 1):
        _hcell(ws_sum, 3, ci, h)
    ws_sum.row_dimensions[3].height = 20

    for i, f in enumerate(fanlar, 1):
        dr  = 3 + i
        alt = (i % 2 == 0)
        # Oxirgi natijani topamiz
        last = mock_natijalari_ol(talaba_kod, exam_key=f["exam_key"], limit=1)
        level = last[0].get("level_label") or "—" if last else "—"
        ball  = last[0].get("umumiy_ball")  if last else None

        _dcell(ws_sum, dr, 1, f.get("exam_label") or f["exam_key"], alt, _LEFT)
        lc = ws_sum.cell(row=dr, column=2, value=level)
        lc.font = _NORMAL_FONT
        lc.fill = _LEVEL_FILL if level != "—" else (_ALT_FILL if alt else PatternFill())
        lc.alignment = _CENTER; lc.border = _thin()
        _dcell(ws_sum, dr, 3,
               round(ball, 2) if isinstance(ball, float) else (ball or "—"), alt)
        _dcell(ws_sum, dr, 4, f.get("soni", "—"), alt)
        _dcell(ws_sum, dr, 5, _sana_str(f.get("oxirgi_sana")), alt)

    _footer(ws_sum, 3 + len(fanlar) + 2, ncols_sum, f"Hisobot sanasi: {_now_str()}")
    for ci, w in enumerate([22, 12, 14, 8, 14], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # ── Har bir fan uchun alohida varaq ───────────────────────
    for f in fanlar:
        ek    = f["exam_key"]
        et    = exam_type_ol(ek) or {}
        lbl   = et.get("label", ek)
        rows  = mock_natijalari_ol(talaba_kod, exam_key=ek, limit=200)
        if not rows:
            continue

        safe  = lbl[:31].replace("/", "-")
        ws    = wb.create_sheet(title=safe)

        # Section ustunlar
        sec_keys   = []
        sec_labels = {}
        for r in rows:
            for k, v in (r.get("sections") or {}).items():
                if k not in sec_keys:
                    sec_keys.append(k)
                if k not in sec_labels:
                    sec_labels[k] = (v.get("label") if isinstance(v, dict) else None) or \
                                     k.replace("_", " ").title()

        has_level = any(r.get("level_label") for r in rows)
        has_notes = any(r.get("notes")       for r in rows)

        ncols = 3 + len(sec_keys) + 1 + int(has_level) + int(has_notes)

        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        tc = ws["A1"]
        tc.value     = f"{lbl} — {talaba['ismlar']}"
        tc.font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
        tc.alignment = _CENTER
        ws.row_dimensions[1].height = 24

        HR  = 2
        col = 1
        cmap = {}
        _hcell(ws, HR, col, "№");         cmap["n"]    = col; col += 1
        _hcell(ws, HR, col, "Sana");      cmap["sana"] = col; col += 1
        for sk in sec_keys:
            _hcell(ws, HR, col, sec_labels[sk], _SUBHEAD_FILL)
            cmap[f"s_{sk}"] = col; col += 1
        _hcell(ws, HR, col, "Umumiy ball"); cmap["ball"]  = col; col += 1
        if has_level:
            _hcell(ws, HR, col, "Daraja", PatternFill("solid", fgColor="375623"))
            cmap["level"] = col; col += 1
        if has_notes:
            _hcell(ws, HR, col, "Izoh"); cmap["notes"] = col; col += 1
        ws.row_dimensions[HR].height = 20

        for i, r in enumerate(rows, 1):
            dr  = HR + i
            alt = (i % 2 == 0)
            secs = r.get("sections") or {}
            _dcell(ws, dr, cmap["n"],    i,                           alt)
            _dcell(ws, dr, cmap["sana"], _sana_str(r.get("test_sanasi")), alt)
            for sk in sec_keys:
                v = _val(secs.get(sk))
                _dcell(ws, dr, cmap[f"s_{sk}"],
                       round(v, 2) if isinstance(v, float) else (v or "—"), alt)
            ball = r.get("umumiy_ball")
            _dcell(ws, dr, cmap["ball"],
                   round(ball, 2) if isinstance(ball, float) else (ball or "—"), alt)
            if has_level:
                lv = r.get("level_label") or "—"
                c  = ws.cell(row=dr, column=cmap["level"], value=lv)
                c.font = _NORMAL_FONT
                c.fill = _LEVEL_FILL if lv != "—" else (_ALT_FILL if alt else PatternFill())
                c.alignment = _CENTER; c.border = _thin()
            if has_notes:
                _dcell(ws, dr, cmap["notes"], r.get("notes") or "—", alt, _LEFT)

        _footer(ws, HR + len(rows) + 2, ncols, f"Hisobot sanasi: {_now_str()}")
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 12
        for sk in sec_keys:
            ws.column_dimensions[get_column_letter(cmap[f"s_{sk}"])].width = 13
        ws.column_dimensions[get_column_letter(cmap["ball"])].width = 13
        if has_level:
            ws.column_dimensions[get_column_letter(cmap["level"])].width = 10
        if has_notes:
            ws.column_dimensions[get_column_letter(cmap["notes"])].width = 24

        ws.freeze_panes = f"A{HR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_all() -> bytes | None:
    """
    Barcha fanlar bo'yicha barcha natijalar — bitta fayl, har fan alohida varaq.
    Birinchi varaqda umumiy statistika.
    """
    from mock_database import mock_exam_statistika, mock_hammasi_by_exam, exam_type_ol

    stats = mock_exam_statistika()
    if not stats:
        return None

    wb = Workbook()

    # ── Umumiy statistika varag'i ──────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Statistika"
    ncols_sum = 5

    ws_sum.merge_cells(f"A1:{get_column_letter(ncols_sum)}1")
    tc = ws_sum["A1"]
    tc.value     = "Mock natijalari — Umumiy statistika"
    tc.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = _CENTER
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells(f"A2:{get_column_letter(ncols_sum)}2")
    sc = ws_sum["A2"]
    sc.value     = f"Eksport: {_now_str()}"
    sc.font      = _ITALIC_FONT
    sc.alignment = _LEFT

    for ci, h in enumerate(["Fan", "Natijalar soni", "O'quvchilar", "Oxirgi sana", "Varaq"], 1):
        _hcell(ws_sum, 3, ci, h)
    ws_sum.row_dimensions[3].height = 20

    total_soni = total_oquv = 0
    for i, s in enumerate(stats, 1):
        dr  = 3 + i
        alt = (i % 2 == 0)
        lbl = s.get("exam_label") or s["exam_key"]
        _dcell(ws_sum, dr, 1, lbl,          alt, _LEFT)
        _dcell(ws_sum, dr, 2, s["soni"],    alt)
        _dcell(ws_sum, dr, 3, s["oquvchilar"], alt)
        _dcell(ws_sum, dr, 4, _sana_str(s.get("oxirgi_sana")), alt)
        _dcell(ws_sum, dr, 5, lbl[:31],     alt)   # varaq nomi
        total_soni += s["soni"]
        total_oquv += s["oquvchilar"]

    # Jami qator
    jami_row = 3 + len(stats) + 1
    for ci in range(1, ncols_sum + 1):
        c = ws_sum.cell(row=jami_row, column=ci)
        c.fill = _TOTAL_FILL; c.border = _thin()
    ws_sum.cell(row=jami_row, column=1, value="JAMI").font = _BOLD_FONT
    ws_sum.cell(row=jami_row, column=1).alignment = _CENTER
    ws_sum.cell(row=jami_row, column=2, value=total_soni).font = _BOLD_FONT
    ws_sum.cell(row=jami_row, column=2).alignment = _CENTER
    ws_sum.cell(row=jami_row, column=3, value=total_oquv).font = _BOLD_FONT
    ws_sum.cell(row=jami_row, column=3).alignment = _CENTER

    _footer(ws_sum, jami_row + 2, ncols_sum, f"Hisobot sanasi: {_now_str()}")
    for ci, w in enumerate([22, 16, 14, 14, 18], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # ── Har bir fan uchun varaq ────────────────────────────────
    for s in stats:
        ek   = s["exam_key"]
        et   = exam_type_ol(ek) or {}
        lbl  = et.get("label", ek)
        rows = mock_hammasi_by_exam(ek, limit=9999)
        if not rows:
            continue

        safe = lbl[:31].replace("/", "-")
        ws   = wb.create_sheet(title=safe)

        sec_keys   = []
        sec_labels = {}
        for r in rows:
            for k, v in (r.get("sections") or {}).items():
                if k not in sec_keys:
                    sec_keys.append(k)
                if k not in sec_labels:
                    sec_labels[k] = (v.get("label") if isinstance(v, dict) else None) or \
                                     k.replace("_", " ").title()

        has_level   = any(r.get("level_label")  for r in rows)
        has_subject = any(r.get("subject_name") for r in rows)
        has_notes   = any(r.get("notes")        for r in rows)

        ncols = 4 + len(sec_keys) + 1 + int(has_level) + int(has_subject) + int(has_notes)

        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        tc = ws["A1"]
        tc.value     = f"{lbl} — {len(rows)} ta natija"
        tc.font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
        tc.alignment = _CENTER
        ws.row_dimensions[1].height = 24

        HR  = 2
        col = 1
        cmap = {}
        _hcell(ws, HR, col, "№");           cmap["n"]    = col; col += 1
        _hcell(ws, HR, col, "O'quvchi");    cmap["ism"]  = col; col += 1
        _hcell(ws, HR, col, "Sinf");        cmap["sinf"] = col; col += 1
        _hcell(ws, HR, col, "Sana");        cmap["sana"] = col; col += 1
        for sk in sec_keys:
            _hcell(ws, HR, col, sec_labels[sk], _SUBHEAD_FILL)
            cmap[f"s_{sk}"] = col; col += 1
        _hcell(ws, HR, col, "Umumiy ball"); cmap["ball"]  = col; col += 1
        if has_level:
            _hcell(ws, HR, col, "Daraja", PatternFill("solid", fgColor="375623"))
            cmap["level"] = col; col += 1
        if has_subject:
            _hcell(ws, HR, col, "Fan nomi"); cmap["subj"] = col; col += 1
        if has_notes:
            _hcell(ws, HR, col, "Izoh");     cmap["notes"] = col; col += 1
        ws.row_dimensions[HR].height = 20

        for i, r in enumerate(rows, 1):
            dr  = HR + i
            alt = (i % 2 == 0)
            secs = r.get("sections") or {}
            _dcell(ws, dr, cmap["n"],    i,                                    alt)
            _dcell(ws, dr, cmap["ism"],  r.get("ismlar") or r["talaba_kod"],   alt, _LEFT)
            sinf_raw = r.get("sinf") or ""
            _dcell(ws, dr, cmap["sinf"], sinf_raw.split(" - ")[0] or "—",     alt)
            _dcell(ws, dr, cmap["sana"], _sana_str(r.get("test_sanasi")),      alt)
            for sk in sec_keys:
                v = _val(secs.get(sk))
                _dcell(ws, dr, cmap[f"s_{sk}"],
                       round(v, 2) if isinstance(v, float) else (v or "—"), alt)
            ball = r.get("umumiy_ball")
            _dcell(ws, dr, cmap["ball"],
                   round(ball, 2) if isinstance(ball, float) else (ball or "—"), alt)
            if has_level:
                lv = r.get("level_label") or "—"
                c  = ws.cell(row=dr, column=cmap["level"], value=lv)
                c.font = _NORMAL_FONT
                c.fill = _LEVEL_FILL if lv != "—" else (_ALT_FILL if alt else PatternFill())
                c.alignment = _CENTER; c.border = _thin()
            if has_subject:
                _dcell(ws, dr, cmap["subj"],  r.get("subject_name") or "—", alt, _LEFT)
            if has_notes:
                _dcell(ws, dr, cmap["notes"], r.get("notes") or "—",        alt, _LEFT)

        _footer(ws, HR + len(rows) + 2, ncols, f"Hisobot sanasi: {_now_str()}")
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions[get_column_letter(cmap["ism"])].width  = 28
        ws.column_dimensions[get_column_letter(cmap["sinf"])].width = 10
        ws.column_dimensions[get_column_letter(cmap["sana"])].width = 12
        ws.column_dimensions[get_column_letter(cmap["ball"])].width = 13
        for sk in sec_keys:
            ws.column_dimensions[get_column_letter(cmap[f"s_{sk}"])].width = 13
        if has_level and "level" in cmap:
            ws.column_dimensions[get_column_letter(cmap["level"])].width = 10
        if has_subject and "subj" in cmap:
            ws.column_dimensions[get_column_letter(cmap["subj"])].width = 18
        if has_notes and "notes" in cmap:
            ws.column_dimensions[get_column_letter(cmap["notes"])].width = 24
        ws.freeze_panes = f"A{HR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════
# TELEGRAM HANDLERLARI
# ════════════════════════════════════════════════════════════

class MockExportFSM(StatesGroup):
    talaba_kod_kutish = State()


async def _admin_ok(src) -> bool:
    from admin import is_admin_id
    uid = src.from_user.id if hasattr(src, "from_user") else src.message.from_user.id
    return is_admin_id(uid)


# ── Asosiy menyu ─────────────────────────────────────────────

@router.message(F.text == "📤 Mock eksport")
async def mock_eksport_menu(message: Message, state: FSMContext):
    if not await _admin_ok(message):
        return
    from mock_database import mock_exam_statistika
    stats = mock_exam_statistika()
    if not stats:
        await message.answer("📭 Bazada hech qanday mock natijasi yo'q.")
        return

    total = sum(s["soni"] for s in stats)
    buttons = [
        [InlineKeyboardButton(
            text=f"📊 Barcha natijalar ({total} ta)",
            callback_data="mock_exp_all"
        )],
        [InlineKeyboardButton(
            text="👤 O'quvchi bo'yicha",
            callback_data="mock_exp_student"
        )],
    ]
    for s in stats:
        lbl = s.get("exam_label") or s["exam_key"]
        buttons.append([InlineKeyboardButton(
            text=f"📋 {lbl} ({s['soni']} ta)",
            callback_data=f"mock_exp_fan:{s['exam_key']}"
        )])
    buttons.append([InlineKeyboardButton(text="❌ Yopish", callback_data="mock_exp_cancel")])

    await message.answer(
        "📤 <b>Mock natijalari — Excel eksport</b>\n\n"
        "Qaysi natijalarni eksport qilmoqchisiz?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


# ── Barcha natijalar ─────────────────────────────────────────

@router.callback_query(F.data == "mock_exp_all")
async def mock_exp_all(cb: CallbackQuery):
    if not await _admin_ok(cb):
        await cb.answer(); return

    await cb.message.edit_text("⏳ Barcha natijalar eksport qilinmoqda...")
    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, export_all)

    if not data:
        await cb.message.edit_text("📭 Eksport qilish uchun natijalar topilmadi.")
        await cb.answer(); return

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    name = f"mock_barcha_{ts}.xlsx"
    await cb.message.answer_document(
        BufferedInputFile(data, filename=name),
        caption="📊 <b>Barcha mock natijalari</b>\nHar bir fan — alohida varaq.",
        parse_mode="HTML",
    )
    await cb.message.delete()
    await cb.answer()


# ── Fan bo'yicha ─────────────────────────────────────────────

@router.callback_query(F.data.startswith("mock_exp_fan:"))
async def mock_exp_fan(cb: CallbackQuery):
    if not await _admin_ok(cb):
        await cb.answer(); return

    exam_key = cb.data.split(":", 1)[1]
    from mock_database import exam_type_ol
    et    = exam_type_ol(exam_key) or {}
    label = et.get("label", exam_key)

    await cb.message.edit_text(f"⏳ <b>{label}</b> natijalari eksport qilinmoqda...", parse_mode="HTML")
    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, export_by_exam, exam_key)

    if not data:
        await cb.message.edit_text(f"📭 <b>{label}</b> bo'yicha natijalar topilmadi.", parse_mode="HTML")
        await cb.answer(); return

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    name = f"mock_{exam_key.lower()}_{ts}.xlsx"
    await cb.message.answer_document(
        BufferedInputFile(data, filename=name),
        caption=f"📋 <b>{label}</b> — barcha natijalar",
        parse_mode="HTML",
    )
    await cb.message.delete()
    await cb.answer()


# ── O'quvchi bo'yicha ────────────────────────────────────────

@router.callback_query(F.data == "mock_exp_student")
async def mock_exp_student_start(cb: CallbackQuery, state: FSMContext):
    if not await _admin_ok(cb):
        await cb.answer(); return

    await state.set_state(MockExportFSM.talaba_kod_kutish)
    await cb.message.edit_text(
        "👤 <b>O'quvchi kodi</b>ni kiriting\n"
        "(masalan: <code>SS1001</code>):",
        parse_mode="HTML",
    )
    await cb.answer()


@router.message(MockExportFSM.talaba_kod_kutish)
async def mock_exp_student_kod(message: Message, state: FSMContext):
    if not await _admin_ok(message):
        return

    kod = message.text.strip().upper()
    wait = await message.answer(f"⏳ <b>{kod}</b> natijalari eksport qilinmoqda...", parse_mode="HTML")

    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, export_by_student, kod)
    await state.clear()

    if not data:
        await wait.edit_text(
            f"❌ <b>{kod}</b> — o'quvchi topilmadi yoki mock natijalari yo'q.",
            parse_mode="HTML",
        )
        return

    from database import talaba_topish
    talaba = talaba_topish(kod)
    ism    = talaba["ismlar"] if talaba else kod

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    name = f"mock_{kod.lower()}_{ts}.xlsx"
    await message.answer_document(
        BufferedInputFile(data, filename=name),
        caption=f"👤 <b>{ism}</b> ({kod}) — mock natijalari\nHar bir fan alohida varaqda.",
        parse_mode="HTML",
    )
    await wait.delete()


# ── Cancel ───────────────────────────────────────────────────

@router.callback_query(F.data == "mock_exp_cancel")
async def mock_exp_cancel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.delete()
    await cb.answer()
