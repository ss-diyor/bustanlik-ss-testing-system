"""
Excel Hisobot — PDF Hisobotga parallel modul.
Sinf yoki maktab bo'yicha o'quvchilar ro'yxati va ballarini .xlsx ga eksport qiladi.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from tz_utils import now as tz_now
from database import (
    talaba_topish,
    talaba_natijalari,
    get_all_students_for_excel,
    maktab_statistikasi,
    maktablar_ol,
    sinf_ol,
    get_all_in_class,
)

# ─── Stil konstantalar ───────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")
ALT_ROW_FILL  = PatternFill("solid", start_color="D6E4F0")
TOTAL_FILL    = PatternFill("solid", start_color="BDD7EE")

WHITE_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
ITALIC_FONT   = Font(name="Arial", italic=True, size=9, color="595959")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _header_row(ws, row_num, cols):
    """cols: list of (value, width_hint) tuples, fills row with header style."""
    for i, (val, _) in enumerate(cols, 1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = _thin_border()

def _data_cell(ws, row, col, value, alt=False, align=CENTER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = NORMAL_FONT
    cell.fill      = ALT_ROW_FILL if alt else PatternFill()
    cell.alignment = align
    cell.border    = _thin_border()
    return cell

def _footer_note(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = ITALIC_FONT
    cell.alignment = LEFT


# ─── ExcelExporter sinfi ─────────────────────────────────────────────────────

class ExcelExporter:
    def __init__(self):
        self.output_dir = "excel_reports"
        os.makedirs(self.output_dir, exist_ok=True)

    # ── 1. Bitta o'quvchi hisoboti ──────────────────────────────────────────
    def create_student_excel(self, talaba_kod: str) -> str | None:
        talaba = talaba_topish(talaba_kod)
        if not talaba:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "O'quvchi hisoboti"

        # Sarlavha
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value     = f"O'quvchi hisoboti — {talaba['ismlar']}"
        title_cell.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
        title_cell.alignment = CENTER
        ws.row_dimensions[1].height = 28

        # Ma'lumot qatorlari
        info = [
            ("Kod",        talaba["kod"]),
            ("Sinf",       talaba["sinf"]),
            ("Yo'nalish",  talaba.get("yonalish") or "—"),
        ]
        for offset, (label, value) in enumerate(info, 2):
            lc = ws.cell(row=offset, column=1, value=label)
            lc.font = BOLD_FONT; lc.fill = SUBHEAD_FILL
            lc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            lc.fill = SUBHEAD_FILL; lc.alignment = LEFT; lc.border = _thin_border()
            ws.merge_cells(start_row=offset, start_column=2,
                           end_row=offset,   end_column=5)
            vc = ws.cell(row=offset, column=2, value=value)
            vc.font = NORMAL_FONT; vc.alignment = LEFT; vc.border = _thin_border()

        # Natijalar jadvali
        cols = [
            ("Sana",         16),
            ("Majburiy",     12),
            ("Asosiy 1",     12),
            ("Asosiy 2",     12),
            ("Umumiy ball",  14),
        ]
        header_row = 6
        _header_row(ws, header_row, cols)
        ws.row_dimensions[header_row].height = 20

        natijalar = talaba_natijalari(talaba_kod)
        data_start = header_row + 1
        for i, n in enumerate(natijalar or []):
            r = data_start + i
            alt = (i % 2 == 1)
            _data_cell(ws, r, 1, n["test_sanasi"].strftime("%d.%m.%Y"), alt)
            _data_cell(ws, r, 2, n["majburiy"],   alt)
            _data_cell(ws, r, 3, n["asosiy_1"],   alt)
            _data_cell(ws, r, 4, n["asosiy_2"],   alt)
            _data_cell(ws, r, 5, n["umumiy_ball"], alt)

        if natijalar:
            # Statistika
            stat_row = data_start + len(natijalar) + 1
            ws.merge_cells(start_row=stat_row, start_column=1,
                           end_row=stat_row,   end_column=5)
            sh = ws.cell(row=stat_row, column=1, value="📊 Statistika")
            sh.font = Font(name="Arial", bold=True, size=11, color="1F4E79")
            sh.alignment = LEFT

            avg  = sum(n["umumiy_ball"] for n in natijalar) / len(natijalar)
            mxb  = max(n["umumiy_ball"] for n in natijalar)
            mnb  = min(n["umumiy_ball"] for n in natijalar)
            stats = [
                ("O'rtacha ball", f"{avg:.1f}"),
                ("Eng yuqori",    mxb),
                ("Eng past",      mnb),
                ("Jami testlar",  f"{len(natijalar)} ta"),
            ]
            for j, (lbl, val) in enumerate(stats, stat_row + 1):
                lc = ws.cell(row=j, column=1, value=lbl)
                lc.font = BOLD_FONT; lc.fill = TOTAL_FILL
                lc.alignment = LEFT; lc.border = _thin_border()
                ws.merge_cells(start_row=j, start_column=2,
                               end_row=j,   end_column=5)
                vc = ws.cell(row=j, column=2, value=val)
                vc.font = NORMAL_FONT; vc.alignment = CENTER; vc.border = _thin_border()

            footer_row = stat_row + len(stats) + 2
        else:
            ws.cell(row=data_start, column=1,
                    value="Natijalar mavjud emas").font = ITALIC_FONT
            footer_row = data_start + 2

        _footer_note(ws, footer_row, 5,
                     f"Hisobot sanasi: {tz_now().strftime('%d.%m.%Y %H:%M')}")

        # Ustun kengliklari
        widths = [16, 12, 12, 12, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        path = os.path.join(self.output_dir, f"student_{talaba_kod}.xlsx")
        wb.save(path)
        return path

    # ── 2. Maktab statistikasi ──────────────────────────────────────────────
    def create_maktab_statistika_excel(self, maktab_id: int = None) -> str | None:
        stats = maktab_statistikasi(maktab_id)
        if not stats:
            return None

        if maktab_id:
            maktablar = maktablar_ol()
            maktab = next((m for m in maktablar if m["id"] == maktab_id), None)
            title = f"{maktab['nomi'] if maktab else 'Maktab'} statistikasi"
        else:
            title = "Barcha maktablar statistikasi"

        wb = Workbook()
        ws = wb.active
        ws.title = "Maktab statistikasi"

        ncols = 6
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        tc = ws["A1"]
        tc.value     = title
        tc.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 30

        cols = [
            ("Maktab",       28),
            ("Sinf",         22),
            ("O'quvchilar",  13),
            ("O'rtacha",     13),
            ("Eng yuqori",   13),
            ("Eng past",     13),
        ]
        _header_row(ws, 2, cols)
        ws.row_dimensions[2].height = 22

        jami_oquv = 0
        for i, s in enumerate(stats, 3):
            alt = (i % 2 == 0)
            _data_cell(ws, i, 1, s["maktab_nomi"],           alt, LEFT)
            _data_cell(ws, i, 2, s["sinf"],                   alt, LEFT)
            _data_cell(ws, i, 3, s["oquvchilar_soni"],        alt)
            _data_cell(ws, i, 4, round(s["ortacha_ball"], 1), alt)
            _data_cell(ws, i, 5, round(s["eng_yuqori_ball"], 1), alt)
            _data_cell(ws, i, 6, round(s["eng_past_ball"], 1),   alt)
            jami_oquv += s["oquvchilar_soni"]

        # Umumiy jami
        total_row = len(stats) + 3
        ws.merge_cells(start_row=total_row, start_column=1,
                       end_row=total_row,   end_column=2)
        tc2 = ws.cell(row=total_row, column=1, value="JAMI O'QUVCHILAR")
        tc2.font = Font(name="Arial", bold=True, size=10)
        tc2.fill = TOTAL_FILL; tc2.alignment = CENTER; tc2.border = _thin_border()
        jc = ws.cell(row=total_row, column=3, value=jami_oquv)
        jc.font = BOLD_FONT; jc.fill = TOTAL_FILL
        jc.alignment = CENTER; jc.border = _thin_border()
        for col in range(4, ncols + 1):
            ec = ws.cell(row=total_row, column=col, value="")
            ec.fill = TOTAL_FILL; ec.border = _thin_border()

        _footer_note(ws, total_row + 2, ncols,
                     f"Hisobot sanasi: {tz_now().strftime('%d.%m.%Y %H:%M')}")

        for i, (_, w) in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ts = tz_now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self.output_dir, f"maktab_stat_{ts}.xlsx")
        wb.save(path)
        return path

    # ── 3. Sinf reytingi ────────────────────────────────────────────────────
    def create_sinf_reyting_excel(self, sinf: str = None,
                                   maktab_id: int = None) -> str | None:
        if sinf:
            sinflar = [sinf]
            title   = f"{sinf} sinf reytingi"
        elif maktab_id:
            from database import sinf_ol_batafsil
            batafsil    = sinf_ol_batafsil()
            maktablar_l = maktablar_ol()
            maktab      = next((m for m in maktablar_l if m["id"] == maktab_id), None)
            maktab_nomi = maktab["nomi"] if maktab else "Maktab"
            sinflar = [
                f"{s['nomi']} - {s['maktab_nomi']}"
                for s in batafsil
                if s["maktab_id"] == maktab_id
            ]
            title = f"{maktab_nomi} — barcha sinflar reytingi"
        else:
            sinflar = sinf_ol()
            title   = "Barcha sinflar reytingi"

        if not sinflar:
            return None

        wb  = Workbook()
        first = True
        all_ws = None  # Umumiy varaq — barcha sinflar

        # Umumiy varaq
        summary_ws = wb.active
        summary_ws.title = "Umumiy"
        summary_ncols = 5

        summary_ws.merge_cells(f"A1:{get_column_letter(summary_ncols)}1")
        sc = summary_ws["A1"]
        sc.value     = title
        sc.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
        sc.alignment = CENTER
        summary_ws.row_dimensions[1].height = 30

        sum_cols = [
            ("№",          6),
            ("Sinf",       20),
            ("Ism",        30),
            ("Yo'nalish",  28),
            ("Ball",       10),
        ]
        _header_row(summary_ws, 2, sum_cols)
        summary_ws.row_dimensions[2].height = 22
        for i, (_, w) in enumerate(sum_cols, 1):
            summary_ws.column_dimensions[get_column_letter(i)].width = w

        summary_row = 3
        global_rank = 0

        for sinf_name in sinflar:
            talabalar = get_all_in_class(sinf_name)
            if not talabalar:
                continue

            # Alohida varaq har bir sinf uchun
            safe_name = sinf_name[:31].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=safe_name)
            ncols_s = 5

            ws.merge_cells(f"A1:{get_column_letter(ncols_s)}1")
            tc = ws["A1"]
            tc.value     = f"{sinf_name} — sinf reytingi"
            tc.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
            tc.alignment = CENTER
            ws.row_dimensions[1].height = 26

            scols = [
                ("№",         6),
                ("Ism",       32),
                ("Kod",       14),
                ("Yo'nalish", 30),
                ("Ball",      10),
            ]
            _header_row(ws, 2, scols)
            ws.row_dimensions[2].height = 20
            for i, (_, w) in enumerate(scols, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            for j, t in enumerate(talabalar, 1):
                r   = j + 2
                alt = (j % 2 == 1)
                _data_cell(ws, r, 1, j,                            alt)
                _data_cell(ws, r, 2, t["ismlar"],                  alt, LEFT)
                _data_cell(ws, r, 3, t["kod"],                     alt)
                _data_cell(ws, r, 4, t.get("yonalish") or "—",    alt, LEFT)
                _data_cell(ws, r, 5, t["umumiy_ball"],             alt)

                # Umumiy varaqqaga ham qo'shamiz
                global_rank += 1
                sr  = summary_row
                alt_s = (global_rank % 2 == 1)
                _data_cell(summary_ws, sr, 1, global_rank,                  alt_s)
                _data_cell(summary_ws, sr, 2, sinf_name,                    alt_s, LEFT)
                _data_cell(summary_ws, sr, 3, t["ismlar"],                  alt_s, LEFT)
                _data_cell(summary_ws, sr, 4, t.get("yonalish") or "—",    alt_s, LEFT)
                _data_cell(summary_ws, sr, 5, t["umumiy_ball"],             alt_s)
                summary_row += 1

            _footer_note(ws, len(talabalar) + 4, ncols_s,
                         f"Hisobot sanasi: {tz_now().strftime('%d.%m.%Y %H:%M')}")

        _footer_note(summary_ws, summary_row + 1, summary_ncols,
                     f"Hisobot sanasi: {tz_now().strftime('%d.%m.%Y %H:%M')}")

        ts   = tz_now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self.output_dir, f"reyting_{ts}.xlsx")
        wb.save(path)
        return path
